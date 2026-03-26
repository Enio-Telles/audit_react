"""
Utilitários de I/O Parquet — audit_engine
Funções auxiliares para leitura, escrita e validação de arquivos Parquet.
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)


def ler_parquet(caminho: Path, colunas: Optional[List[str]] = None):
    """Lê um arquivo Parquet e retorna um DataFrame Polars."""
    try:
        import polars as pl
    except ImportError:
        raise RuntimeError("Polars não instalado. Execute: pip install polars")

    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    return pl.read_parquet(caminho, columns=colunas)


def escrever_parquet(df, caminho: Path, compressao: str = "zstd"):
    """Escreve um DataFrame Polars em formato Parquet."""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(caminho, compression=compressao)
    logger.info(f"Parquet salvo: {caminho} ({len(df)} registros)")


def info_parquet(caminho: Path) -> Dict:
    """Retorna informações sobre um arquivo Parquet."""
    try:
        import polars as pl
    except ImportError:
        raise RuntimeError("Polars não instalado")

    if not caminho.exists():
        return {"existe": False}

    df = pl.read_parquet(caminho)
    return {
        "existe": True,
        "registros": len(df),
        "colunas": df.columns,
        "schema": {col: str(df[col].dtype) for col in df.columns},
        "tamanho_bytes": caminho.stat().st_size,
    }


def listar_parquets(diretorio: Path) -> List[Dict]:
    """Lista todos os arquivos Parquet em um diretório."""
    if not diretorio.exists():
        return []

    resultado = []
    for arquivo in sorted(diretorio.glob("*.parquet")):
        try:
            info = info_parquet(arquivo)
            info["nome"] = arquivo.stem
            info["caminho"] = str(arquivo)
            resultado.append(info)
        except Exception as e:
            resultado.append({
                "nome": arquivo.stem,
                "caminho": str(arquivo),
                "existe": True,
                "erro": str(e),
            })

    return resultado


def exportar_excel(caminho_parquet: Path, caminho_excel: Path, formatado: bool = True):
    """Exporta um Parquet para Excel com formatação opcional."""
    try:
        import polars as pl
    except ImportError:
        raise RuntimeError("Polars não instalado")

    df = pl.read_parquet(caminho_parquet)
    df_pandas = df.to_pandas()

    if formatado:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            df_pandas.to_excel(caminho_excel, index=False, engine="openpyxl")
            
            wb = openpyxl.load_workbook(caminho_excel)
            ws = wb.active
            
            # Formatar header
            header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Auto-ajustar largura
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)
            
            wb.save(caminho_excel)
        except ImportError:
            df_pandas.to_excel(caminho_excel, index=False)
    else:
        df_pandas.to_excel(caminho_excel, index=False)

    logger.info(f"Excel exportado: {caminho_excel}")


def exportar_csv(caminho_parquet: Path, caminho_csv: Path, separador: str = ";"):
    """Exporta um Parquet para CSV."""
    try:
        import polars as pl
    except ImportError:
        raise RuntimeError("Polars não instalado")

    df = pl.read_parquet(caminho_parquet)
    df.write_csv(caminho_csv, separator=separador)
    logger.info(f"CSV exportado: {caminho_csv}")
