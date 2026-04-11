from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import base64
import inspect
import re
from typing import Callable

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font as OpenPyxlFont
from PySide6.QtCore import QDate, QThread, Qt, Signal, QUrl, QByteArray, QTimer
from PySide6.QtGui import QAction, QDesktopServices, QFont, QGuiApplication, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QMenu,
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QDateEdit,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QInputDialog,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QComboBox,
    QScrollArea,
    QSplitter,
    QStatusBar,
    QTabWidget,
    QTableView,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QToolBar,
    QVBoxLayout,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
    QStyledItemDelegate,
    QDoubleSpinBox,
)

from interface_grafica.config import (
    APP_NAME,
    CNPJ_ROOT,
    CONSULTAS_ROOT,
)
from interface_grafica.models.table_model import PolarsTableModel
from interface_grafica.services.aggregation_service import ServicoAgregacao
from interface_grafica.services.export_service import ExportService
from interface_grafica.services.parquet_service import FilterCondition, ParquetService
from interface_grafica.services.pipeline_funcoes_service import ResultadoPipeline, ServicoPipelineCompleto
from interface_grafica.services.pipeline_service import PipelineService
from interface_grafica.services.profile_utils import ordenar_colunas_perfil, ordenar_colunas_visiveis
from interface_grafica.services.query_worker import QueryWorker
from interface_grafica.services.registry_service import RegistryService
from interface_grafica.services.selection_persistence_service import SelectionPersistenceService
from interface_grafica.services.sql_service import SqlService, ParamInfo, WIDGET_DATE
from interface_grafica.ui.dialogs import (
    ColumnSelectorDialog,
    DialogoSelecaoConsultas,
    DialogoSelecaoTabelas,
)
from utilitarios.text import (
    display_cell,
    formatar_identificador_excel_texto,
    is_excel_date_column_name,
    is_excel_datetime_column_name,
    is_excel_text_identifier_column_name,
    is_year_column_name,
    normalize_text,
    parse_data_iso_texto,
    remove_accents,
)


class FloatDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QDoubleSpinBox(parent)
        editor.setDecimals(6)
        editor.setMinimum(-999999999.0)
        editor.setMaximum(999999999.0)
        return editor

class PipelineWorker(QThread):
    finished_ok = Signal(object)
    failed = Signal(str)
    progress = Signal(str)

    def __init__(
        self,
        service: ServicoPipelineCompleto,
        cnpj: str,
        consultas: list[str | Path],
        tabelas: list[str],
        data_limite: str | None = None,
    ) -> None:
        super().__init__()
        self.service = service
        self.cnpj = cnpj
        self.consultas = consultas
        self.tabelas = tabelas
        self.data_limite = data_limite

    def run(self) -> None:
        try:
            result = self.service.executar_completo(
                self.cnpj, 
                self.consultas, 
                self.tabelas, 
                self.data_limite,
                progresso=self.progress.emit
            )
        except Exception as exc:  # pragma: nao cover - UI
            from utilitarios.perf_monitor import registrar_evento_performance
            registrar_evento_performance(
                "pipeline_worker.erro",
                contexto={"cnpj": self.cnpj, "erro": str(exc)},
                status="error",
            )
            self.failed.emit("Ocorreu um erro no pipeline. Verifique os logs internos.")
            return
        
        if result.ok:
            self.finished_ok.emit(result)
        else:
            message = "\n".join(result.erros) if result.erros else "Falha nao pipeline."
            self.failed.emit(message or "Falha sem detalhes.")


class ServiceTaskWorker(QThread):
    finished_ok = Signal(object)
    failed = Signal(str)
    progress = Signal(str)

    def __init__(self, func: Callable, *args, **kwargs) -> None:
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs

    def run(self) -> None:
        try:
            call_kwargs = dict(self.kwargs)
            try:
                assinatura = inspect.signature(self.func)
                if "progresso" in assinatura.parameters and "progresso" not in call_kwargs:
                    call_kwargs["progresso"] = self.progress.emit
            except Exception:
                pass
            resultado = self.func(*self.args, **call_kwargs)
        except Exception as exc:
            from utilitarios.perf_monitor import registrar_evento_performance
            registrar_evento_performance(
                "service_task_worker.erro",
                contexto={"func": getattr(self.func, '__name__', str(self.func)), "erro": str(exc)},
                status="error",
            )
            self.failed.emit("Ocorreu um erro no processamento. Verifique os logs internos.")
            return
        self.finished_ok.emit(resultado)


class DetachedTableWindow(QMainWindow):
    closed = Signal(str)

    def __init__(
        self,
        titulo: str,
        contexto: str,
        table_model: PolarsTableModel,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._contexto = contexto
        self._source_model = table_model
        self._table_model = table_model.clone_configuration(pl.DataFrame())
        self.setAttribute(Qt.WA_DeleteOnClose, True)
        self.setWindowTitle(titulo)
        self.resize(1200, 720)

        central = QWidget(self)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(8, 8, 8, 8)

        self.lbl_titulo = QLabel(titulo)
        self.lbl_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; "
            "border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_titulo)

        filtros = QHBoxLayout()
        self.filter_column = QComboBox(self)
        self.filter_column.setMinimumWidth(180)
        self.filter_column.addItem("Todas")
        self.filter_text = QLineEdit(self)
        self.filter_text.setPlaceholderText("Filtrar na janela destacada...")
        self.profile_combo = QComboBox(self)
        self.profile_combo.setMinimumWidth(140)
        self.btn_apply_profile = QPushButton("Perfil", self)
        self.btn_save_profile = QPushButton("Salvar perfil", self)
        self.btn_columns = QPushButton("Colunas", self)
        self.btn_apply_filter = QPushButton("Aplicar filtros", self)
        self.btn_clear_filter = QPushButton("Limpar filtros", self)
        filtros.addWidget(self.filter_column)
        filtros.addWidget(self.filter_text, 1)
        filtros.addWidget(self.profile_combo)
        filtros.addWidget(self.btn_apply_profile)
        filtros.addWidget(self.btn_save_profile)
        filtros.addWidget(self.btn_columns)
        filtros.addWidget(self.btn_apply_filter)
        filtros.addWidget(self.btn_clear_filter)
        layout.addLayout(filtros)

        self.lbl_status = QLabel("Filtros ativos: nenhum", self)
        self.lbl_status.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #dbeafe; background: #0f1b33; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_status)

        self.table = QTableView(self)
        self.table.setModel(self._table_model)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.setWordWrap(True)
        self.table.verticalHeader().setDefaultSectionSize(40)
        self.table.horizontalHeader().setMinimumSectionSize(40)
        self.table.horizontalHeader().setDefaultSectionSize(150)
        self.table.horizontalHeader().setMaximumSectionSize(420)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.table, 1)

        self.setCentralWidget(central)

        self.btn_apply_filter.clicked.connect(self.apply_filters)
        self.btn_clear_filter.clicked.connect(self.clear_filters)
        self.filter_text.returnPressed.connect(self.apply_filters)
        self.filter_column.currentIndexChanged.connect(lambda _i: self.apply_filters())

        self._on_source_model_reset = self._refresh_from_source
        self._on_source_layout_changed = self._refresh_from_source
        self._on_source_data_changed = lambda *_args: self._refresh_from_source()
        self._source_model.modelReset.connect(self._on_source_model_reset)
        self._source_model.layoutChanged.connect(self._on_source_layout_changed)
        self._source_model.dataChanged.connect(self._on_source_data_changed)

        self._refresh_from_source()

    @property
    def contexto(self) -> str:
        return self._contexto

    @property
    def table_model(self) -> PolarsTableModel:
        return self._table_model

    def _expr_texto_coluna(self, df: pl.DataFrame, coluna: str) -> pl.Expr:
        dtype = df.schema.get(coluna)
        if dtype is not None and dtype.base_type() == pl.List:
            return (
                pl.col(coluna)
                .cast(pl.List(pl.Utf8), strict=False)
                .list.join(" | ")
                .fill_null("")
                .str.to_lowercase()
            )
        return (
            pl.col(coluna)
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.to_lowercase()
        )

    def _refresh_from_source(self) -> None:
        if not hasattr(self, "filter_column") or self.filter_column is None:
            return
        df = self._source_model.get_dataframe()
        col_atual = self.filter_column.currentText()
        self.filter_column.blockSignals(True)
        self.filter_column.clear()
        self.filter_column.addItem("Todas")
        self.filter_column.addItems(df.columns)
        idx = self.filter_column.findText(col_atual)
        if idx >= 0:
            self.filter_column.setCurrentIndex(idx)
        self.filter_column.blockSignals(False)
        self.apply_filters()

    def apply_filters(self) -> None:
        df = self._source_model.get_dataframe()
        termo = self.filter_text.text().strip().lower()
        coluna = self.filter_column.currentText().strip()

        if not df.is_empty() and termo:
            if coluna and coluna != "Todas" and coluna in df.columns:
                df = df.filter(
                    self._expr_texto_coluna(df, coluna).str.contains(termo, literal=True)
                )
            else:
                exprs = [
                    self._expr_texto_coluna(df, col).str.contains(termo, literal=True)
                    for col in df.columns
                ]
                if exprs:
                    df = df.filter(pl.any_horizontal(exprs))

        self._table_model.set_dataframe(df)
        self.lbl_status.setText(
            "Filtros ativos: nenhum" if not termo else f"Filtros ativos: {coluna or 'Todas'} contem '{termo}'"
        )
        self.table.resizeColumnsToContents()

    def clear_filters(self) -> None:
        self.filter_text.clear()
        self.filter_column.setCurrentIndex(0)
        self.apply_filters()

    def closeEvent(self, event) -> None:
        try:
            self._source_model.modelReset.disconnect(self._on_source_model_reset)
        except Exception:
            pass
        try:
            self._source_model.layoutChanged.disconnect(self._on_source_layout_changed)
        except Exception:
            pass
        try:
            self._source_model.dataChanged.disconnect(self._on_source_data_changed)
        except Exception:
            pass
        self.closed.emit(self._contexto)
        super().closeEvent(event)


@dataclass
class ViewState:
    current_cnpj: str | None = None
    current_file: Path | None = None
    all_columns: list[str] | None = None
    visible_columns: list[str] | None = None
    filters: list[FilterCondition] | None = None
    total_rows: int = 0



@dataclass
class OracleConnectionContext:
    f_host: QLineEdit
    f_port: QLineEdit
    f_service: QLineEdit
    f_user: QLineEdit
    f_password: QLineEdit
    lbl: QLabel
    worker_attr: str
    btn: QPushButton | None = None


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.resize(1560, 920)

        self.registry_service = RegistryService()
        self.selection_service = SelectionPersistenceService()
        self.parquet_service = ParquetService(root=CNPJ_ROOT)
        self.pipeline_service = PipelineService(output_root=CONSULTAS_ROOT)
        self.servico_pipeline_funcoes = ServicoPipelineCompleto()
        self.export_service = ExportService()
        self.servico_agregacao = ServicoAgregacao()
        self.sql_service = SqlService()

        self.state = ViewState(filters=[])
        self.current_page_df_all = pl.DataFrame()
        self.current_page_df_visible = pl.DataFrame()
        self.table_model = PolarsTableModel()
        self.aggregation_table_model = PolarsTableModel(checkable=True)
        self.results_table_model = PolarsTableModel(checkable=True)
        self.conversion_model = PolarsTableModel()
        self.conversion_model.set_editable_columns({"fator", "unid_ref"})
        self.mov_estoque_model = PolarsTableModel(
            foreground_resolver=self._mov_estoque_foreground,
            background_resolver=self._mov_estoque_background,
            font_resolver=self._mov_estoque_font,
        )
        self.sql_result_model = PolarsTableModel()
        self.aggregation_basket: list[dict] = []
        self.aggregation_results: list[dict] = []
        self.pipeline_worker: PipelineWorker | None = None
        self.query_worker: QueryWorker | None = None
        self.service_worker: ServiceTaskWorker | None = None
        self._sql_files: list = []
        self._sql_param_widgets: dict[str, QWidget] = {}
        self._sql_current_sql: str = ""
        self._sql_result_df: pl.DataFrame = pl.DataFrame()
        self._conversion_df_full: pl.DataFrame = pl.DataFrame()
        self._recalculando_conversao = False
        self._conversion_recalc_pending = False
        self._mov_estoque_file_path: Path | None = None
        self._mov_estoque_df: pl.DataFrame = pl.DataFrame()
        self.aba_anual_model = PolarsTableModel(
            checkable=True,
            foreground_resolver=self._aba_anual_foreground,
            background_resolver=self._aba_anual_background,
        )
        self._aba_anual_file_path: Path | None = None
        self._aba_anual_df: pl.DataFrame = pl.DataFrame()
        self.aba_mensal_model = PolarsTableModel(
            checkable=True,
            foreground_resolver=self._aba_mensal_foreground,
            background_resolver=self._aba_mensal_background,
        )
        self._aba_mensal_file_path: Path | None = None
        self._aba_mensal_df: pl.DataFrame = pl.DataFrame()
        self.nfe_entrada_model = PolarsTableModel()
        self._nfe_entrada_file_path: Path | None = None
        self._nfe_entrada_df: pl.DataFrame = pl.DataFrame()
        self.id_agrupados_model = PolarsTableModel()
        self._id_agrupados_file_path: Path | None = None
        self._id_agrupados_df: pl.DataFrame = pl.DataFrame()
        self.produtos_selecionados_model = PolarsTableModel(checkable=True)
        self._produtos_selecionados_df: pl.DataFrame = pl.DataFrame()
        self._produtos_selecionados_mov_df: pl.DataFrame = pl.DataFrame()
        self._produtos_selecionados_mensal_df: pl.DataFrame = pl.DataFrame()
        self._produtos_selecionados_anual_df: pl.DataFrame = pl.DataFrame()
        self.resumo_global_model = PolarsTableModel()
        self._resumo_global_df: pl.DataFrame = pl.DataFrame()
        self._produtos_sel_preselecionado_cnpj: str | None = None
        self._filtro_cruzado_anuais_ids: list[str] = []
        self._aggregation_file_path: Path | None = None
        self._aggregation_filters: list[FilterCondition] = []
        self._aggregation_results_filters: list[FilterCondition] = []
        self._aggregation_relational_mode: str | None = None
        self._aggregation_results_relational_mode: str | None = None
        self._sync_resumos_estoque_cnpj: str | None = None
        self._debounce_timers: dict[str, QTimer] = {}
        self._debounce_callbacks: dict[str, Callable[[], None]] = {}
        self._auto_resized_tables: set[str] = set()
        self._detached_windows: dict[str, DetachedTableWindow] = {}
        self._closing_after_workers = False

        self._build_ui()
        self._connect_signals()
        self._setup_copy_shortcut()
        self._refresh_profile_combos()
        self.refresh_cnpjs()
        self.refresh_logs()
        self._populate_sql_combo()
        # verifica conexão Oracle automaticamente na abertura da aplicação
        QTimer.singleShot(800, self._verificar_conexoes)

    def _executar_callback_debounce(self, key: str) -> None:
        callback = self._debounce_callbacks.get(key)
        if callback is None:
            return
        callback()

    def _schedule_debounced(self, key: str, callback: Callable[[], None], delay_ms: int = 280) -> None:
        timer = self._debounce_timers.get(key)
        if timer is None:
            timer = QTimer(self)
            timer.setSingleShot(True)
            timer.timeout.connect(lambda key=key: self._executar_callback_debounce(key))
            self._debounce_timers[key] = timer
        self._debounce_callbacks[key] = callback
        timer.start(delay_ms)

    def _registrar_limpeza_worker(self, attr_name: str, worker: QThread) -> None:
        def _cleanup() -> None:
            if getattr(self, attr_name, None) is worker:
                setattr(self, attr_name, None)
            worker.deleteLater()
            self._atualizar_estado_botao_nfe_entrada()
            if self._closing_after_workers:
                self._tentar_fechar_apos_workers()

        worker.finished.connect(_cleanup)

    def _workers_ativos(self) -> list[QThread]:
        ativos: list[QThread] = []
        for worker in (self.pipeline_worker, self.query_worker, self.service_worker):
            if worker is not None and worker.isRunning():
                ativos.append(worker)
        return ativos

    def _atualizar_estado_botao_nfe_entrada(self) -> None:
        if not hasattr(self, "btn_extract_nfe_entrada"):
            return
        habilitado = bool(self.state.current_cnpj) and not self._workers_ativos()
        self.btn_extract_nfe_entrada.setEnabled(habilitado)

    def _tentar_fechar_apos_workers(self) -> None:
        if self._workers_ativos():
            return
        self._closing_after_workers = False
        self.close()

    def closeEvent(self, event) -> None:
        ativos = self._workers_ativos()
        if not ativos:
            super().closeEvent(event)
            return

        if not self._closing_after_workers:
            self._closing_after_workers = True
            self.status.showMessage("Aguardando o termino das operacoes em execucao para fechar a janela...")
            self.setEnabled(False)
            for worker in ativos:
                worker.finished.connect(self._tentar_fechar_apos_workers)
            QTimer.singleShot(100, self._tentar_fechar_apos_workers)
        event.ignore()

    def _resize_table_once(self, table: QTableView, key: str) -> None:
        if key in self._auto_resized_tables:
            return
        table.resizeColumnsToContents()
        self._auto_resized_tables.add(key)

    def _reset_table_resize_flag(self, key: str) -> None:
        self._auto_resized_tables.discard(key)

    def _estilo_botao_destacar(self) -> str:
        return (
            "QPushButton { background: #0e639c; color: #ffffff; border: 1px solid #1177bb; "
            "border-radius: 4px; padding: 6px 10px; font-weight: bold; }"
            "QPushButton:hover { background: #1177bb; }"
            "QPushButton:pressed { background: #0b4f7c; }"
        )

    def _criar_botao_destacar(self, texto: str = "Destacar") -> QPushButton:
        botao = QPushButton(texto)
        botao.setStyleSheet(self._estilo_botao_destacar())
        return botao

    def _build_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)

        self.main_splitter = QSplitter(Qt.Horizontal)
        root_layout.addWidget(self.main_splitter)

        self.left_panel_widget = self._build_left_panel()
        self.main_splitter.addWidget(self.left_panel_widget)
        self.main_splitter.addWidget(self._build_right_panel())
        self.main_splitter.setSizes([310, 1200])

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("Pronto.")

    def _build_left_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)

        cnpj_box = QGroupBox("CPF/CNPJ")
        cnpj_layout = QVBoxLayout(cnpj_box)
        input_line = QHBoxLayout()
        self.cnpj_input = QLineEdit()
        self.cnpj_input.setPlaceholderText("Digite o CPF ou CNPJ com ou sem mascara")
        self.btn_run_pipeline = QPushButton("Extrair + Processar")
        input_line.addWidget(self.cnpj_input)
        input_line.addWidget(self.btn_run_pipeline)
        cnpj_layout.addLayout(input_line)

        date_line = QHBoxLayout()
        date_line.addWidget(QLabel("Data limite EFD:"))
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setDisplayFormat("dd/MM/yyyy")
        date_line.addWidget(self.date_input)
        cnpj_layout.addLayout(date_line)

        actions_row1 = QHBoxLayout()
        self.btn_extrair_brutas = QPushButton("Extrair Tabelas Brutas")
        self.btn_processamento = QPushButton("Processamento")
        actions_row1.addWidget(self.btn_extrair_brutas)
        actions_row1.addWidget(self.btn_processamento)
        cnpj_layout.addLayout(actions_row1)

        actions_row2 = QHBoxLayout()
        self.btn_refresh_cnpjs = QPushButton("Atualizar lista")
        self.btn_open_cnpj_folder = QPushButton("Abrir pasta")
        actions_row2.addWidget(self.btn_refresh_cnpjs)
        actions_row2.addWidget(self.btn_open_cnpj_folder)
        cnpj_layout.addLayout(actions_row2)

        actions_row3 = QHBoxLayout()
        self.btn_apagar_dados = QPushButton("Apagar Dados do CNPJ")
        self.btn_apagar_dados.setStyleSheet("QPushButton { color: #e57373; }")
        self.btn_apagar_cnpj = QPushButton("Apagar CNPJ")
        self.btn_apagar_cnpj.setStyleSheet("QPushButton { color: #ef5350; font-weight: bold; }")
        actions_row3.addWidget(self.btn_apagar_dados)
        actions_row3.addWidget(self.btn_apagar_cnpj)
        cnpj_layout.addLayout(actions_row3)

        self.cnpj_list = QListWidget()
        cnpj_layout.addWidget(self.cnpj_list)
        layout.addWidget(cnpj_box)

        files_box = QGroupBox("Arquivos Parquet do CNPJ")
        files_layout = QVBoxLayout(files_box)
        self.file_tree = QTreeWidget()
        self.file_tree.setHeaderLabels(["Arquivo", "Local"])
        files_layout.addWidget(self.file_tree)
        layout.addWidget(files_box)

        notes = QLabel(
            "Fluxo recomendado: analise um CNPJ, abra a tabela desejada, filtre, selecione colunas e exporte. "
            "Para agregacao, trabalhe sobre a tabela desagregada e monte o lote na aba Agregacao."
        )
        notes.setWordWrap(True)
        layout.addWidget(notes)
        return panel

    def _build_right_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)

        header = QHBoxLayout()
        self.lbl_context = QLabel("Nenhum arquivo selecionado")
        self.lbl_context.setWordWrap(True)
        header.addWidget(self.lbl_context)
        header.addStretch()
        layout.addLayout(header)

        self.tabs = QTabWidget()
        self.btn_toggle_panel = QPushButton("<< Ocultar Painel Lateral")
        self.btn_toggle_panel.setCheckable(True)
        self.tabs.setCornerWidget(self.btn_toggle_panel, Qt.TopRightCorner)

        self.tabs.addTab(self._build_tab_configuracoes(), "Configurações")
        self.tabs.addTab(self._build_tab_consulta(), "Consulta")
        self.tabs.addTab(self._build_tab_sql_query(), "Consulta SQL")
        self.tabs.addTab(self._build_tab_agregacao(), "Agregacao")
        self.tab_conversao = self._build_tab_conversao()
        self.tabs.addTab(self.tab_conversao, "Conversao")
        self.tabs.addTab(self._build_tab_estoque(), "Estoque")
        self.tab_nfe_entrada = self._build_tab_nfe_entrada()
        self.tabs.addTab(self.tab_nfe_entrada, "NFe Entrada")
        self.tabs.addTab(self._build_tab_analise_lote_cnpj(), "Análise Lote CNPJ")
        self.tabs.addTab(self._build_tab_logs(), "Logs")
        layout.addWidget(self.tabs)
        return panel

    # ------------------------------------------------------------------
    # Aba: Configurações Oracle / Aplicativo
    # ------------------------------------------------------------------

    def _build_tab_configuracoes(self) -> QWidget:
        """Aba de configuração de conexões Oracle e opções gerais da aplicação."""
        from dotenv import dotenv_values
        from interface_grafica.fisconforme.path_resolver import get_env_path

        env_path = get_env_path()
        env_vars = dotenv_values(env_path) if env_path.exists() else {}

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        scroll.setWidget(container)
        root = QVBoxLayout(container)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        # ── helpers ──────────────────────────────────────────────────
        def _field(key: str, placeholder: str = "", password: bool = False) -> QLineEdit:
            le = QLineEdit()
            le.setText(str(env_vars.get(key, "")))
            if placeholder:
                le.setPlaceholderText(placeholder)
            if password:
                le.setEchoMode(QLineEdit.Password)
            return le

        def _status_label() -> QLabel:
            lbl = QLabel("—")
            lbl.setWordWrap(True)
            lbl.setMinimumHeight(36)
            return lbl

        def _test_button(texto: str = "Testar Conexão") -> QPushButton:
            btn = QPushButton(texto)
            btn.setFixedWidth(160)
            return btn

        # ── Status de Conexão — painel de destaque no topo ────────────
        grp_status = QGroupBox("Status da Conexão Oracle")
        sl = QVBoxLayout(grp_status)
        sl.setSpacing(6)
        sl.setContentsMargins(12, 8, 12, 8)

        r1 = QHBoxLayout()
        lbl_c1_title = QLabel("Conexão 1 — Principal:")
        lbl_c1_title.setFixedWidth(190)
        self._cfg_conn_lbl_1 = QLabel("— não verificado")
        self._cfg_conn_lbl_1.setWordWrap(True)
        r1.addWidget(lbl_c1_title)
        r1.addWidget(self._cfg_conn_lbl_1)
        r1.addStretch()
        sl.addLayout(r1)

        r2 = QHBoxLayout()
        lbl_c2_title = QLabel("Conexão 2 — Secundária:")
        lbl_c2_title.setFixedWidth(190)
        self._cfg_conn_lbl_2 = QLabel("— não verificado")
        self._cfg_conn_lbl_2.setWordWrap(True)
        r2.addWidget(lbl_c2_title)
        r2.addWidget(self._cfg_conn_lbl_2)
        r2.addStretch()
        sl.addLayout(r2)

        btn_verify_all = QPushButton("↺  Verificar Conexões")
        btn_verify_all.setFixedWidth(180)
        btn_verify_all.clicked.connect(self._verificar_conexoes)
        sl.addWidget(btn_verify_all)
        root.addWidget(grp_status)

        # ── Conexão Oracle 1 (Principal) ─────────────────────────────
        grp1 = QGroupBox("Conexão Oracle 1 — Principal")
        form1 = QFormLayout(grp1)
        form1.setLabelAlignment(Qt.AlignRight)
        form1.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        self._cfg_host = _field("ORACLE_HOST", "ex: exa01-scan.sefin.ro.gov.br")
        self._cfg_port = _field("ORACLE_PORT", "1521")
        self._cfg_service = _field("ORACLE_SERVICE", "ex: sefindw")
        self._cfg_user = _field("DB_USER", "CPF ou usuário")
        self._cfg_password = _field("DB_PASSWORD", "Senha", password=True)
        form1.addRow("Host:", self._cfg_host)
        form1.addRow("Porta:", self._cfg_port)
        form1.addRow("Serviço:", self._cfg_service)
        form1.addRow("Usuário:", self._cfg_user)
        form1.addRow("Senha:", self._cfg_password)

        self._cfg_test_status_1 = _status_label()
        self._cfg_btn_test_1 = _test_button()
        self._cfg_btn_test_1.clicked.connect(lambda: self._testar_conexao(OracleConnectionContext(
            f_host=self._cfg_host, f_port=self._cfg_port, f_service=self._cfg_service,
            f_user=self._cfg_user, f_password=self._cfg_password,
            btn=self._cfg_btn_test_1, lbl=self._cfg_test_status_1,
            worker_attr="_oracle_test_worker_1",
        )))
        test_row1 = QHBoxLayout()
        test_row1.addWidget(self._cfg_btn_test_1)
        test_row1.addWidget(self._cfg_test_status_1)
        test_row1.addStretch()
        form1.addRow("Teste:", test_row1)
        root.addWidget(grp1)

        # ── Conexão Oracle 2 (Secundária) ────────────────────────────
        grp2 = QGroupBox("Conexão Oracle 2 — Secundária")
        form2 = QFormLayout(grp2)
        form2.setLabelAlignment(Qt.AlignRight)
        form2.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        self._cfg_host_1 = _field("ORACLE_HOST_1", "ex: exacc-x10-sefinscan.sefin.ro.gov.br")
        self._cfg_port_1 = _field("ORACLE_PORT_1", "1521")
        self._cfg_service_1 = _field("ORACLE_SERVICE_1", "ex: svc.bi.users")
        self._cfg_user_1 = _field("DB_USER_1", "CPF ou usuário")
        self._cfg_password_1 = _field("DB_PASSWORD_1", "Senha", password=True)
        form2.addRow("Host:", self._cfg_host_1)
        form2.addRow("Porta:", self._cfg_port_1)
        form2.addRow("Serviço:", self._cfg_service_1)
        form2.addRow("Usuário:", self._cfg_user_1)
        form2.addRow("Senha:", self._cfg_password_1)

        self._cfg_test_status_2 = _status_label()
        self._cfg_btn_test_2 = _test_button()
        self._cfg_btn_test_2.clicked.connect(lambda: self._testar_conexao(OracleConnectionContext(
            f_host=self._cfg_host_1, f_port=self._cfg_port_1, f_service=self._cfg_service_1,
            f_user=self._cfg_user_1, f_password=self._cfg_password_1,
            btn=self._cfg_btn_test_2, lbl=self._cfg_test_status_2,
            worker_attr="_oracle_test_worker_2",
        )))
        test_row2 = QHBoxLayout()
        test_row2.addWidget(self._cfg_btn_test_2)
        test_row2.addWidget(self._cfg_test_status_2)
        test_row2.addStretch()
        form2.addRow("Teste:", test_row2)
        root.addWidget(grp2)

        # ── Configurações do Aplicativo ───────────────────────────────
        grp3 = QGroupBox("Configurações do Aplicativo")
        form3 = QFormLayout(grp3)
        form3.setLabelAlignment(Qt.AlignRight)
        form3.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self._cfg_log_level = QComboBox()
        self._cfg_log_level.addItems(["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
        current_level = env_vars.get("LOG_LEVEL", "INFO").upper()
        idx = self._cfg_log_level.findText(current_level)
        if idx >= 0:
            self._cfg_log_level.setCurrentIndex(idx)

        self._cfg_cache_enabled = QCheckBox("Ativar cache")
        self._cfg_cache_enabled.setChecked(env_vars.get("CACHE_ENABLED", "true").lower() == "true")

        self._cfg_cache_ttl = _field("CACHE_TTL", "3600 (segundos)")

        self._cfg_theme = QComboBox()
        self._cfg_theme.addItems(["dark", "light"])
        current_theme = env_vars.get("DASHBOARD_THEME", "dark").lower()
        theme_idx = self._cfg_theme.findText(current_theme)
        if theme_idx >= 0:
            self._cfg_theme.setCurrentIndex(theme_idx)

        form3.addRow("Nível de log:", self._cfg_log_level)
        form3.addRow("Cache:", self._cfg_cache_enabled)
        form3.addRow("TTL do cache (s):", self._cfg_cache_ttl)
        form3.addRow("Tema do dashboard:", self._cfg_theme)
        root.addWidget(grp3)

        # ── Botão salvar ──────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self._cfg_status_label = QLabel("")
        btn_salvar = QPushButton("Salvar Configurações")
        btn_salvar.setStyleSheet(self._estilo_botao_destacar())
        btn_salvar.clicked.connect(self._salvar_configuracoes)
        btn_row.addStretch()
        btn_row.addWidget(self._cfg_status_label)
        btn_row.addWidget(btn_salvar)
        root.addLayout(btn_row)
        root.addStretch()

        # init worker slots
        self._oracle_test_worker_1: object | None = None
        self._oracle_test_worker_2: object | None = None
        self._oracle_verify_worker_1: object | None = None
        self._oracle_verify_worker_2: object | None = None

        return scroll

    def _verificar_conexoes(self) -> None:
        """Testa ambas as conexões Oracle e atualiza o painel de status no topo da aba."""
        if not hasattr(self, "_cfg_host"):
            return  # aba ainda não construída
        self._testar_conexao_para_status(OracleConnectionContext(
            f_host=self._cfg_host, f_port=self._cfg_port, f_service=self._cfg_service,
            f_user=self._cfg_user, f_password=self._cfg_password,
            lbl=self._cfg_conn_lbl_1, worker_attr="_oracle_verify_worker_1",
        ))
        self._testar_conexao_para_status(OracleConnectionContext(
            f_host=self._cfg_host_1, f_port=self._cfg_port_1, f_service=self._cfg_service_1,
            f_user=self._cfg_user_1, f_password=self._cfg_password_1,
            lbl=self._cfg_conn_lbl_2, worker_attr="_oracle_verify_worker_2",
        ))

    def _testar_conexao_para_status(self, ctx: OracleConnectionContext) -> None:
        """Worker isolado que atualiza apenas o label de status (sem botão dedicado)."""
        from interface_grafica.services.oracle_test_worker import OracleConnectionTestWorker

        existing = getattr(self, ctx.worker_attr, None)
        if existing is not None and existing.isRunning():
            return

        ctx.lbl.setText("⏳ verificando…")
        ctx.lbl.setStyleSheet("color: #ccaa00;")

        worker = OracleConnectionTestWorker(
            host=ctx.f_host.text(), port=ctx.f_port.text(),
            service=ctx.f_service.text(), user=ctx.f_user.text(),
            password=ctx.f_password.text(), parent=self,
        )
        setattr(self, ctx.worker_attr, worker)

        def _on(ok: bool, msg: str, ms: int) -> None:
            first_line = msg.splitlines()[0] if msg else ""
            if ok:
                ctx.lbl.setText(f"✔ {first_line}")
                ctx.lbl.setStyleSheet("color: #4caf50; font-weight: bold;")
                self.status.showMessage(
                    f"[Oracle] {ctx.lbl.parent().parent().parent().title() if False else ctx.worker_attr.replace('_oracle_verify_worker_','Conexão ')} — OK ({ms} ms)",
                    5000,
                )
            else:
                short = first_line[:100]
                ctx.lbl.setText(f"✖ {short}")
                ctx.lbl.setStyleSheet("color: #e57373;")
            worker.deleteLater()
            setattr(self, ctx.worker_attr, None)

        worker.resultado.connect(_on)
        worker.start()

    def _testar_conexao(self, ctx: OracleConnectionContext) -> None:
        """Lança o teste de conexão Oracle em background (não bloqueia a UI)."""
        from interface_grafica.services.oracle_test_worker import OracleConnectionTestWorker

        # evitar múltiplos testes simultâneos no mesmo slot
        existing: OracleConnectionTestWorker | None = getattr(self, ctx.worker_attr, None)
        if existing is not None and existing.isRunning():
            return

        if ctx.btn:
            ctx.btn.setEnabled(False)
        ctx.lbl.setText("⏳ Testando…")
        ctx.lbl.setStyleSheet("color: #ccaa00;")

        worker = OracleConnectionTestWorker(
            host=ctx.f_host.text(),
            port=ctx.f_port.text(),
            service=ctx.f_service.text(),
            user=ctx.f_user.text(),
            password=ctx.f_password.text(),
            parent=self,
        )
        setattr(self, ctx.worker_attr, worker)

        def _on_result(ok: bool, msg: str, _ms: int) -> None:
            if ok:
                ctx.lbl.setText(f"✔ {msg}")
                ctx.lbl.setStyleSheet("color: #4caf50;")
            else:
                ctx.lbl.setText(f"✖ {msg}")
                ctx.lbl.setStyleSheet("color: #e57373;")
            if ctx.btn:
                ctx.btn.setEnabled(True)
            worker.deleteLater()
            setattr(self, ctx.worker_attr, None)

        worker.resultado.connect(_on_result)
        worker.start()

    def _salvar_configuracoes(self) -> None:
        """Escreve todos os campos do painel de configurações no arquivo .env."""
        from interface_grafica.fisconforme.path_resolver import get_env_path

        env_path = get_env_path()
        conteudo = env_path.read_text(encoding="utf-8") if env_path.exists() else ""

        campos: dict[str, str] = {
            "ORACLE_HOST": self._cfg_host.text().strip(),
            "ORACLE_PORT": self._cfg_port.text().strip(),
            "ORACLE_SERVICE": self._cfg_service.text().strip(),
            "DB_USER": self._cfg_user.text().strip(),
            "DB_PASSWORD": self._cfg_password.text().strip(),
            "ORACLE_HOST_1": self._cfg_host_1.text().strip(),
            "ORACLE_PORT_1": self._cfg_port_1.text().strip(),
            "ORACLE_SERVICE_1": self._cfg_service_1.text().strip(),
            "DB_USER_1": self._cfg_user_1.text().strip(),
            "DB_PASSWORD_1": self._cfg_password_1.text().strip(),
            "LOG_LEVEL": self._cfg_log_level.currentText(),
            "CACHE_ENABLED": "true" if self._cfg_cache_enabled.isChecked() else "false",
            "CACHE_TTL": self._cfg_cache_ttl.text().strip(),
            "DASHBOARD_THEME": self._cfg_theme.currentText(),
        }

        for chave, valor in campos.items():
            if re.search(rf"^{chave}=", conteudo, flags=re.MULTILINE):
                conteudo = re.sub(
                    rf"^{chave}=.*$",
                    f"{chave}={valor}",
                    conteudo,
                    flags=re.MULTILINE,
                )
            else:
                conteudo = conteudo.rstrip() + f"\n{chave}={valor}\n"

        env_path.parent.mkdir(parents=True, exist_ok=True)
        env_path.write_text(conteudo.strip() + "\n", encoding="utf-8")
        self._cfg_status_label.setText("✔ Configurações salvas.")
        self.status.showMessage("Configurações Oracle salvas com sucesso.", 4000)



    def _build_tab_consulta(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        filter_box = QGroupBox("Filtros")
        filter_layout = QVBoxLayout(filter_box)
        form = QHBoxLayout()
        self.filter_column = QComboBox()
        self.filter_operator = QComboBox()
        self.filter_operator.addItems(["contem", "igual", "comeca com", "termina com", ">", ">=", "<", "<=", "e nulo", "nao e nulo"])
        self.filter_value = QLineEdit()
        self.filter_value.setPlaceholderText("Valor do filtro")
        self.btn_add_filter = QPushButton("Adicionar filtro")
        self.btn_clear_filters = QPushButton("Limpar filtros")
        form.addWidget(QLabel("Coluna"))
        form.addWidget(self.filter_column)
        form.addWidget(QLabel("Operador"))
        form.addWidget(self.filter_operator)
        form.addWidget(QLabel("Valor"))
        form.addWidget(self.filter_value)
        form.addWidget(self.btn_add_filter)
        form.addWidget(self.btn_clear_filters)
        filter_layout.addLayout(form)

        self.filter_list = QListWidget()
        self.filter_list.setMaximumHeight(90)
        filter_layout.addWidget(self.filter_list)

        filter_actions = QHBoxLayout()
        self.btn_remove_filter = QPushButton("Remover filtro selecionado")
        self.btn_choose_columns = QPushButton("Selecionar colunas")
        self.consulta_profile = QComboBox()
        self.consulta_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_apply_consulta_profile = QPushButton("Aplicar perfil")
        self.btn_save_consulta_profile = QPushButton("Salvar perfil")
        self.btn_consulta_destacar = self._criar_botao_destacar()
        self.lbl_page = QLabel("Linhas filtradas: 0")
        filter_actions.addWidget(self.btn_remove_filter)
        filter_actions.addWidget(self.btn_choose_columns)
        filter_actions.addWidget(self.consulta_profile)
        filter_actions.addWidget(self.btn_apply_consulta_profile)
        filter_actions.addWidget(self.btn_save_consulta_profile)
        filter_actions.addWidget(self.btn_consulta_destacar)
        filter_actions.addStretch()
        filter_actions.addWidget(self.lbl_page)
        filter_layout.addLayout(filter_actions)
        layout.addWidget(filter_box)

        export_box = QGroupBox("Exportacao")
        export_layout = QHBoxLayout(export_box)
        self.btn_export_excel_full = QPushButton("Excel - tabela completa")
        self.btn_export_excel_filtered = QPushButton("Excel - tabela filtrada")
        self.btn_export_excel_visible = QPushButton("Excel - colunas visiveis")
        self.btn_export_docx = QPushButton("Relatorio Word")
        self.btn_export_html_txt = QPushButton("TXT com HTML")
        for btn in [
            self.btn_export_excel_full,
            self.btn_export_excel_filtered,
            self.btn_export_excel_visible,
            self.btn_export_docx,
            self.btn_export_html_txt,
        ]:
            export_layout.addWidget(btn)
        layout.addWidget(export_box)

        quick_filter_layout = QHBoxLayout()
        self.qf_norm = QLineEdit()
        self.qf_norm.setPlaceholderText("Filtrar Desc. Norm")
        self.qf_desc = QLineEdit()
        self.qf_desc.setPlaceholderText("Filtrar Descricao (ex.: buch 18)")
        self.qf_ncm = QLineEdit()
        self.qf_ncm.setPlaceholderText("Filtrar NCM")
        self.qf_cest = QLineEdit()
        self.qf_cest.setPlaceholderText("Filtrar CEST")
        
        for w in [self.qf_norm, self.qf_desc, self.qf_ncm, self.qf_cest]:
            w.setMaximumWidth(200)
            quick_filter_layout.addWidget(w)
        quick_filter_layout.addStretch()
        layout.addLayout(quick_filter_layout)

        self.table_view = QTableView()
        self.table_view.setModel(self.table_model)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table_view.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSortingEnabled(False)
        self.table_view.setWordWrap(True)
        self.table_view.verticalHeader().setDefaultSectionSize(60)
        self.table_view.horizontalHeader().setMinimumSectionSize(40)
        self.table_view.horizontalHeader().setDefaultSectionSize(200)
        self.table_view.horizontalHeader().setMaximumSectionSize(300)
        self.table_view.setStyleSheet("QTableView::item { padding: 4px 2px; }")
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(lambda pos: self._abrir_menu_contexto_celula("consulta", self.table_view, pos))
        self.table_view.horizontalHeader().setSectionsMovable(True)
        self.table_view.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.table_view, 1)
        return tab

    def _build_tab_agregacao(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 0, 0, 0)
        
        splitter = QSplitter(Qt.Vertical)

        # Top box
        top_box = QGroupBox("Tabela Agrupada Filtravel (Selecione linhas para agregar)")
        top_layout = QVBoxLayout(top_box)
        top_layout.setContentsMargins(4, 12, 4, 4)

        toolbar = QHBoxLayout()
        self.btn_abrir_grup_sql = QPushButton("Abrir tabela agrupada")
        self.btn_abrir_grup_sql.clicked.connect(self._abrir_tabela_agrupada)
        toolbar.addWidget(self.btn_abrir_grup_sql)

        self.btn_agregar_descricoes = QPushButton("Agregar Descricoes (da selecao)")
        toolbar.addWidget(self.btn_agregar_descricoes)

        self.btn_reprocessar_agregacao = self._criar_botao_destacar("Reprocessar")
        toolbar.addWidget(self.btn_reprocessar_agregacao)

        toolbar.addStretch()
        top_layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.top_filter_desc = QLineEdit()
        self.top_filter_desc.setPlaceholderText("Filtrar Descricao (ex.: buch 18)")
        self.top_filter_ncm = QLineEdit()
        self.top_filter_ncm.setPlaceholderText("Filtrar NCM")
        self.top_filter_cest = QLineEdit()
        self.top_filter_cest.setPlaceholderText("Filtrar CEST")
        self.top_filter_texto = QLineEdit()
        self.top_filter_texto.setPlaceholderText("Busca global...")
        self.btn_top_match_ncm_cest = QPushButton("NCM+CEST iguais")
        self.btn_top_match_ncm_cest_gtin = QPushButton("NCM+CEST+GTIN iguais")
        self.btn_clear_top_agg_filters = QPushButton("Limpar filtros")
        self.top_profile = QComboBox()
        self.top_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_apply_top_profile = QPushButton("Perfil")
        self.btn_save_top_profile = QPushButton("Salvar perfil")
        self.btn_top_colunas = QPushButton("Colunas")
        self.btn_top_destacar = self._criar_botao_destacar()
        filtros.addWidget(self.top_filter_desc)
        filtros.addWidget(self.top_filter_ncm)
        filtros.addWidget(self.top_filter_cest)
        filtros.addWidget(self.top_filter_texto)
        filtros.addWidget(self.btn_top_match_ncm_cest)
        filtros.addWidget(self.btn_top_match_ncm_cest_gtin)
        filtros.addWidget(self.top_profile)
        filtros.addWidget(self.btn_apply_top_profile)
        filtros.addWidget(self.btn_save_top_profile)
        filtros.addWidget(self.btn_top_colunas)
        filtros.addWidget(self.btn_top_destacar)
        filtros.addWidget(self.btn_clear_top_agg_filters)
        top_layout.addLayout(filtros)

        self.lbl_top_table_status = QLabel("Nenhum dado.")
        top_layout.addWidget(self.lbl_top_table_status)

        self.aggregation_table = QTableView()
        self.aggregation_table.setModel(self.aggregation_table_model)
        self.aggregation_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.aggregation_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.aggregation_table.setAlternatingRowColors(True)
        self.aggregation_table.setSortingEnabled(True)
        self.aggregation_table.setWordWrap(True)
        self.aggregation_table.verticalHeader().setDefaultSectionSize(40)
        self.aggregation_table.horizontalHeader().setMinimumSectionSize(40)
        self.aggregation_table.horizontalHeader().setDefaultSectionSize(150)
        self.aggregation_table.horizontalHeader().setMaximumSectionSize(400)
        self.aggregation_table.horizontalHeader().setSectionsMovable(True)
        self.aggregation_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        top_layout.addWidget(self.aggregation_table)

        splitter.addWidget(top_box)

        # Bottom box
        bottom_box = QGroupBox("Linhas Agregadas (Mesma Tabela de Referencia)")
        bottom_layout = QVBoxLayout(bottom_box)
        bottom_layout.setContentsMargins(4, 12, 4, 4)

        bottom_filtros = QHBoxLayout()
        self.bot_filter_desc_norm = QLineEdit()
        self.bot_filter_desc_norm.setPlaceholderText("Filtrar Desc. Norm")
        self.bot_filter_desc_orig = QLineEdit()
        self.bot_filter_desc_orig.setPlaceholderText("Filtrar Descricao (ex.: whisky 12)")
        self.bot_filter_ncm = QLineEdit()
        self.bot_filter_ncm.setPlaceholderText("Filtrar NCM")
        self.bot_filter_cest = QLineEdit()
        self.bot_filter_cest.setPlaceholderText("Filtrar CEST")
        self.btn_bottom_match_ncm_cest = QPushButton("NCM+CEST iguais")
        self.btn_bottom_match_ncm_cest_gtin = QPushButton("NCM+CEST+GTIN iguais")
        self.btn_clear_bottom_agg_filters = QPushButton("Limpar filtros")
        self.bottom_profile = QComboBox()
        self.bottom_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_apply_bottom_profile = QPushButton("Perfil")
        self.btn_save_bottom_profile = QPushButton("Salvar perfil")
        self.btn_bottom_colunas = QPushButton("Colunas")
        self.btn_bottom_destacar = self._criar_botao_destacar()
        bottom_filtros.addWidget(self.bot_filter_desc_norm)
        bottom_filtros.addWidget(self.bot_filter_desc_orig)
        bottom_filtros.addWidget(self.bot_filter_ncm)
        bottom_filtros.addWidget(self.bot_filter_cest)
        bottom_filtros.addWidget(self.btn_bottom_match_ncm_cest)
        bottom_filtros.addWidget(self.btn_bottom_match_ncm_cest_gtin)
        bottom_filtros.addWidget(self.bottom_profile)
        bottom_filtros.addWidget(self.btn_apply_bottom_profile)
        bottom_filtros.addWidget(self.btn_save_bottom_profile)
        bottom_filtros.addWidget(self.btn_bottom_colunas)
        bottom_filtros.addWidget(self.btn_bottom_destacar)
        bottom_filtros.addWidget(self.btn_clear_bottom_agg_filters)
        bottom_layout.addLayout(bottom_filtros)

        self.lbl_bot_table_status = QLabel("Nenhuma linha agrupada.")
        bottom_layout.addWidget(self.lbl_bot_table_status)

        self.results_table = QTableView()
        self.results_table.setModel(self.results_table_model)
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setSortingEnabled(True)
        self.results_table.setWordWrap(True)
        self.results_table.verticalHeader().setDefaultSectionSize(40)
        self.results_table.horizontalHeader().setMinimumSectionSize(40)
        self.results_table.horizontalHeader().setDefaultSectionSize(150)
        self.results_table.horizontalHeader().setMaximumSectionSize(400)
        self.results_table.horizontalHeader().setSectionsMovable(True)
        self.results_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        bottom_layout.addWidget(self.results_table)

        tb_acoes = QToolBar()
        self.btn_reverter_agregacao = QPushButton(
            QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_BrowserReload),
            "Reverter agrupamento",
        )
        self.btn_desfazer_agregacao = QPushButton(
            QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_ArrowLeft),
            "Desfazer selecao",
        )
        self.btn_reverter_agregacao.clicked.connect(self.reverter_agregacao)
        tb_acoes.addWidget(self.btn_reverter_agregacao)
        self.btn_desfazer_agregacao.clicked.connect(self._desfazer_agregacao)
        tb_acoes.addWidget(self.btn_desfazer_agregacao)
        bottom_layout.addWidget(tb_acoes)

        splitter.addWidget(bottom_box)
        splitter.setSizes([500, 300])

        # Aliases legados usados em outros trechos da tela.
        self.btn_open_editable_table = self.btn_abrir_grup_sql
        self.btn_execute_aggregation = self.btn_agregar_descricoes
        self.btn_recalc_defaults = self.btn_reprocessar_agregacao
        self.btn_recalc_totals = self.btn_reprocessar_agregacao
        self.aggregation_table_view = self.aggregation_table
        self.results_table_view = self.results_table
        self.aqf_norm = self.top_filter_texto
        self.aqf_desc = self.top_filter_desc
        self.aqf_ncm = self.top_filter_ncm
        self.aqf_cest = self.top_filter_cest
        self.bqf_norm = self.bot_filter_desc_norm
        self.bqf_desc = self.bot_filter_desc_orig
        self.bqf_ncm = self.bot_filter_ncm
        self.bqf_cest = self.bot_filter_cest

        layout.addWidget(splitter)
        return tab

    # ------------------------------------------------------------------
    # Aba Consulta SQL
    # ------------------------------------------------------------------
    def _build_tab_sql_query(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # --- Linha superior: seletor de SQL + botAes ---
        top_bar = QHBoxLayout()
        top_bar.addWidget(QLabel("SQL:"))
        self.sql_combo = QComboBox()
        self.sql_combo.setMinimumWidth(300)
        top_bar.addWidget(self.sql_combo, 1)
        self.btn_sql_execute = QPushButton("Executar Consulta")
        self.btn_sql_execute.setStyleSheet("QPushButton { font-weight: bold; padding: 6px 16px; }")
        self.btn_sql_export = QPushButton("Exportar Excel")
        self.btn_sql_destacar = self._criar_botao_destacar()
        top_bar.addWidget(self.btn_sql_execute)
        top_bar.addWidget(self.btn_sql_export)
        top_bar.addWidget(self.btn_sql_destacar)
        layout.addLayout(top_bar)

        # --- Splitter: SQL + parametros (esquerda) | resultados (direita) ---
        splitter = QSplitter(Qt.Vertical)

        # Parte superior: SQL + parametros
        upper_widget = QWidget()
        upper_layout = QHBoxLayout(upper_widget)
        upper_layout.setContentsMargins(0, 0, 0, 0)

        # Visualizador SQL
        sql_group = QGroupBox("Texto SQL")
        sql_group_layout = QVBoxLayout(sql_group)
        self.sql_text_view = QPlainTextEdit()
        self.sql_text_view.setReadOnly(True)
        self.sql_text_view.setStyleSheet(
            "QPlainTextEdit { font-family: 'Consolas', 'Courier New', monospace; "
            "font-size: 12px; background: #1e1e2e; color: #cdd6f4; "
            "border: 1px solid #45475a; border-radius: 4px; padding: 8px; }"
        )
        self.sql_text_view.setMinimumHeight(120)
        sql_group_layout.addWidget(self.sql_text_view)
        upper_layout.addWidget(sql_group, 3)

        # Painel de parametros
        param_group = QGroupBox("Parametros")
        param_outer_layout = QVBoxLayout(param_group)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.sql_param_container = QWidget()
        self.sql_param_form = QFormLayout(self.sql_param_container)
        self.sql_param_form.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        scroll.setWidget(self.sql_param_container)
        param_outer_layout.addWidget(scroll)
        upper_layout.addWidget(param_group, 1)

        splitter.addWidget(upper_widget)

        # Parte inferior: resultados
        result_widget = QWidget()
        result_layout = QVBoxLayout(result_widget)
        result_layout.setContentsMargins(0, 0, 0, 0)

        # Status
        self.sql_status_label = QLabel("Selecione um SQL e clique em Executar.")
        self.sql_status_label.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #f0f4ff; border-radius: 4px; "
            "border: 1px solid #d0d8e8; color: #334155; font-weight: bold; }"
        )
        result_layout.addWidget(self.sql_status_label)

        # Filtro rApido nos resultados
        sql_filter_bar = QHBoxLayout()
        self.sql_result_search = QLineEdit()
        self.sql_result_search.setPlaceholderText("Buscar nos resultados...")
        sql_filter_bar.addWidget(self.sql_result_search)
        self.sql_result_page_label = QLabel("Total: 0")
        sql_filter_bar.addWidget(self.sql_result_page_label)
        result_layout.addLayout(sql_filter_bar)

        # Tabela de resultados
        self.sql_result_table = QTableView()
        self.sql_result_table.setModel(self.sql_result_model)
        self.sql_result_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.sql_result_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.sql_result_table.setAlternatingRowColors(True)
        self.sql_result_table.setSortingEnabled(False)
        self.sql_result_table.setWordWrap(True)
        self.sql_result_table.verticalHeader().setDefaultSectionSize(60)
        self.sql_result_table.horizontalHeader().setMinimumSectionSize(40)
        self.sql_result_table.horizontalHeader().setDefaultSectionSize(200)
        self.sql_result_table.horizontalHeader().setMaximumSectionSize(400)
        self.sql_result_table.horizontalHeader().setStretchLastSection(True)
        self.sql_result_table.setStyleSheet("QTableView::item { padding: 4px 2px; }")
        result_layout.addWidget(self.sql_result_table, 1)

        splitter.addWidget(result_widget)
        splitter.setSizes([280, 500])

        layout.addWidget(splitter, 1)
        return tab

    def _build_tab_conversao(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        toolbar = QHBoxLayout()
        self.btn_refresh_conversao = QPushButton("Recarregar")
        self.btn_refresh_conversao.setIcon(QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_BrowserReload))
        self.chk_show_single_unit = QCheckBox("Mostrar itens de unidade unica")
        self.chk_show_single_unit.setChecked(False)
        self.btn_export_conversao = QPushButton("Exportar Excel")
        self.btn_import_conversao = QPushButton("Importar Excel")
        self.btn_conversao_destacar = self._criar_botao_destacar()
        self.btn_recalcular_fatores = self._criar_botao_destacar("Recalcular fatores")
        self.btn_recalcular_fatores.setEnabled(False)
        self.conversao_profile = QComboBox()
        self.conversao_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_apply_conversao_profile = QPushButton("Perfil")
        self.btn_save_conversao_profile = QPushButton("Salvar perfil")
        self.btn_conversao_colunas = QPushButton("Colunas")
        
        toolbar.addWidget(self.btn_refresh_conversao)
        toolbar.addWidget(self.chk_show_single_unit)
        toolbar.addStretch()
        toolbar.addWidget(self.btn_recalcular_fatores)
        toolbar.addWidget(self.conversao_profile)
        toolbar.addWidget(self.btn_apply_conversao_profile)
        toolbar.addWidget(self.btn_save_conversao_profile)
        toolbar.addWidget(self.btn_conversao_colunas)
        toolbar.addWidget(self.btn_conversao_destacar)
        toolbar.addWidget(self.btn_import_conversao)
        layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.conv_filter_id = QComboBox()
        self.conv_filter_id.setEditable(True)
        self.conv_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.conv_filter_id.setMinimumWidth(220)
        self.conv_filter_id.lineEdit().setPlaceholderText("Filtrar id_agrupado")
        self.conv_filter_desc = QLineEdit()
        self.conv_filter_desc.setPlaceholderText("Filtrar descr_padrao")
        filtros.addWidget(self.conv_filter_id)
        filtros.addWidget(self.conv_filter_desc)
        layout.addLayout(filtros)

        self.panel_unid_ref = QGroupBox("Alterar Unidade de Referencia do Produto Selecionado")
        panel_layout = QHBoxLayout(self.panel_unid_ref)
        self.lbl_produto_sel = QLabel("Nenhum produto selecionado")
        self.lbl_produto_sel.setStyleSheet("font-weight: bold; color: #1e40af;")
        self.combo_unid_ref = QComboBox()
        self.btn_apply_unid_ref = QPushButton("Aplicar a todos os itens")
        self.btn_apply_unid_ref.setStyleSheet("font-weight: bold;")
        self.btn_apply_unid_ref.setEnabled(False)
        self.combo_unid_ref.setEnabled(False)
        panel_layout.addWidget(self.lbl_produto_sel)
        panel_layout.addWidget(QLabel("   -> Nova unid_ref:"))
        panel_layout.addWidget(self.combo_unid_ref)
        panel_layout.addWidget(self.btn_apply_unid_ref)
        panel_layout.addStretch()
        layout.addWidget(self.panel_unid_ref)

        self.conversion_table = QTableView()
        self.conversion_table.setModel(self.conversion_model)
        self.conversion_table.setAlternatingRowColors(True)
        self.conversion_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.conversion_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.conversion_table.setSortingEnabled(True)
        self.conversion_table.horizontalHeader().setSectionsMovable(True)
        self.conversion_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.conversion_table)

        return tab

    def _build_tab_analise_lote_cnpj(self) -> QWidget:
        """Retorna o painel Fisconforme não Atendido como aba do QTabWidget."""
        try:
            from ..fisconforme import FisconformeNaoAtendidoPanel
            return FisconformeNaoAtendidoPanel()
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning(
                "Não foi possível carregar o painel Fisconforme: %s", exc
            )
            from PySide6.QtWidgets import QLabel
            lbl = QLabel(f"Painel Fisconforme indisponível: {exc}")
            lbl.setWordWrap(True)
            return lbl

    def _build_tab_logs(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        layout.addWidget(self.log_view)
        return tab

    def _build_tab_mov_estoque(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.lbl_mov_estoque_titulo = QLabel("Tabela: mov_estoque")
        self.lbl_mov_estoque_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #e2e8f0; background: #1e293b; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_mov_estoque_titulo)

        filtros = QHBoxLayout()
        self.mov_filter_id = QComboBox()
        self.mov_filter_id.setEditable(True)
        self.mov_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.mov_filter_id.setMinimumWidth(250)
        self.mov_filter_id.lineEdit().setPlaceholderText("Filtrar id_agrupado")
        
        self.mov_filter_desc = QLineEdit()
        self.mov_filter_desc.setPlaceholderText("Filtrar descriCAo")
        
        self.mov_filter_ncm = QLineEdit()
        self.mov_filter_ncm.setPlaceholderText("Filtrar NCM")
        
        self.mov_filter_tipo = QComboBox()
        self.mov_filter_tipo.addItems(["Todos", "Entradas", "Saidas"])
        
        self.mov_filter_texto = QLineEdit()
        self.mov_filter_texto.setPlaceholderText("Busca geral...")

        for widget in [self.mov_filter_id, self.mov_filter_desc, self.mov_filter_ncm, self.mov_filter_tipo, self.mov_filter_texto]:
            filtros.addWidget(widget)
        layout.addLayout(filtros)

        filtros_avancados = QHBoxLayout()
        self.mov_filter_data_col = QComboBox()
        self.mov_filter_data_col.addItems(["Dt_doc", "Dt_e_s"])
        self.mov_filter_data_ini = QDateEdit()
        self.mov_filter_data_ini.setCalendarPopup(True)
        self.mov_filter_data_ini.setDisplayFormat("dd/MM/yyyy")
        self.mov_filter_data_ini.setSpecialValueText("Data inicial")
        self.mov_filter_data_ini.setMinimumDate(QDate(1900, 1, 1))
        self.mov_filter_data_ini.setDate(self.mov_filter_data_ini.minimumDate())
        self.mov_filter_data_fim = QDateEdit()
        self.mov_filter_data_fim.setCalendarPopup(True)
        self.mov_filter_data_fim.setDisplayFormat("dd/MM/yyyy")
        self.mov_filter_data_fim.setSpecialValueText("Data final")
        self.mov_filter_data_fim.setMinimumDate(QDate(1900, 1, 1))
        self.mov_filter_data_fim.setDate(self.mov_filter_data_fim.minimumDate())
        self.mov_filter_num_col = QComboBox()
        self.mov_filter_num_col.addItems(["saldo_estoque_anual", "custo_medio_anual", "entr_desac_anual", "q_conv", "preco_item", "preco_unit"])
        self.mov_filter_num_min = QLineEdit()
        self.mov_filter_num_min.setPlaceholderText("Min numerico")
        self.mov_filter_num_max = QLineEdit()
        self.mov_filter_num_max.setPlaceholderText("Max numerico")
        self.mov_profile = QComboBox()
        self.mov_profile.addItems(["Padrao", "Contribuinte", "Auditoria", "Auditoria Fiscal", "Estoque", "Custos"])
        self.btn_mov_profile = QPushButton("Perfil")
        self.btn_mov_save_profile = QPushButton("Salvar perfil")
        self.btn_mov_colunas = QPushButton("Colunas")
        self.btn_mov_destacar = self._criar_botao_destacar()
        self.btn_export_mov_estoque = QPushButton("Exportar Excel")
        for widget in [
            QLabel("Data"),
            self.mov_filter_data_col,
            self.mov_filter_data_ini,
            self.mov_filter_data_fim,
            QLabel("Numero"),
            self.mov_filter_num_col,
            self.mov_filter_num_min,
            self.mov_filter_num_max,
            self.mov_profile,
            self.btn_mov_profile,
            self.btn_mov_save_profile,
            self.btn_mov_colunas,
            self.btn_mov_destacar,
            self.btn_export_mov_estoque,
        ]:
            filtros_avancados.addWidget(widget)
        layout.addLayout(filtros_avancados)

        self.lbl_mov_estoque_status = QLabel("Selecione um CNPJ para carregar as movimentacoes.")
        self.lbl_mov_estoque_status.setStyleSheet("QLabel { padding: 4px; color: #475569; }")
        layout.addWidget(self.lbl_mov_estoque_status)

        self.lbl_mov_estoque_filtros = QLabel("Filtros ativos: nenhum")
        self.lbl_mov_estoque_filtros.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #cbd5e1; background: #0f172a; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_mov_estoque_filtros)

        self.mov_estoque_table = QTableView()
        self.mov_estoque_table.setModel(self.mov_estoque_model)
        self.mov_estoque_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.mov_estoque_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.mov_estoque_table.setAlternatingRowColors(True)
        self.mov_estoque_table.setSortingEnabled(True)
        self.mov_estoque_table.setWordWrap(True)
        self.mov_estoque_table.verticalHeader().setDefaultSectionSize(40)
        self.mov_estoque_table.horizontalHeader().setMinimumSectionSize(40)
        self.mov_estoque_table.horizontalHeader().setDefaultSectionSize(110)
        self.mov_estoque_table.horizontalHeader().setMaximumSectionSize(400)
        self.mov_estoque_table.horizontalHeader().setSectionsMovable(True)
        self.mov_estoque_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.mov_estoque_table.customContextMenuRequested.connect(lambda pos: self._abrir_menu_contexto_celula("mov_estoque", self.mov_estoque_table, pos))
        self.mov_estoque_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.mov_estoque_table)

        return tab

    def _build_tab_estoque(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.estoque_tabs = QTabWidget()
        
        self.tab_mov_estoque = self._build_tab_mov_estoque()
        self.estoque_tabs.addTab(self.tab_mov_estoque, "Tabela mov_estoque")

        self.tab_aba_mensal = self._build_tab_aba_mensal()
        self.estoque_tabs.addTab(self.tab_aba_mensal, "Tabela mensal")

        self.tab_aba_anual = self._build_tab_aba_anual()
        self.estoque_tabs.addTab(self.tab_aba_anual, "Tabela anual")

        self.tab_resumo_global = self._build_tab_resumo_global()
        self.estoque_tabs.addTab(self.tab_resumo_global, "Resumo Global")

        self.tab_produtos_selecionados = self._build_tab_produtos_selecionados()
        self.estoque_tabs.addTab(self.tab_produtos_selecionados, "Produtos selecionados")

        self.tab_id_agrupados = self._build_tab_id_agrupados()
        self.estoque_tabs.addTab(self.tab_id_agrupados, "id_agrupados")

        layout.addWidget(self.estoque_tabs)
        return tab

    def _build_tab_produtos_selecionados(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        tab.setStyleSheet(
            """
            QWidget {
                background: #252526;
                color: #f3f4f6;
            }
            QLabel {
                color: #e5e7eb;
            }
            QLineEdit, QComboBox, QDateEdit {
                background: #323438;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 8px;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QPushButton {
                background: #34373d;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background: #40444b;
            }
            QPushButton:pressed {
                background: #2f3338;
            }
            QTableView {
                background: #1f1f1f;
                alternate-background-color: #262626;
                color: #f9fafb;
                gridline-color: #3f3f46;
                border: 1px solid #3f3f46;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QHeaderView::section {
                background: #18181b;
                color: #f9fafb;
                border: 1px solid #3f3f46;
                padding: 6px 8px;
                font-weight: bold;
            }
            """
        )

        self.lbl_produtos_sel_titulo = QLabel("Tabela: produtos_selecionados")
        self.lbl_produtos_sel_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_produtos_sel_titulo)

        toolbar = QHBoxLayout()
        self.btn_refresh_produtos_sel = QPushButton("Atualizar resumo")
        self.btn_apply_produtos_sel_filters = QPushButton("Aplicar filtros")
        self.btn_clear_produtos_sel_filters = QPushButton("Limpar filtros")
        self.produtos_sel_profile = QComboBox()
        self.produtos_sel_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_produtos_sel_profile = QPushButton("Perfil")
        self.btn_produtos_sel_save_profile = QPushButton("Salvar perfil")
        self.btn_colunas_produtos_sel = QPushButton("Colunas")
        self.btn_destacar_produtos_sel = self._criar_botao_destacar()
        self.btn_export_produtos_sel = QPushButton("Exportar Excel")
        toolbar.addWidget(self.btn_refresh_produtos_sel)
        toolbar.addWidget(self.btn_apply_produtos_sel_filters)
        toolbar.addWidget(self.btn_clear_produtos_sel_filters)
        toolbar.addStretch()
        toolbar.addWidget(self.produtos_sel_profile)
        toolbar.addWidget(self.btn_produtos_sel_profile)
        toolbar.addWidget(self.btn_produtos_sel_save_profile)
        toolbar.addWidget(self.btn_colunas_produtos_sel)
        toolbar.addWidget(self.btn_destacar_produtos_sel)
        toolbar.addWidget(self.btn_export_produtos_sel)
        layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.produtos_sel_filter_id = QComboBox()
        self.produtos_sel_filter_id.setEditable(True)
        self.produtos_sel_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.produtos_sel_filter_id.setMinimumWidth(220)
        self.produtos_sel_filter_id.lineEdit().setPlaceholderText("Filtrar id_agregado")
        self.produtos_sel_filter_desc = QLineEdit()
        self.produtos_sel_filter_desc.setPlaceholderText("Filtrar descricao")
        self.produtos_sel_filter_ano_ini = QComboBox()
        self.produtos_sel_filter_ano_ini.addItem("Todos")
        self.produtos_sel_filter_ano_fim = QComboBox()
        self.produtos_sel_filter_ano_fim.addItem("Todos")
        self.produtos_sel_filter_data_ini = QDateEdit()
        self.produtos_sel_filter_data_ini.setCalendarPopup(True)
        self.produtos_sel_filter_data_ini.setDisplayFormat("dd/MM/yyyy")
        self.produtos_sel_filter_data_ini.setSpecialValueText("Data inicial")
        self.produtos_sel_filter_data_ini.setMinimumDate(QDate(1900, 1, 1))
        self.produtos_sel_filter_data_ini.setDate(self.produtos_sel_filter_data_ini.minimumDate())
        self.produtos_sel_filter_data_fim = QDateEdit()
        self.produtos_sel_filter_data_fim.setCalendarPopup(True)
        self.produtos_sel_filter_data_fim.setDisplayFormat("dd/MM/yyyy")
        self.produtos_sel_filter_data_fim.setSpecialValueText("Data final")
        self.produtos_sel_filter_data_fim.setMinimumDate(QDate(1900, 1, 1))
        self.produtos_sel_filter_data_fim.setDate(self.produtos_sel_filter_data_fim.minimumDate())
        self.produtos_sel_filter_texto = QLineEdit()
        self.produtos_sel_filter_texto.setPlaceholderText("Busca ampla...")
        for widget in [
            self.produtos_sel_filter_id,
            self.produtos_sel_filter_desc,
            QLabel("Ano inicial"),
            self.produtos_sel_filter_ano_ini,
            QLabel("Ano final"),
            self.produtos_sel_filter_ano_fim,
            QLabel("Data inicial"),
            self.produtos_sel_filter_data_ini,
            QLabel("Data final"),
            self.produtos_sel_filter_data_fim,
            self.produtos_sel_filter_texto,
        ]:
            filtros.addWidget(widget)
        layout.addLayout(filtros)

        self.lbl_produtos_sel_status = QLabel("Selecione um CNPJ para consolidar os produtos analisados.")
        self.lbl_produtos_sel_status.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #101827; border: 1px solid #374151; border-radius: 4px; color: #e5e7eb; }"
        )
        layout.addWidget(self.lbl_produtos_sel_status)

        self.lbl_produtos_sel_filtros = QLabel("Filtros ativos: nenhum")
        self.lbl_produtos_sel_filtros.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #dbeafe; background: #0f1b33; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_produtos_sel_filtros)

        self.lbl_produtos_sel_resumo = QLabel("Recorte atual: mov_estoque 0 | mensal 0 | anual 0")
        self.lbl_produtos_sel_resumo.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #fef3c7; background: #2a1f0f; border: 1px solid #7c5a18; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_produtos_sel_resumo)

        self.produtos_sel_table = QTableView()
        self.produtos_sel_table.setModel(self.produtos_selecionados_model)
        self.produtos_sel_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.produtos_sel_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.produtos_sel_table.setAlternatingRowColors(True)
        self.produtos_sel_table.setSortingEnabled(True)
        self.produtos_sel_table.setWordWrap(True)
        self.produtos_sel_table.verticalHeader().setDefaultSectionSize(40)
        self.produtos_sel_table.horizontalHeader().setMinimumSectionSize(40)
        self.produtos_sel_table.horizontalHeader().setDefaultSectionSize(180)
        self.produtos_sel_table.horizontalHeader().setMaximumSectionSize(420)
        self.produtos_sel_table.horizontalHeader().setStretchLastSection(True)
        self.produtos_sel_table.horizontalHeader().setSectionsMovable(True)
        self.produtos_sel_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.produtos_sel_table.setStyleSheet(
            "QTableView::item { padding: 4px 2px; }"
            "QTableCornerButton::section { background: #18181b; border: 1px solid #3f3f46; }"
        )
        layout.addWidget(self.produtos_sel_table, 1)
        return tab

    def _build_tab_id_agrupados(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.lbl_id_agrupados_titulo = QLabel("Tabela: id_agrupados")
        self.lbl_id_agrupados_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_id_agrupados_titulo)

        toolbar = QHBoxLayout()
        self.btn_refresh_id_agrupados = QPushButton("Recarregar")
        self.btn_apply_id_agrupados_filters = QPushButton("Aplicar filtros")
        self.btn_clear_id_agrupados_filters = QPushButton("Limpar filtros")
        self.id_agrupados_profile = QComboBox()
        self.id_agrupados_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_id_agrupados_profile = QPushButton("Perfil")
        self.btn_id_agrupados_save_profile = QPushButton("Salvar perfil")
        self.btn_id_agrupados_colunas = QPushButton("Colunas")
        self.btn_destacar_id_agrupados = self._criar_botao_destacar()
        self.btn_export_id_agrupados = QPushButton("Exportar Excel")
        toolbar.addWidget(self.btn_refresh_id_agrupados)
        toolbar.addWidget(self.btn_apply_id_agrupados_filters)
        toolbar.addWidget(self.btn_clear_id_agrupados_filters)
        toolbar.addStretch()
        toolbar.addWidget(self.id_agrupados_profile)
        toolbar.addWidget(self.btn_id_agrupados_profile)
        toolbar.addWidget(self.btn_id_agrupados_save_profile)
        toolbar.addWidget(self.btn_id_agrupados_colunas)
        toolbar.addWidget(self.btn_destacar_id_agrupados)
        toolbar.addWidget(self.btn_export_id_agrupados)
        layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.id_agrupados_filter_id = QComboBox()
        self.id_agrupados_filter_id.setEditable(True)
        self.id_agrupados_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.id_agrupados_filter_id.setMinimumWidth(240)
        self.id_agrupados_filter_id.lineEdit().setPlaceholderText("Filtrar id_agrupado")
        self.id_agrupados_filter_texto = QLineEdit()
        self.id_agrupados_filter_texto.setPlaceholderText("Busca ampla...")
        filtros.addWidget(self.id_agrupados_filter_id)
        filtros.addWidget(self.id_agrupados_filter_texto)
        layout.addLayout(filtros)

        self.lbl_id_agrupados_status = QLabel("Selecione um CPF/CNPJ para carregar os id_agrupados.")
        self.lbl_id_agrupados_status.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #101827; border: 1px solid #374151; border-radius: 4px; color: #e5e7eb; }"
        )
        layout.addWidget(self.lbl_id_agrupados_status)

        self.lbl_id_agrupados_filtros = QLabel("Filtros ativos: nenhum")
        self.lbl_id_agrupados_filtros.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #dbeafe; background: #0f1b33; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_id_agrupados_filtros)

        self.id_agrupados_table = QTableView()
        self.id_agrupados_table.setModel(self.id_agrupados_model)
        self.id_agrupados_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.id_agrupados_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.id_agrupados_table.setAlternatingRowColors(True)
        self.id_agrupados_table.setSortingEnabled(True)
        self.id_agrupados_table.setWordWrap(True)
        self.id_agrupados_table.verticalHeader().setDefaultSectionSize(40)
        self.id_agrupados_table.horizontalHeader().setMinimumSectionSize(40)
        self.id_agrupados_table.horizontalHeader().setDefaultSectionSize(180)
        self.id_agrupados_table.horizontalHeader().setMaximumSectionSize(420)
        self.id_agrupados_table.horizontalHeader().setStretchLastSection(True)
        self.id_agrupados_table.horizontalHeader().setSectionsMovable(True)
        self.id_agrupados_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.id_agrupados_table, 1)
        return tab

    def _build_tab_nfe_entrada(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        self.lbl_nfe_entrada_titulo = QLabel("Tabela: nfe_entrada")
        self.lbl_nfe_entrada_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_nfe_entrada_titulo)

        toolbar = QHBoxLayout()
        self.btn_extract_nfe_entrada = QPushButton("Extrair")
        self.btn_refresh_nfe_entrada = QPushButton("Recarregar")
        self.btn_apply_nfe_entrada_filters = QPushButton("Aplicar filtros")
        self.btn_clear_nfe_entrada_filters = QPushButton("Limpar filtros")
        self.nfe_entrada_profile = QComboBox()
        self.nfe_entrada_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_nfe_entrada_profile = QPushButton("Perfil")
        self.btn_nfe_entrada_save_profile = QPushButton("Salvar perfil")
        self.btn_nfe_entrada_colunas = QPushButton("Colunas")
        self.btn_nfe_entrada_destacar = self._criar_botao_destacar()
        self.btn_export_nfe_entrada = QPushButton("Exportar Excel")
        for widget in [
            self.btn_extract_nfe_entrada,
            self.btn_refresh_nfe_entrada,
            self.btn_apply_nfe_entrada_filters,
            self.btn_clear_nfe_entrada_filters,
            self.nfe_entrada_profile,
            self.btn_nfe_entrada_profile,
            self.btn_nfe_entrada_save_profile,
            self.btn_nfe_entrada_colunas,
        ]:
            toolbar.addWidget(widget)
        toolbar.addStretch()
        toolbar.addWidget(self.btn_nfe_entrada_destacar)
        toolbar.addWidget(self.btn_export_nfe_entrada)
        layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.nfe_entrada_filter_id = QComboBox()
        self.nfe_entrada_filter_id.setEditable(True)
        self.nfe_entrada_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.nfe_entrada_filter_id.setMinimumWidth(220)
        self.nfe_entrada_filter_id.lineEdit().setPlaceholderText("Filtrar id_agrupado")
        self.nfe_entrada_filter_desc = QLineEdit()
        self.nfe_entrada_filter_desc.setPlaceholderText("Filtrar descricao")
        self.nfe_entrada_filter_ncm = QLineEdit()
        self.nfe_entrada_filter_ncm.setPlaceholderText("Filtrar NCM")
        self.nfe_entrada_filter_sefin = QLineEdit()
        self.nfe_entrada_filter_sefin.setPlaceholderText("Filtrar co_sefin")
        self.nfe_entrada_filter_texto = QLineEdit()
        self.nfe_entrada_filter_texto.setPlaceholderText("Busca ampla...")
        for widget in [
            self.nfe_entrada_filter_id,
            self.nfe_entrada_filter_desc,
            self.nfe_entrada_filter_ncm,
            self.nfe_entrada_filter_sefin,
            self.nfe_entrada_filter_texto,
        ]:
            filtros.addWidget(widget)
        layout.addLayout(filtros)

        filtros_datas = QHBoxLayout()
        self.nfe_entrada_filter_data_ini = QDateEdit()
        self.nfe_entrada_filter_data_ini.setCalendarPopup(True)
        self.nfe_entrada_filter_data_ini.setDisplayFormat("dd/MM/yyyy")
        self.nfe_entrada_filter_data_ini.setSpecialValueText("Data inicial")
        self.nfe_entrada_filter_data_ini.setMinimumDate(QDate(1900, 1, 1))
        self.nfe_entrada_filter_data_ini.setDate(self.nfe_entrada_filter_data_ini.minimumDate())
        self.nfe_entrada_filter_data_fim = QDateEdit()
        self.nfe_entrada_filter_data_fim.setCalendarPopup(True)
        self.nfe_entrada_filter_data_fim.setDisplayFormat("dd/MM/yyyy")
        self.nfe_entrada_filter_data_fim.setSpecialValueText("Data final")
        self.nfe_entrada_filter_data_fim.setMinimumDate(QDate(1900, 1, 1))
        self.nfe_entrada_filter_data_fim.setDate(self.nfe_entrada_filter_data_fim.minimumDate())
        filtros_datas.addWidget(QLabel("Data inicial"))
        filtros_datas.addWidget(self.nfe_entrada_filter_data_ini)
        filtros_datas.addWidget(QLabel("Data final"))
        filtros_datas.addWidget(self.nfe_entrada_filter_data_fim)
        filtros_datas.addStretch()
        layout.addLayout(filtros_datas)

        self.lbl_nfe_entrada_status = QLabel("Selecione um CNPJ para carregar as NFes de entrada.")
        self.lbl_nfe_entrada_status.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #101827; border: 1px solid #374151; border-radius: 4px; color: #e5e7eb; }"
        )
        layout.addWidget(self.lbl_nfe_entrada_status)

        self.lbl_nfe_entrada_filtros = QLabel("Filtros ativos: nenhum")
        self.lbl_nfe_entrada_filtros.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #dbeafe; background: #0f1b33; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_nfe_entrada_filtros)

        self.nfe_entrada_table = QTableView()
        self.nfe_entrada_table.setModel(self.nfe_entrada_model)
        self.nfe_entrada_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.nfe_entrada_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.nfe_entrada_table.setAlternatingRowColors(True)
        self.nfe_entrada_table.setSortingEnabled(True)
        self.nfe_entrada_table.setWordWrap(True)
        self.nfe_entrada_table.verticalHeader().setDefaultSectionSize(40)
        self.nfe_entrada_table.horizontalHeader().setMinimumSectionSize(40)
        self.nfe_entrada_table.horizontalHeader().setDefaultSectionSize(170)
        self.nfe_entrada_table.horizontalHeader().setMaximumSectionSize(420)
        self.nfe_entrada_table.horizontalHeader().setSectionsMovable(True)
        self.nfe_entrada_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        layout.addWidget(self.nfe_entrada_table, 1)
        return tab

    def _build_tab_aba_anual(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        tab.setStyleSheet(
            """
            QWidget {
                background: #252526;
                color: #f3f4f6;
            }
            QLabel {
                color: #e5e7eb;
            }
            QLineEdit, QComboBox, QDateEdit {
                background: #323438;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 8px;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QPushButton {
                background: #34373d;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background: #40444b;
            }
            QPushButton:pressed {
                background: #2f3338;
            }
            QTableView {
                background: #1f1f1f;
                alternate-background-color: #262626;
                color: #f9fafb;
                gridline-color: #3f3f46;
                border: 1px solid #3f3f46;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QHeaderView::section {
                background: #18181b;
                color: #f9fafb;
                border: 1px solid #3f3f46;
                padding: 6px 8px;
                font-weight: bold;
            }
            QScrollBar:vertical, QScrollBar:horizontal {
                background: #252526;
                border: none;
            }
            QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
                background: #4b5563;
                border-radius: 5px;
            }
            """
        )

        self.lbl_aba_anual_titulo = QLabel("Tabela: aba_anual")
        self.lbl_aba_anual_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_aba_anual_titulo)

        toolbar = QHBoxLayout()
        self.btn_refresh_aba_anual = QPushButton("Recarregar")
        self.btn_refresh_aba_anual.setIcon(QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_BrowserReload))
        self.btn_apply_aba_anual_filters = QPushButton("Aplicar filtros")
        self.btn_clear_aba_anual_filters = QPushButton("Limpar filtros")
        self.btn_filtrar_estoque_anual = QPushButton("Filtrar Estoque (SeleCAo)")
        self.btn_limpar_filtro_cruzado = QPushButton("Limpar Filtro Cruzado")
        self.btn_export_aba_anual = QPushButton("Exportar Excel")
        self.btn_destacar_aba_anual = self._criar_botao_destacar()
        
        toolbar.addWidget(self.btn_refresh_aba_anual)
        toolbar.addWidget(self.btn_apply_aba_anual_filters)
        toolbar.addWidget(self.btn_clear_aba_anual_filters)
        toolbar.addStretch()
        toolbar.addWidget(self.btn_filtrar_estoque_anual)
        toolbar.addWidget(self.btn_limpar_filtro_cruzado)
        toolbar.addWidget(self.btn_destacar_aba_anual)
        toolbar.addWidget(self.btn_export_aba_anual)
        layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.anual_filter_id = QComboBox()
        self.anual_filter_id.setEditable(True)
        self.anual_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.anual_filter_id.setMinimumWidth(220)
        self.anual_filter_id.lineEdit().setPlaceholderText("Filtrar id_agregado")
        self.anual_filter_desc = QLineEdit()
        self.anual_filter_desc.setPlaceholderText("Filtrar descriCAo")
        self.anual_filter_ano = QComboBox()
        self.anual_filter_ano.addItem("Todos")
        self.anual_filter_ano.setMinimumWidth(100)
        self.anual_filter_texto = QLineEdit()
        self.anual_filter_texto.setPlaceholderText("Busca ampla...")

        for widget in [self.anual_filter_id, self.anual_filter_desc, self.anual_filter_ano, self.anual_filter_texto]:
            filtros.addWidget(widget)
        layout.addLayout(filtros)

        filtros_avancados = QHBoxLayout()
        self.anual_filter_num_col = QComboBox()
        self.anual_filter_num_col.addItems(["entradas_desacob", "saidas_desacob", "estoque_final_desacob", "saldo_final", "estoque_final"])
        self.anual_filter_num_min = QLineEdit()
        self.anual_filter_num_min.setPlaceholderText("Min numerico")
        self.anual_filter_num_max = QLineEdit()
        self.anual_filter_num_max.setPlaceholderText("Max numerico")
        self.anual_profile = QComboBox()
        self.anual_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_anual_profile = QPushButton("Perfil")
        self.btn_anual_save_profile = QPushButton("Salvar perfil")
        self.btn_anual_colunas = QPushButton("Colunas")
        for widget in [
            QLabel("Numero"),
            self.anual_filter_num_col,
            self.anual_filter_num_min,
            self.anual_filter_num_max,
            self.anual_profile,
            self.btn_anual_profile,
            self.btn_anual_save_profile,
            self.btn_anual_colunas,
        ]:
            filtros_avancados.addWidget(widget)
        filtros_avancados.addStretch()
        layout.addLayout(filtros_avancados)

        self.lbl_aba_anual_status = QLabel("Selecione um CNPJ para carregar a aba anual.")
        self.lbl_aba_anual_status.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #101827; border: 1px solid #374151; border-radius: 4px; color: #e5e7eb; }"
        )
        layout.addWidget(self.lbl_aba_anual_status)

        self.lbl_aba_anual_filtros = QLabel("Filtros ativos: nenhum")
        self.lbl_aba_anual_filtros.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #dbeafe; background: #0f1b33; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_aba_anual_filtros)

        self.aba_anual_table = QTableView()
        self.aba_anual_table.setModel(self.aba_anual_model)
        self.aba_anual_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.aba_anual_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.aba_anual_table.setAlternatingRowColors(True)
        self.aba_anual_table.setSortingEnabled(True)
        self.aba_anual_table.setWordWrap(True)
        self.aba_anual_table.verticalHeader().setDefaultSectionSize(40)
        self.aba_anual_table.horizontalHeader().setMinimumSectionSize(40)
        self.aba_anual_table.horizontalHeader().setDefaultSectionSize(180)
        self.aba_anual_table.horizontalHeader().setMaximumSectionSize(380)
        self.aba_anual_table.horizontalHeader().setStretchLastSection(True)
        self.aba_anual_table.horizontalHeader().setSectionsMovable(True)
        self.aba_anual_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.aba_anual_table.setStyleSheet(
            "QTableView::item { padding: 4px 2px; }"
            "QTableCornerButton::section { background: #18181b; border: 1px solid #3f3f46; }"
        )
        layout.addWidget(self.aba_anual_table, 1)

        return tab

    def _build_tab_resumo_global(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        tab.setStyleSheet(
            """
            QWidget {
                background: #252526;
                color: #f3f4f6;
            }
            QLabel {
                color: #e5e7eb;
            }
            QPushButton {
                background: #34373d;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background: #40444b;
            }
            QPushButton:pressed {
                background: #2f3338;
            }
            QTableView {
                background: #1f1f1f;
                alternate-background-color: #262626;
                color: #f9fafb;
                gridline-color: #3f3f46;
                border: 1px solid #3f3f46;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QHeaderView::section {
                background: #18181b;
                color: #f9fafb;
                border: 1px solid #3f3f46;
                padding: 6px 8px;
                font-weight: bold;
            }
            """
        )

        self.lbl_resumo_global_titulo = QLabel("Tabela: Resumo Global (Mensal e Anual)")
        self.lbl_resumo_global_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_resumo_global_titulo)

        toolbar = QHBoxLayout()
        self.btn_refresh_resumo_global = QPushButton("Atualizar Resumo Global")
        self.btn_export_resumo_global = QPushButton("Exportar Excel")
        toolbar.addWidget(self.btn_refresh_resumo_global)
        toolbar.addStretch()
        toolbar.addWidget(self.btn_export_resumo_global)
        layout.addLayout(toolbar)

        self.lbl_resumo_global_status = QLabel("Aguardando carregamento da aba mensal e anual...")
        self.lbl_resumo_global_status.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #101827; border: 1px solid #374151; border-radius: 4px; color: #e5e7eb; }"
        )
        layout.addWidget(self.lbl_resumo_global_status)

        self.resumo_global_table = QTableView()
        self.resumo_global_table.setModel(self.resumo_global_model)
        self.resumo_global_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.resumo_global_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.resumo_global_table.setAlternatingRowColors(True)
        self.resumo_global_table.setSortingEnabled(True)
        self.resumo_global_table.setWordWrap(True)
        self.resumo_global_table.verticalHeader().setDefaultSectionSize(40)
        self.resumo_global_table.horizontalHeader().setMinimumSectionSize(80)
        self.resumo_global_table.horizontalHeader().setDefaultSectionSize(180)
        self.resumo_global_table.horizontalHeader().setStretchLastSection(True)
        self.resumo_global_table.setStyleSheet(
            "QTableView::item { padding: 4px 2px; }"
            "QTableCornerButton::section { background: #18181b; border: 1px solid #3f3f46; }"
        )
        layout.addWidget(self.resumo_global_table, 1)

        return tab

    def _build_tab_aba_mensal(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        tab.setStyleSheet(
            """
            QWidget {
                background: #252526;
                color: #f3f4f6;
            }
            QLabel {
                color: #e5e7eb;
            }
            QLineEdit, QComboBox, QDateEdit {
                background: #323438;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 8px;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QPushButton {
                background: #34373d;
                color: #f9fafb;
                border: 1px solid #4b5563;
                border-radius: 4px;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background: #40444b;
            }
            QPushButton:pressed {
                background: #2f3338;
            }
            QTableView {
                background: #1f1f1f;
                alternate-background-color: #262626;
                color: #f9fafb;
                gridline-color: #3f3f46;
                border: 1px solid #3f3f46;
                selection-background-color: #0e639c;
                selection-color: #ffffff;
            }
            QHeaderView::section {
                background: #18181b;
                color: #f9fafb;
                border: 1px solid #3f3f46;
                padding: 6px 8px;
                font-weight: bold;
            }
            """
        )

        self.lbl_aba_mensal_titulo = QLabel("Tabela: aba_mensal")
        self.lbl_aba_mensal_titulo.setStyleSheet(
            "QLabel { font-weight: bold; color: #f8fafc; background: #1f2a44; border: 1px solid #334155; border-radius: 4px; padding: 6px 10px; }"
        )
        layout.addWidget(self.lbl_aba_mensal_titulo)

        toolbar = QHBoxLayout()
        self.btn_refresh_aba_mensal = QPushButton("Recarregar")
        self.btn_refresh_aba_mensal.setIcon(QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_BrowserReload))
        self.btn_apply_aba_mensal_filters = QPushButton("Aplicar filtros")
        self.btn_clear_aba_mensal_filters = QPushButton("Limpar filtros")
        self.btn_export_aba_mensal = QPushButton("Exportar Excel")
        self.btn_destacar_aba_mensal = self._criar_botao_destacar()
        toolbar.addWidget(self.btn_refresh_aba_mensal)
        toolbar.addWidget(self.btn_apply_aba_mensal_filters)
        toolbar.addWidget(self.btn_clear_aba_mensal_filters)
        toolbar.addStretch()
        toolbar.addWidget(self.btn_destacar_aba_mensal)
        toolbar.addWidget(self.btn_export_aba_mensal)
        layout.addLayout(toolbar)

        filtros = QHBoxLayout()
        self.mensal_filter_id = QComboBox()
        self.mensal_filter_id.setEditable(True)
        self.mensal_filter_id.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.mensal_filter_id.setMinimumWidth(220)
        self.mensal_filter_id.lineEdit().setPlaceholderText("Filtrar id_agregado")
        self.mensal_filter_desc = QLineEdit()
        self.mensal_filter_desc.setPlaceholderText("Filtrar descricao")
        self.mensal_filter_ano = QComboBox()
        self.mensal_filter_ano.addItem("Todos")
        self.mensal_filter_ano.setMinimumWidth(100)
        self.mensal_filter_ano.setToolTip("Filtrar por ano")
        self.mensal_filter_mes = QComboBox()
        self.mensal_filter_mes.addItems(["Todos"] + [str(i) for i in range(1, 13)])
        self.mensal_filter_mes.setMinimumWidth(100)
        self.mensal_filter_mes.setToolTip("Filtrar por mes")
        self.mensal_filter_texto = QLineEdit()
        self.mensal_filter_texto.setPlaceholderText("Busca ampla...")
        for widget in [
            self.mensal_filter_id,
            self.mensal_filter_desc,
            QLabel("Ano"),
            self.mensal_filter_ano,
            QLabel("Mes"),
            self.mensal_filter_mes,
            self.mensal_filter_texto,
        ]:
            filtros.addWidget(widget)
        layout.addLayout(filtros)

        filtros_avancados = QHBoxLayout()
        self.mensal_filter_num_col = QComboBox()
        self.mensal_filter_num_col.addItems(["valor_entradas", "qtd_entradas", "pme_mes", "valor_saidas", "qtd_saidas", "pms_mes", "entradas_desacob", "ICMS_entr_desacob", "saldo_mes", "custo_medio_mes", "valor_estoque"])
        self.mensal_filter_num_min = QLineEdit()
        self.mensal_filter_num_min.setPlaceholderText("Min numerico")
        self.mensal_filter_num_max = QLineEdit()
        self.mensal_filter_num_max.setPlaceholderText("Max numerico")
        self.mensal_profile = QComboBox()
        self.mensal_profile.addItems(["Padrao", "Auditoria", "Estoque", "Custos"])
        self.btn_mensal_profile = QPushButton("Perfil")
        self.btn_mensal_save_profile = QPushButton("Salvar perfil")
        self.btn_mensal_colunas = QPushButton("Colunas")
        for widget in [
            QLabel("Numero"),
            self.mensal_filter_num_col,
            self.mensal_filter_num_min,
            self.mensal_filter_num_max,
            self.mensal_profile,
            self.btn_mensal_profile,
            self.btn_mensal_save_profile,
            self.btn_mensal_colunas,
        ]:
            filtros_avancados.addWidget(widget)
        filtros_avancados.addStretch()
        layout.addLayout(filtros_avancados)

        self.lbl_aba_mensal_status = QLabel("Selecione um CNPJ para carregar a aba mensal.")
        self.lbl_aba_mensal_status.setStyleSheet(
            "QLabel { padding: 4px 8px; background: #101827; border: 1px solid #374151; border-radius: 4px; color: #e5e7eb; }"
        )
        layout.addWidget(self.lbl_aba_mensal_status)

        self.lbl_aba_mensal_filtros = QLabel("Filtros ativos: nenhum")
        self.lbl_aba_mensal_filtros.setStyleSheet(
            "QLabel { padding: 4px 8px; color: #dbeafe; background: #0f1b33; border: 1px solid #334155; border-radius: 4px; }"
        )
        layout.addWidget(self.lbl_aba_mensal_filtros)

        self.aba_mensal_table = QTableView()
        self.aba_mensal_table.setModel(self.aba_mensal_model)
        self.aba_mensal_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.aba_mensal_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.aba_mensal_table.setAlternatingRowColors(True)
        self.aba_mensal_table.setSortingEnabled(True)
        self.aba_mensal_table.setWordWrap(True)
        self.aba_mensal_table.verticalHeader().setDefaultSectionSize(40)
        self.aba_mensal_table.horizontalHeader().setMinimumSectionSize(40)
        self.aba_mensal_table.horizontalHeader().setDefaultSectionSize(170)
        self.aba_mensal_table.horizontalHeader().setMaximumSectionSize(380)
        self.aba_mensal_table.horizontalHeader().setStretchLastSection(True)
        self.aba_mensal_table.horizontalHeader().setSectionsMovable(True)
        self.aba_mensal_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.aba_mensal_table.setStyleSheet(
            "QTableView::item { padding: 4px 2px; }"
            "QTableCornerButton::section { background: #18181b; border: 1px solid #3f3f46; }"
        )
        layout.addWidget(self.aba_mensal_table, 1)
        return tab

    def _connect_signals(self) -> None:
        schedule_mov = lambda: self._schedule_debounced("mov_filters", self.aplicar_filtros_mov_estoque)
        schedule_anual = lambda: self._schedule_debounced("anual_filters", self.aplicar_filtros_aba_anual)
        schedule_mensal = lambda: self._schedule_debounced("mensal_filters", self.aplicar_filtros_aba_mensal)
        schedule_nfe_entrada = lambda: self._schedule_debounced("nfe_entrada_filters", self.aplicar_filtros_nfe_entrada)
        schedule_produtos_sel = lambda: self._schedule_debounced("produtos_sel_filters", self.aplicar_filtros_produtos_selecionados)
        schedule_id_agrupados = lambda: self._schedule_debounced("id_agrupados_filters", self.aplicar_filtros_id_agrupados)
        schedule_conv = lambda: self._schedule_debounced("conversao_filters", self.aplicar_filtros_conversao)
        schedule_consulta_quick = lambda: self._schedule_debounced("consulta_quick_filters", self.apply_quick_filters)
        schedule_agregacao_bottom = lambda: self._schedule_debounced("agregacao_bottom_filters", self.apply_aggregation_results_filters)
        schedule_sql_search = lambda: self._schedule_debounced("sql_result_search", self._filter_sql_results)

        self.btn_refresh_cnpjs.clicked.connect(self.refresh_cnpjs)
        self.btn_run_pipeline.clicked.connect(self.run_pipeline_for_input)
        self.btn_extrair_brutas.clicked.connect(self.extrair_tabelas_brutas)
        self.btn_processamento.clicked.connect(self.executar_processamento)
        self.btn_apagar_dados.clicked.connect(self.apagar_dados_cnpj)
        self.btn_apagar_cnpj.clicked.connect(self.apagar_cnpj_completo)
        self.cnpj_list.itemSelectionChanged.connect(self.on_cnpj_selected)
        self.file_tree.itemClicked.connect(self.on_file_activated)
        self.file_tree.itemDoubleClicked.connect(self.on_file_activated)
        self.btn_open_cnpj_folder.clicked.connect(self.open_cnpj_folder)
        self.btn_toggle_panel.toggled.connect(self._toggle_left_panel)
        self.tabs.currentChanged.connect(self._on_main_tab_changed)

        # --- Estoque Tab signals ---
        self.mov_filter_id.currentTextChanged.connect(lambda _value: schedule_mov())
        self.mov_filter_desc.textChanged.connect(lambda _value: schedule_mov())
        self.mov_filter_ncm.textChanged.connect(lambda _value: schedule_mov())
        self.mov_filter_tipo.currentIndexChanged.connect(lambda _index: schedule_mov())
        self.mov_filter_texto.textChanged.connect(lambda _value: schedule_mov())
        self.mov_filter_data_col.currentIndexChanged.connect(lambda _index: schedule_mov())
        self.mov_filter_data_ini.dateChanged.connect(lambda _date: schedule_mov())
        self.mov_filter_data_fim.dateChanged.connect(lambda _date: schedule_mov())
        self.mov_filter_num_col.currentIndexChanged.connect(lambda _index: schedule_mov())
        self.mov_filter_num_min.textChanged.connect(lambda _value: schedule_mov())
        self.mov_filter_num_max.textChanged.connect(lambda _value: schedule_mov())
        self.btn_mov_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("mov_estoque", self.mov_estoque_table, self.mov_estoque_model, self.mov_profile.currentText(), "mov_estoque"))
        self.btn_mov_save_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("mov_estoque", self.mov_estoque_table, self.mov_estoque_model, self.mov_profile, ["Exportar", "Padrao", "Auditoria", "Auditoria Fiscal", "Estoque", "Custos"])
        )
        self.btn_mov_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("mov_estoque", self.mov_estoque_table))
        self.btn_mov_destacar.clicked.connect(lambda: self._destacar_tabela("mov_estoque"))
        self.mov_estoque_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("mov_estoque", self.mov_estoque_table, pos)
        )
        self.mov_estoque_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("mov_estoque", self.mov_estoque_table, self.mov_estoque_model)
        )
        self.mov_estoque_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("mov_estoque", self.mov_estoque_table, self.mov_estoque_model)
        )
        self.mov_estoque_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("mov_estoque", self.mov_estoque_table, self.mov_estoque_model)
        )

        self.btn_export_mov_estoque.clicked.connect(self.exportar_mov_estoque_excel)
        self.btn_refresh_aba_anual.clicked.connect(self.atualizar_aba_anual)
        self.btn_apply_aba_anual_filters.clicked.connect(self.aplicar_filtros_aba_anual)
        self.btn_clear_aba_anual_filters.clicked.connect(self.limpar_filtros_aba_anual)
        self.btn_filtrar_estoque_anual.clicked.connect(self.filtrar_estoque_pela_selecao_anual)
        self.btn_limpar_filtro_cruzado.clicked.connect(self.limpar_filtro_cruzado_anual)
        self.btn_export_aba_anual.clicked.connect(self.exportar_aba_anual_excel)
        self.anual_filter_id.currentTextChanged.connect(lambda _value: schedule_anual())
        self.anual_filter_desc.textChanged.connect(lambda _value: schedule_anual())
        self.anual_filter_ano.currentIndexChanged.connect(lambda _index: schedule_anual())
        self.anual_filter_texto.textChanged.connect(lambda _value: schedule_anual())
        self.anual_filter_num_col.currentIndexChanged.connect(lambda _index: schedule_anual())
        self.anual_filter_num_min.textChanged.connect(lambda _value: schedule_anual())
        self.anual_filter_num_max.textChanged.connect(lambda _value: schedule_anual())
        self.btn_anual_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("aba_anual", self.aba_anual_table, self.aba_anual_model, self.anual_profile.currentText(), "aba_anual"))
        self.btn_anual_save_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("aba_anual", self.aba_anual_table, self.aba_anual_model, self.anual_profile, ["Exportar", "Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_anual_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("aba_anual", self.aba_anual_table))
        self.btn_destacar_aba_anual.clicked.connect(lambda: self._destacar_tabela("aba_anual"))
        self.aba_anual_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("aba_anual", self.aba_anual_table, pos)
        )
        self.aba_anual_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("aba_anual", self.aba_anual_table, self.aba_anual_model)
        )
        self.aba_anual_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("aba_anual", self.aba_anual_table, self.aba_anual_model)
        )
        self.aba_anual_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("aba_anual", self.aba_anual_table, self.aba_anual_model)
        )

        self.btn_refresh_resumo_global.clicked.connect(self.atualizar_aba_resumo_global)
        self.btn_export_resumo_global.clicked.connect(self.exportar_resumo_global_excel)

        self.btn_refresh_aba_mensal.clicked.connect(self.atualizar_aba_mensal)
        self.btn_apply_aba_mensal_filters.clicked.connect(self.aplicar_filtros_aba_mensal)
        self.btn_clear_aba_mensal_filters.clicked.connect(self.limpar_filtros_aba_mensal)
        self.btn_export_aba_mensal.clicked.connect(self.exportar_aba_mensal_excel)
        self.mensal_filter_num_col.currentIndexChanged.connect(lambda _index: schedule_mensal())
        self.mensal_filter_num_min.textChanged.connect(lambda _value: schedule_mensal())
        self.mensal_filter_num_max.textChanged.connect(lambda _value: schedule_mensal())
        self.mensal_filter_id.currentTextChanged.connect(lambda _value: schedule_mensal())
        self.mensal_filter_desc.textChanged.connect(lambda _value: schedule_mensal())
        self.mensal_filter_ano.currentIndexChanged.connect(lambda _index: schedule_mensal())
        self.mensal_filter_mes.currentIndexChanged.connect(lambda _index: schedule_mensal())
        self.mensal_filter_texto.textChanged.connect(lambda _value: schedule_mensal())
        self.btn_mensal_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("aba_mensal", self.aba_mensal_table, self.aba_mensal_model, self.mensal_profile.currentText(), "aba_mensal"))
        self.btn_mensal_save_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("aba_mensal", self.aba_mensal_table, self.aba_mensal_model, self.mensal_profile, ["Exportar", "Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_mensal_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("aba_mensal", self.aba_mensal_table))
        self.btn_destacar_aba_mensal.clicked.connect(lambda: self._destacar_tabela("aba_mensal"))
        self.aba_mensal_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("aba_mensal", self.aba_mensal_table, pos)
        )
        self.aba_mensal_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("aba_mensal", self.aba_mensal_table, self.aba_mensal_model)
        )
        self.aba_mensal_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("aba_mensal", self.aba_mensal_table, self.aba_mensal_model)
        )
        self.aba_mensal_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("aba_mensal", self.aba_mensal_table, self.aba_mensal_model)
        )

        self.btn_extract_nfe_entrada.clicked.connect(self.extrair_dados_nfe_entrada)
        self.btn_refresh_nfe_entrada.clicked.connect(self.atualizar_aba_nfe_entrada)
        self.btn_apply_nfe_entrada_filters.clicked.connect(self.aplicar_filtros_nfe_entrada)
        self.btn_clear_nfe_entrada_filters.clicked.connect(self.limpar_filtros_nfe_entrada)
        self.btn_nfe_entrada_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model, self.nfe_entrada_profile.currentText(), "nfe_entrada"))
        self.btn_nfe_entrada_save_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model, self.nfe_entrada_profile, ["Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_nfe_entrada_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("nfe_entrada", self.nfe_entrada_table))
        self.btn_nfe_entrada_destacar.clicked.connect(lambda: self._destacar_tabela("nfe_entrada"))
        self.btn_export_nfe_entrada.clicked.connect(self.exportar_nfe_entrada_excel)
        self.nfe_entrada_filter_id.currentTextChanged.connect(lambda _value: schedule_nfe_entrada())
        self.nfe_entrada_filter_desc.textChanged.connect(lambda _value: schedule_nfe_entrada())
        self.nfe_entrada_filter_ncm.textChanged.connect(lambda _value: schedule_nfe_entrada())
        self.nfe_entrada_filter_sefin.textChanged.connect(lambda _value: schedule_nfe_entrada())
        self.nfe_entrada_filter_texto.textChanged.connect(lambda _value: schedule_nfe_entrada())
        self.nfe_entrada_filter_data_ini.dateChanged.connect(lambda _date: schedule_nfe_entrada())
        self.nfe_entrada_filter_data_fim.dateChanged.connect(lambda _date: schedule_nfe_entrada())
        self.nfe_entrada_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("nfe_entrada", self.nfe_entrada_table, pos)
        )
        self.nfe_entrada_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model)
        )
        self.nfe_entrada_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model)
        )
        self.nfe_entrada_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model)
        )

        self.btn_refresh_id_agrupados.clicked.connect(self.atualizar_aba_id_agrupados)
        self.btn_apply_id_agrupados_filters.clicked.connect(self.aplicar_filtros_id_agrupados)
        self.btn_clear_id_agrupados_filters.clicked.connect(self.limpar_filtros_id_agrupados)
        self.btn_id_agrupados_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("id_agrupados", self.id_agrupados_table, self.id_agrupados_model, self.id_agrupados_profile.currentText(), "id_agrupados"))
        self.btn_id_agrupados_save_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("id_agrupados", self.id_agrupados_table, self.id_agrupados_model, self.id_agrupados_profile, ["Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_id_agrupados_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("id_agrupados", self.id_agrupados_table))
        self.btn_destacar_id_agrupados.clicked.connect(lambda: self._destacar_tabela("id_agrupados"))
        self.btn_export_id_agrupados.clicked.connect(self.exportar_id_agrupados_excel)
        self.id_agrupados_filter_id.currentTextChanged.connect(lambda _value: schedule_id_agrupados())
        self.id_agrupados_filter_texto.textChanged.connect(lambda _value: schedule_id_agrupados())
        self.id_agrupados_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("id_agrupados", self.id_agrupados_table, pos)
        )
        self.id_agrupados_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("id_agrupados", self.id_agrupados_table, self.id_agrupados_model)
        )
        self.id_agrupados_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("id_agrupados", self.id_agrupados_table, self.id_agrupados_model)
        )
        self.id_agrupados_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("id_agrupados", self.id_agrupados_table, self.id_agrupados_model)
        )

        self.btn_refresh_produtos_sel.clicked.connect(self.atualizar_aba_produtos_selecionados)
        self.btn_apply_produtos_sel_filters.clicked.connect(self.aplicar_filtros_produtos_selecionados)
        self.btn_clear_produtos_sel_filters.clicked.connect(self.limpar_filtros_produtos_selecionados)
        self.btn_produtos_sel_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model, self.produtos_sel_profile.currentText(), "produtos_selecionados"))
        self.btn_produtos_sel_save_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model, self.produtos_sel_profile, ["Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_colunas_produtos_sel.clicked.connect(lambda: self._abrir_menu_colunas_tabela("produtos_selecionados", self.produtos_sel_table))
        self.btn_destacar_produtos_sel.clicked.connect(lambda: self._destacar_tabela("produtos_selecionados"))
        self.btn_export_produtos_sel.clicked.connect(self.exportar_produtos_selecionados_excel)
        self.produtos_sel_filter_id.currentTextChanged.connect(lambda _value: schedule_produtos_sel())
        self.produtos_sel_filter_desc.textChanged.connect(lambda _value: schedule_produtos_sel())
        self.produtos_sel_filter_ano_ini.currentIndexChanged.connect(lambda _index: schedule_produtos_sel())
        self.produtos_sel_filter_ano_fim.currentIndexChanged.connect(lambda _index: schedule_produtos_sel())
        self.produtos_sel_filter_data_ini.dateChanged.connect(lambda _date: schedule_produtos_sel())
        self.produtos_sel_filter_data_fim.dateChanged.connect(lambda _date: schedule_produtos_sel())
        self.produtos_sel_filter_texto.textChanged.connect(lambda _value: schedule_produtos_sel())
        self.produtos_sel_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("produtos_selecionados", self.produtos_sel_table, pos)
        )
        self.produtos_sel_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model)
        )
        self.produtos_sel_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model)
        )
        self.produtos_sel_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model)
        )

        self.btn_add_filter.clicked.connect(self.add_filter_from_form)
        self.btn_clear_filters.clicked.connect(self.clear_filters)
        self.btn_remove_filter.clicked.connect(self.remove_selected_filter)
        self.btn_choose_columns.clicked.connect(self.choose_columns)
        self.btn_apply_consulta_profile.clicked.connect(self._aplicar_perfil_consulta)
        self.btn_save_consulta_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("consulta", self.table_view, self.table_model, self.consulta_profile, ["Padrao", "Auditoria", "Estoque", "Custos"], self._consulta_scope())
        )
        self.btn_consulta_destacar.clicked.connect(lambda: self._destacar_tabela("consulta"))
        self.table_view.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("consulta", self.table_view, pos, scope=self._consulta_scope())
        )
        self.table_view.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("consulta", self.table_view, self.table_model, scope=self._consulta_scope())
        )
        self.table_view.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("consulta", self.table_view, self.table_model, scope=self._consulta_scope())
        )
        self.btn_export_excel_full.clicked.connect(lambda: self.export_excel("full"))
        self.btn_export_excel_filtered.clicked.connect(lambda: self.export_excel("filtered"))
        self.btn_export_excel_visible.clicked.connect(lambda: self.export_excel("visible"))
        self.btn_export_docx.clicked.connect(self.export_docx)
        self.btn_export_html_txt.clicked.connect(self.export_txt_html)

        self.btn_open_editable_table.clicked.connect(self.open_editable_aggregation_table)
        self.btn_execute_aggregation.clicked.connect(self.execute_aggregation)
        self.btn_reprocessar_agregacao.clicked.connect(self.reprocessar_agregacao)
        self.btn_clear_top_agg_filters.clicked.connect(self.clear_top_aggregation_filters)
        self.btn_clear_bottom_agg_filters.clicked.connect(self.clear_bottom_aggregation_filters)
        self.btn_top_match_ncm_cest.clicked.connect(lambda: self._aplicar_filtro_relacional_agregacao("top", include_gtin=False))
        self.btn_top_match_ncm_cest_gtin.clicked.connect(lambda: self._aplicar_filtro_relacional_agregacao("top", include_gtin=True))
        self.btn_apply_top_profile.clicked.connect(lambda: self._aplicar_perfil_agregacao("agregacao_top", self.aggregation_table, self.aggregation_table_model, self.top_profile.currentText()))
        self.btn_save_top_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("agregacao_top", self.aggregation_table, self.aggregation_table_model, self.top_profile, ["Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_top_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("agregacao_top", self.aggregation_table))
        self.btn_top_destacar.clicked.connect(lambda: self._destacar_tabela("agregacao_top"))
        self.aggregation_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("agregacao_top", self.aggregation_table, pos)
        )
        self.aggregation_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("agregacao_top", self.aggregation_table, self.aggregation_table_model)
        )
        self.aggregation_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("agregacao_top", self.aggregation_table, self.aggregation_table_model)
        )
        self.aggregation_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda *_: self._salvar_preferencias_tabela("agregacao_top", self.aggregation_table, self.aggregation_table_model)
        )
        self.btn_apply_bottom_profile.clicked.connect(lambda: self._aplicar_perfil_agregacao("agregacao_bottom", self.results_table, self.results_table_model, self.bottom_profile.currentText()))
        self.btn_save_bottom_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("agregacao_bottom", self.results_table, self.results_table_model, self.bottom_profile, ["Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_bottom_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("agregacao_bottom", self.results_table))
        self.results_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("agregacao_bottom", self.results_table, pos)
        )
        self.results_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("agregacao_bottom", self.results_table, self.results_table_model)
        )
        self.results_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("agregacao_bottom", self.results_table, self.results_table_model)
        )
        self.results_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda *_: self._salvar_preferencias_tabela("agregacao_bottom", self.results_table, self.results_table_model)
        )
        self.btn_bottom_match_ncm_cest.clicked.connect(lambda: self._aplicar_filtro_relacional_agregacao("bottom", include_gtin=False))
        self.btn_bottom_match_ncm_cest_gtin.clicked.connect(lambda: self._aplicar_filtro_relacional_agregacao("bottom", include_gtin=True))
        self.btn_bottom_destacar.clicked.connect(lambda: self._destacar_tabela("agregacao_bottom"))
        for tabela, contexto in [
            (self.aggregation_table, "agregacao_top"),
            (self.results_table, "agregacao_bottom"),
            (self.sql_result_table, "sql_result"),
            (self.conversion_table, "conversao"),
            (self.aba_mensal_table, "aba_mensal"),
            (self.aba_anual_table, "aba_anual"),
            (self.nfe_entrada_table, "nfe_entrada"),
            (self.produtos_sel_table, "produtos_selecionados"),
            (self.id_agrupados_table, "id_agrupados"),
        ]:
            tabela.setContextMenuPolicy(Qt.CustomContextMenu)
            tabela.customContextMenuRequested.connect(
                lambda pos, t=tabela, ctx=contexto: self._abrir_menu_contexto_celula(ctx, t, pos)
            )

        for qf in [self.qf_norm, self.qf_desc, self.qf_ncm, self.qf_cest,
                   self.aqf_norm, self.aqf_desc, self.aqf_ncm, self.aqf_cest]:
            qf.returnPressed.connect(self.apply_quick_filters)
            qf.textChanged.connect(lambda _value: schedule_consulta_quick())
        for qf in [self.bqf_norm, self.bqf_desc, self.bqf_ncm, self.bqf_cest]:
            qf.returnPressed.connect(self.apply_aggregation_results_filters)
            qf.textChanged.connect(lambda _value: schedule_agregacao_bottom())

        # --- Consulta SQL tab ---
        self.sql_combo.currentIndexChanged.connect(self._on_sql_selected)
        self.btn_sql_execute.clicked.connect(self._execute_sql_query)
        self.btn_sql_export.clicked.connect(self._export_sql_results)
        self.btn_sql_destacar.clicked.connect(lambda: self._destacar_tabela("sql_result"))
        self.sql_result_search.returnPressed.connect(self._filter_sql_results)
        self.sql_result_search.textChanged.connect(lambda _value: schedule_sql_search())
        # --- Conversao tab ---
        self.btn_refresh_conversao.clicked.connect(self.atualizar_aba_conversao)
        self.chk_show_single_unit.stateChanged.connect(lambda _state: self.atualizar_aba_conversao())
        self.btn_export_conversao.clicked.connect(self.exportar_conversao_excel)
        self.btn_import_conversao.clicked.connect(self.importar_conversao_excel)
        self.btn_conversao_destacar.clicked.connect(lambda: self._destacar_tabela("conversao"))
        self.btn_recalcular_fatores.clicked.connect(lambda: self.recalcular_derivados_conversao())
        self.btn_apply_conversao_profile.clicked.connect(lambda: self._aplicar_perfil_tabela("conversao", self.conversion_table, self.conversion_model, self.conversao_profile.currentText(), "conversao"))
        self.btn_save_conversao_profile.clicked.connect(
            lambda: self._salvar_perfil_tabela_com_dialogo("conversao", self.conversion_table, self.conversion_model, self.conversao_profile, ["Padrao", "Auditoria", "Estoque", "Custos"])
        )
        self.btn_conversao_colunas.clicked.connect(lambda: self._abrir_menu_colunas_tabela("conversao", self.conversion_table))
        self.conversion_table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos: self._abrir_menu_colunas_tabela("conversao", self.conversion_table, pos)
        )
        self.conversion_table.horizontalHeader().sectionMoved.connect(
            lambda *_: self._salvar_preferencias_tabela("conversao", self.conversion_table, self.conversion_model)
        )
        self.conversion_table.horizontalHeader().sectionResized.connect(
            lambda *_: self._salvar_preferencias_tabela("conversao", self.conversion_table, self.conversion_model)
        )
        self.conversion_table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order: self._salvar_preferencias_tabela("conversao", self.conversion_table, self.conversion_model)
        )
        self.conv_filter_id.currentTextChanged.connect(lambda _value: schedule_conv())
        self.conv_filter_desc.textChanged.connect(lambda _value: schedule_conv())
        self.conversion_model.dataChanged.connect(self._on_conversion_model_changed)
        
        self.conversion_table.selectionModel().selectionChanged.connect(self._on_conversion_selection_changed)
        self.btn_apply_unid_ref.clicked.connect(self._apply_unid_ref_to_all)

    def _abrir_fio_de_ouro(self, id_agrupado: str) -> None:
        if not self.state.current_cnpj:
            return
            
        pasta_analises = CNPJ_ROOT / self.state.current_cnpj / "analises" / "produtos"
        arquivos = list(pasta_analises.glob(f"*_enriquecido_{self.state.current_cnpj}.parquet"))
        dfs = []
        filtro_id = [FilterCondition(column="id_agrupado", operator="igual", value=id_agrupado)]
        for arq in arquivos:
            try:
                schema = self.parquet_service.get_schema(arq)
                if "id_agrupado" not in schema:
                    continue
                df = self.parquet_service.load_dataset(arq, filtro_id)
                if not df.is_empty():
                    df = df.with_columns(pl.lit(arq.name.split("_enriquecido")[0].upper()).alias("origem_fio_ouro"))
                    dfs.append(df)
            except Exception:
                pass
                
        if not dfs:
            self.show_info("Fio de Ouro", f"Nenhum registro enriquecido encontrado para: {id_agrupado}.")
            return
            
        try:
            df_final = pl.concat(dfs, how="diagonal_relaxed")
            from interface_grafica.ui.dialogs import DialogoFioDeOuro
            dlg = DialogoFioDeOuro(df_final, self)
            dlg.exec()
        except Exception as e:
            self.show_error("Fio de Ouro", f"Erro ao gerar trilha de auditoria: {e}")

    def _copiar_valor_celula(self, table: QTableView, index) -> None:
        if not index or not index.isValid():
            return
        valor = index.data(Qt.DisplayRole)
        QGuiApplication.clipboard().setText("" if valor is None else str(valor))

    def _abrir_menu_contexto_celula(self, contexto: str, table: QTableView, pos) -> None:
        index = table.indexAt(pos)
        if not index.isValid():
            return

        menu = QMenu(self)
        acao_copiar = menu.addAction("Copiar valor")
        acao_copiar.triggered.connect(lambda: self._copiar_valor_celula(table, index))

        model = table.model()
        if (
            contexto == "mov_estoque"
            and isinstance(model, PolarsTableModel)
            and not model.get_dataframe().is_empty()
            and "id_agrupado" in model.get_dataframe().columns
        ):
            try:
                id_agrupado = model.get_dataframe()["id_agrupado"][index.row()]
            except Exception:
                id_agrupado = None
            if id_agrupado:
                menu.addSeparator()
                acao = menu.addAction(f"Auditoria 'Fio de Ouro' ({id_agrupado})")
                acao.triggered.connect(lambda: self._abrir_fio_de_ouro(id_agrupado))

        menu.exec(table.viewport().mapToGlobal(pos))

    def show_error(self, title: str, message: str) -> None:
        QMessageBox.critical(self, title, message)

    def show_info(self, title: str, message: str) -> None:
        QMessageBox.information(self, title, message)

    def _setup_copy_shortcut(self) -> None:
        self.shortcut_copy = QShortcut(QKeySequence.StandardKey.Copy, self)
        self.shortcut_copy.activated.connect(self._copy_selection_from_active_table)

    def _copy_selection_from_active_table(self) -> None:
        tables = [
            self.table_view,
            self.aggregation_table_view,
            self.results_table_view,
            self.sql_result_table,
            self.conversion_table,
            self.mov_estoque_table,
            self.aba_mensal_table,
            self.aba_anual_table,
            self.nfe_entrada_table,
            self.id_agrupados_table,
            self.produtos_sel_table,
        ]
        tables.extend(janela.table for janela in self._detached_windows.values() if janela is not None)
        active_table = next((t for t in tables if t and t.hasFocus()), None)
        if active_table is None:
            return

        selected_indexes = active_table.selectedIndexes()
        if not selected_indexes:
            return

        selected_indexes = sorted(selected_indexes, key=lambda i: (i.row(), i.column()))
        row_min = min(i.row() for i in selected_indexes)
        row_max = max(i.row() for i in selected_indexes)
        col_min = min(i.column() for i in selected_indexes)
        col_max = max(i.column() for i in selected_indexes)
        selected_map = {(i.row(), i.column()): i for i in selected_indexes}

        lines: list[str] = []
        for r in range(row_min, row_max + 1):
            vals: list[str] = []
            for c in range(col_min, col_max + 1):
                idx = selected_map.get((r, c))
                vals.append(str(idx.data() if idx is not None else ""))
            lines.append("\t".join(vals))

        QGuiApplication.clipboard().setText("\n".join(lines))

    def _detached_title(self, contexto: str) -> str:
        cnpj = self.state.current_cnpj or "sem CNPJ"
        mapa = {
            "consulta": f"Consulta - {cnpj}",
            "sql_result": f"Consulta SQL - {cnpj}",
            "agregacao_top": f"Agregacao - Tabela Superior - {cnpj}",
            "agregacao_bottom": f"Agregacao - Tabela Inferior - {cnpj}",
            "conversao": f"Conversao - {cnpj}",
            "mov_estoque": f"Movimentacao de Estoque - {cnpj}",
            "aba_mensal": f"Tabela Mensal - {cnpj}",
            "aba_anual": f"Tabela Anual - {cnpj}",
            "nfe_entrada": f"NFe Entrada - {cnpj}",
            "id_agrupados": f"id_agrupados - {cnpj}",
            "produtos_selecionados": f"Produtos Selecionados - {cnpj}",
        }
        return mapa.get(contexto, f"Tabela Destacada - {cnpj}")

    def _detached_assets(self, contexto: str) -> tuple[QTableView | None, PolarsTableModel | None]:
        mapa = {
            "consulta": (self.table_view, self.table_model),
            "sql_result": (self.sql_result_table, self.sql_result_model),
            "agregacao_top": (self.aggregation_table, self.aggregation_table_model),
            "agregacao_bottom": (self.results_table, self.results_table_model),
            "conversao": (self.conversion_table, self.conversion_model),
            "mov_estoque": (self.mov_estoque_table, self.mov_estoque_model),
            "aba_mensal": (self.aba_mensal_table, self.aba_mensal_model),
            "aba_anual": (self.aba_anual_table, self.aba_anual_model),
            "nfe_entrada": (self.nfe_entrada_table, self.nfe_entrada_model),
            "id_agrupados": (self.id_agrupados_table, self.id_agrupados_model),
            "produtos_selecionados": (self.produtos_sel_table, self.produtos_selecionados_model),
        }
        return mapa.get(contexto, (None, None))

    def _detached_scope(self, contexto: str) -> str | None:
        if contexto == "consulta":
            return self._consulta_scope()
        return None

    def _marcar_recalculo_conversao_pendente(self, motivo: str | None = None) -> None:
        self._conversion_recalc_pending = True
        if hasattr(self, "btn_recalcular_fatores"):
            self.btn_recalcular_fatores.setEnabled(True)
        mensagem = "Alteracoes em fatores salvas. Recalculo pendente."
        if motivo:
            mensagem += f" {motivo}"
        self.status.showMessage(mensagem)

    def _limpar_recalculo_conversao_pendente(self) -> None:
        self._conversion_recalc_pending = False
        if hasattr(self, "btn_recalcular_fatores"):
            self.btn_recalcular_fatores.setEnabled(False)

    def _on_main_tab_changed(self, current_index: int) -> None:
        if not hasattr(self, "tab_conversao"):
            return
        idx_conversao = self.tabs.indexOf(self.tab_conversao)
        if idx_conversao < 0:
            return
        if current_index != idx_conversao and self._conversion_recalc_pending:
            self.recalcular_derivados_conversao(show_popup=False)

    def _on_detached_window_closed(self, contexto: str) -> None:
        self._detached_windows.pop(contexto, None)

    def _destacar_tabela(self, contexto: str) -> None:
        table, source_model = self._detached_assets(contexto)
        if table is None or source_model is None:
            self.show_error("Tabela indisponivel", "Nao foi possivel localizar a tabela para destacar.")
            return
        if source_model.dataframe.is_empty():
            self.show_error("Tabela vazia", "Nao ha dados carregados nessa tabela para destacar.")
            return

        janela_existente = self._detached_windows.get(contexto)
        if janela_existente is not None:
            janela_existente.show()
            janela_existente.raise_()
            janela_existente.activateWindow()
            return

        janela = DetachedTableWindow(self._detached_title(contexto), contexto, source_model, self)
        self._atualizar_combo_perfis_tabela(
            janela.profile_combo,
            contexto,
            ["Padrao", "Auditoria", "Estoque", "Custos"],
            scope=self._detached_scope(contexto),
        )
        janela.btn_apply_profile.clicked.connect(
            lambda _checked=False, ctx=contexto, t=janela.table, m=janela.table_model, combo=janela.profile_combo:
            self._aplicar_perfil_tabela(ctx, t, m, combo.currentText(), ctx, scope=self._detached_scope(ctx))
        )
        janela.btn_save_profile.clicked.connect(
            lambda _checked=False, ctx=contexto, t=janela.table, m=janela.table_model, combo=janela.profile_combo:
            self._salvar_perfil_tabela_com_dialogo(ctx, t, m, combo, ["Padrao", "Auditoria", "Estoque", "Custos"], scope=self._detached_scope(ctx))
        )
        janela.btn_columns.clicked.connect(
            lambda _checked=False, ctx=contexto, t=janela.table:
            self._abrir_menu_colunas_tabela(ctx, t, scope=self._detached_scope(ctx))
        )
        janela.closed.connect(self._on_detached_window_closed)
        janela.table.customContextMenuRequested.connect(
            lambda pos, t=janela.table, ctx=contexto: self._abrir_menu_contexto_celula(ctx, t, pos)
        )
        janela.table.horizontalHeader().customContextMenuRequested.connect(
            lambda pos, t=janela.table, m=janela.table_model, ctx=contexto: self._abrir_menu_colunas_tabela(ctx, t, pos, scope=self._detached_scope(ctx))
        )
        janela.table.horizontalHeader().sectionMoved.connect(
            lambda *_args, t=janela.table, m=janela.table_model, ctx=contexto: self._salvar_preferencias_tabela(ctx, t, m, scope=self._detached_scope(ctx))
        )
        janela.table.horizontalHeader().sectionResized.connect(
            lambda *_args, t=janela.table, m=janela.table_model, ctx=contexto: self._salvar_preferencias_tabela(ctx, t, m, scope=self._detached_scope(ctx))
        )
        janela.table.horizontalHeader().sortIndicatorChanged.connect(
            lambda _index, _order, t=janela.table, m=janela.table_model, ctx=contexto: self._salvar_preferencias_tabela(ctx, t, m, scope=self._detached_scope(ctx))
        )
        self._aplicar_preferencias_tabela(contexto, janela.table, janela.table_model, scope=self._detached_scope(contexto))
        janela.show()
        self._detached_windows[contexto] = janela

    def _destacar_tabela_estoque(self, contexto: str) -> None:
        self._destacar_tabela(contexto)

    def refresh_cnpjs(self) -> None:
        known = {record.cnpj for record in self.registry_service.list_records()}
        known.update(self.parquet_service.list_cnpjs())
        current = self.state.current_cnpj
        self.cnpj_list.clear()
        for cnpj in sorted(known):
            self.cnpj_list.addItem(cnpj)
        if current:
            matches = self.cnpj_list.findItems(current, Qt.MatchExactly)
            if matches:
                self.cnpj_list.setCurrentItem(matches[0])

    def run_pipeline_for_input(self) -> None:
        try:
            cnpj = self.servico_pipeline_funcoes.servico_extracao.sanitizar_cnpj(self.cnpj_input.text())
        except Exception as exc:
            self.show_error("CPF/CNPJ invalido", str(exc))
            return

        # 1. Selecionar Consultas SQL
        consultas_disp = self.servico_pipeline_funcoes.servico_extracao.listar_consultas()
        if not consultas_disp:
            self.show_error("Sem consultas", "Nenhum arquivo .sql encontrado na pasta sql/")
            return
            
        pre_sql = self.selection_service.get_selections("ultimas_consultas")
        dlg_sql = DialogoSelecaoConsultas(consultas_disp, self, pre_selecionados=pre_sql)
        if not dlg_sql.exec():
            return
        sql_selecionados = dlg_sql.consultas_selecionadas()
        self.selection_service.set_selections("ultimas_consultas", sql_selecionados)

        # 2. Selecionar Tabelas
        tabelas_disp = self.servico_pipeline_funcoes.servico_tabelas.listar_tabelas()
        pre_tabs = self.selection_service.get_selections("ultimas_tabelas")
        dlg_tab = DialogoSelecaoTabelas(tabelas_disp, self, pre_selecionados=pre_tabs)
        if not dlg_tab.exec():
            return
        tabelas_selecionadas = dlg_tab.tabelas_selecionadas()
        self.selection_service.set_selections("ultimas_tabelas", tabelas_selecionadas)

        if not sql_selecionados and not tabelas_selecionadas:
            return

        self.btn_run_pipeline.setEnabled(False)
        self.status.showMessage(f"Executando pipeline para {cnpj}...")
        
        data_limite = self.date_input.date().toString("dd/MM/yyyy")
        self.pipeline_worker = PipelineWorker(
            self.servico_pipeline_funcoes, 
            cnpj, 
            sql_selecionados, 
            tabelas_selecionadas, 
            data_limite
        )
        self.pipeline_worker.finished_ok.connect(self.on_pipeline_finished)
        self.pipeline_worker.failed.connect(self.on_pipeline_failed)
        self.pipeline_worker.progress.connect(self.status.showMessage)
        self._registrar_limpeza_worker("pipeline_worker", self.pipeline_worker)
        self.pipeline_worker.start()

    def on_pipeline_finished(self, result: ResultadoPipeline) -> None:
        self.btn_run_pipeline.setEnabled(True)
        self.registry_service.upsert(result.cnpj, ran_now=True)
        self.status.showMessage(f"Pipeline concluido para {result.cnpj}.")
        self.refresh_cnpjs()
        matches = self.cnpj_list.findItems(result.cnpj, Qt.MatchExactly)
        if matches:
            self.cnpj_list.setCurrentItem(matches[0])
            self.refresh_file_tree(result.cnpj)
            self.atualizar_aba_conversao()
            
        msg = "\n".join(result.mensagens[-10:]) if result.mensagens else "Processado com sucesso."
        self.show_info("Pipeline concluido", f"CNPJ {result.cnpj} processado.\n\nUltimas mensagens:\n{msg}")

    def on_pipeline_failed(self, message: str) -> None:
        self.btn_run_pipeline.setEnabled(True)
        self.status.showMessage("Falha na execucao do pipeline.")
        self.show_error("Falha nao pipeline", message)

    # ------------------------------------------------------------------
    # BotAes individuais: Extrair Brutas, Processamento, Apagar
    # ------------------------------------------------------------------
    def _obter_cnpj_valido(self) -> str | None:
        """Obtem CPF/CNPJ valido da input box ou da selecao da lista."""
        texto = self.cnpj_input.text().strip()
        if not texto:
            item = self.cnpj_list.currentItem()
            if item:
                texto = item.text()
        if not texto:
            self.show_error("CPF/CNPJ nao informado", "Digite ou selecione um CPF/CNPJ.")
            return None
        try:
            return self.servico_pipeline_funcoes.servico_extracao.sanitizar_cnpj(texto)
        except Exception as exc:
            self.show_error("CPF/CNPJ invalido", str(exc))
            return None

    def extrair_tabelas_brutas(self) -> None:
        """Executa apenas a extracao SQL (fase 1 do pipeline)."""
        cnpj = self._obter_cnpj_valido()
        if not cnpj:
            return

        consultas_disp = self.servico_pipeline_funcoes.servico_extracao.listar_consultas()
        if not consultas_disp:
            self.show_error("Sem consultas", "Nenhum arquivo .sql encontrado na pasta sql/")
            return

        pre_sql = self.selection_service.get_selections("ultimas_consultas")
        dlg_sql = DialogoSelecaoConsultas(consultas_disp, self, pre_selecionados=pre_sql)
        if not dlg_sql.exec():
            return
        sql_selecionados = dlg_sql.consultas_selecionadas()
        self.selection_service.set_selections("ultimas_consultas", sql_selecionados)

        self.btn_extrair_brutas.setEnabled(False)
        self.status.showMessage(f"Extraindo tabelas brutas para {cnpj}...")

        data_limite = self.date_input.date().toString("dd/MM/yyyy")
        self.pipeline_worker = PipelineWorker(
            self.servico_pipeline_funcoes,
            cnpj,
            sql_selecionados,
            [],  # sem tabelas a apenas extracao
            data_limite,
        )
        self.pipeline_worker.finished_ok.connect(self._on_extracao_finished)
        self.pipeline_worker.failed.connect(self._on_extracao_failed)
        self.pipeline_worker.progress.connect(self.status.showMessage)
        self._registrar_limpeza_worker("pipeline_worker", self.pipeline_worker)
        self.pipeline_worker.start()

    def _on_extracao_finished(self, result: ResultadoPipeline) -> None:
        self.btn_extrair_brutas.setEnabled(True)
        self._atualizar_estado_botao_nfe_entrada()
        self.status.showMessage(f"Extracao concluida para {result.cnpj}.")
        self.refresh_cnpjs()
        matches = self.cnpj_list.findItems(result.cnpj, Qt.MatchExactly)
        if matches:
            self.cnpj_list.setCurrentItem(matches[0])
            self.refresh_file_tree(result.cnpj)
        msg = "\n".join(result.mensagens[-10:]) if result.mensagens else "Extracao concluida."
        self.show_info("Extracao concluida", f"CNPJ {result.cnpj}.\n\n{msg}")

    def _on_extracao_failed(self, message: str) -> None:
        self.btn_extrair_brutas.setEnabled(True)
        self._atualizar_estado_botao_nfe_entrada()
        self.status.showMessage("Falha na extracao.")
        self.show_error("Falha na extracao", message)

    def extrair_dados_nfe_entrada(self) -> None:
        cnpj = self._obter_cnpj_valido()
        if not cnpj:
            return
        if self.pipeline_worker is not None and self.pipeline_worker.isRunning():
            self.show_error("Aguarde", "Ja existe uma extracao/processamento em execucao.")
            return
        if self.service_worker is not None and self.service_worker.isRunning():
            self.show_error("Aguarde", "Ja existe um processamento pesado em execucao.")
            return
        if self.query_worker is not None and self.query_worker.isRunning():
            self.show_error("Aguarde", "Ja existe uma consulta SQL em execucao.")
            return

        consultas_disp = self.servico_pipeline_funcoes.servico_extracao.listar_consultas()
        sql_nfe = next((sql_id for sql_id in consultas_disp if sql_id.lower().endswith("/nfe.sql")), None)
        sql_nfce = next((sql_id for sql_id in consultas_disp if sql_id.lower().endswith("/nfce.sql")), None)
        consultas_nfe_entrada = [p for p in [sql_nfe, sql_nfce] if p is not None]
        if not consultas_nfe_entrada:
            self.show_error("SQL nao encontrada", "Nao foi possivel localizar as consultas NFe.sql/NFCe.sql na pasta sql/.")
            return

        tabelas_necessarias = [
            "item_unidades",
            "itens",
            "descricao_produtos",
            "produtos_final",
            "fontes_produtos",
        ]
        self.btn_extract_nfe_entrada.setEnabled(False)
        self.status.showMessage(f"Extraindo dados da NFe Entrada para {cnpj}...")
        data_limite = self.date_input.date().toString("dd/MM/yyyy")
        self.pipeline_worker = PipelineWorker(
            self.servico_pipeline_funcoes,
            cnpj,
            consultas_nfe_entrada,
            tabelas_necessarias,
            data_limite,
        )
        self.pipeline_worker.finished_ok.connect(self._on_nfe_entrada_extract_finished)
        self.pipeline_worker.failed.connect(self._on_nfe_entrada_extract_failed)
        self.pipeline_worker.progress.connect(self.status.showMessage)
        self._registrar_limpeza_worker("pipeline_worker", self.pipeline_worker)
        self.pipeline_worker.start()

    def _on_nfe_entrada_extract_finished(self, result: ResultadoPipeline) -> None:
        self.btn_extract_nfe_entrada.setEnabled(True)
        self._atualizar_estado_botao_nfe_entrada()
        self.registry_service.upsert(result.cnpj, ran_now=True)
        self.status.showMessage(f"Extracao da NFe Entrada concluida para {result.cnpj}.")
        self.refresh_cnpjs()
        matches = self.cnpj_list.findItems(result.cnpj, Qt.MatchExactly)
        if matches:
            self.cnpj_list.setCurrentItem(matches[0])
            self.refresh_file_tree(result.cnpj)
        self.atualizar_aba_nfe_entrada()
        self.atualizar_aba_id_agrupados()
        msg = "\n".join(result.mensagens[-10:]) if result.mensagens else "Dados da NFe Entrada preparados com sucesso."
        self.show_info("NFe Entrada concluida", f"CNPJ {result.cnpj}.\n\n{msg}")

    def _on_nfe_entrada_extract_failed(self, message: str) -> None:
        self.btn_extract_nfe_entrada.setEnabled(True)
        self._atualizar_estado_botao_nfe_entrada()
        self.status.showMessage("Falha na extracao da NFe Entrada.")
        self.show_error("Falha na NFe Entrada", message)

    def executar_processamento(self) -> None:
        """Executa apenas a geracao de tabelas (fase 2 do pipeline)."""
        cnpj = self._obter_cnpj_valido()
        if not cnpj:
            return

        tabelas_disp = self.servico_pipeline_funcoes.servico_tabelas.listar_tabelas()
        pre_tabs = self.selection_service.get_selections("ultimas_tabelas")
        dlg_tab = DialogoSelecaoTabelas(tabelas_disp, self, pre_selecionados=pre_tabs)
        if not dlg_tab.exec():
            return
        tabelas_selecionadas = dlg_tab.tabelas_selecionadas()
        self.selection_service.set_selections("ultimas_tabelas", tabelas_selecionadas)

        self.btn_processamento.setEnabled(False)
        self.status.showMessage(f"Gerando tabelas para {cnpj}...")

        self.pipeline_worker = PipelineWorker(
            self.servico_pipeline_funcoes,
            cnpj,
            [],  # sem consultas SQL a apenas processamento
            tabelas_selecionadas,
            None,
        )
        self.pipeline_worker.finished_ok.connect(self._on_processamento_finished)
        self.pipeline_worker.failed.connect(self._on_processamento_failed)
        self.pipeline_worker.progress.connect(self.status.showMessage)
        self._registrar_limpeza_worker("pipeline_worker", self.pipeline_worker)
        self.pipeline_worker.start()

    def _on_processamento_finished(self, result: ResultadoPipeline) -> None:
        self.btn_processamento.setEnabled(True)
        self._atualizar_estado_botao_nfe_entrada()
        self.status.showMessage(f"Processamento concluido para {result.cnpj}.")
        self.refresh_cnpjs()
        matches = self.cnpj_list.findItems(result.cnpj, Qt.MatchExactly)
        if matches:
            self.cnpj_list.setCurrentItem(matches[0])
            self.refresh_file_tree(result.cnpj)
            self.atualizar_aba_conversao()
            self.atualizar_aba_id_agrupados()
        msg = "\n".join(result.mensagens[-10:]) if result.mensagens else "Processamento concluido."
        self.show_info("Processamento concluido", f"CNPJ {result.cnpj}.\n\n{msg}")

    def _on_processamento_failed(self, message: str) -> None:
        self.btn_processamento.setEnabled(True)
        self._atualizar_estado_botao_nfe_entrada()
        self.status.showMessage("Falha nao processamento.")
        self.show_error("Falha nao processamento", message)

    def apagar_dados_cnpj(self) -> None:
        """Apaga analises/ e arquivos_parquet/ do CNPJ selecionado (mantem pasta raiz)."""
        cnpj = self._obter_cnpj_valido()
        if not cnpj:
            return

        resp = QMessageBox.warning(
            self,
            "Apagar dados",
            f"Deseja apagar todos os dados (parquets e analises) do CNPJ {cnpj}?\n\nEsta acao nao pode ser desfeita.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        ok = self.servico_pipeline_funcoes.servico_extracao.apagar_dados_cnpj(cnpj)
        if ok:
            self.show_info("Dados apagados", f"Os dados do CNPJ {cnpj} foram removidos.")
            self.refresh_file_tree(cnpj)

    def apagar_cnpj_completo(self) -> None:
        """Remove a pasta inteira do CNPJ do filesystem e do registro SQL."""
        cnpj = self._obter_cnpj_valido()
        if not cnpj:
            return

        ret = QMessageBox.critical(
            self,
            "PERIGO: Apagar CNPJ",
            f"Isso removera permanentemente TODA a pasta do CNPJ {cnpj} e seus registros no banco.\n\nTem certeza absoluta?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return

        try:
            self.servico_pipeline_funcoes.servico_extracao.apagar_cnpj_total(cnpj)
            self.registry_service.delete_by_cnpj(cnpj)
            self.show_info("Removido", f"CNPJ {cnpj} removido com sucesso.")
            self.refresh_cnpjs()
            self.file_tree.clear()
        except Exception as e:
            self.show_error("Erro ao apagar", str(e))

    def _toggle_left_panel(self, checked: bool) -> None:
        if checked:
            self.left_panel_widget.hide()
            self.btn_toggle_panel.setText(">> Mostrar Painel Lateral")
        else:
            self.left_panel_widget.show()
            self.btn_toggle_panel.setText("<< Ocultar Painel Lateral")

    def on_cnpj_selected(self) -> None:
        item = self.cnpj_list.currentItem()
        if not item:
            return
        cnpj = item.text()
        self.state.current_cnpj = cnpj
        self._reset_table_resize_flag("conversao")
        self._reset_table_resize_flag("mov_estoque")
        self._reset_table_resize_flag("aba_mensal")
        self._reset_table_resize_flag("aba_anual")
        self._reset_table_resize_flag("nfe_entrada")
        self._reset_table_resize_flag("produtos_selecionados")
        self._reset_table_resize_flag("id_agrupados")
        self._reset_table_resize_flag("agregacao_top")
        self._reset_table_resize_flag("agregacao_bottom")
        self.status.showMessage(f"CNPJ selecionado: {cnpj}")
        self._refresh_profile_combos()
        self.refresh_file_tree(cnpj)
        self.atualizar_aba_conversao()
        self.atualizar_aba_mov_estoque()
        self.atualizar_aba_mensal()
        self.atualizar_aba_anual()
        self.atualizar_aba_nfe_entrada()
        self.atualizar_aba_id_agrupados()

    def atualizar_aba_mov_estoque(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            self._atualizar_titulo_aba_mov_estoque()
            return

        path = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"mov_estoque_{cnpj}.parquet"
        if not path.exists():
            self.mov_estoque_model.set_dataframe(pl.DataFrame())
            self._mov_estoque_df = pl.DataFrame()
            self.lbl_mov_estoque_status.setText("Arquivo 'mov_estoque' nao encontrado para este CNPJ.")
            self._atualizar_titulo_aba_mov_estoque()
            self.atualizar_aba_produtos_selecionados()
            return

        try:
            self._mov_estoque_df = pl.read_parquet(path)
            self._mov_estoque_file_path = path
            self._reset_table_resize_flag("mov_estoque")

            # Popular combo de id_agrupado
            id_atual = self.mov_filter_id.currentText()
            ids = self._mov_estoque_df.get_column("id_agrupado").unique().sort().to_list()
            self._popular_combo_texto(self.mov_filter_id, [str(i) for i in ids], id_atual, "")

            self.aplicar_filtros_mov_estoque()
            self.atualizar_aba_produtos_selecionados()
        except Exception as e:
            self.show_error("Erro de leitura", f"Falha ao ler mov_estoque: {e}")

    def aplicar_filtros_mov_estoque(self) -> None:
        if self._mov_estoque_df.is_empty():
            return
        try:
            id_agrup = self.mov_filter_id.currentText().strip()
            desc = self.mov_filter_desc.text().strip().lower()
            ncm = self.mov_filter_ncm.text().strip()
            tipo = self.mov_filter_tipo.currentText()
            texto = self.mov_filter_texto.text().strip().lower()
            data_col = self.mov_filter_data_col.currentText().strip()
            data_ini = self._valor_qdate_ativo(self.mov_filter_data_ini.date())
            data_fim = self._valor_qdate_ativo(self.mov_filter_data_fim.date())
            num_col = self.mov_filter_num_col.currentText().strip()
            num_min = self.mov_filter_num_min.text().strip()
            num_max = self.mov_filter_num_max.text().strip()

            df_filtrado = self._mov_estoque_df

            # Filtro Cruzado
            if self._filtro_cruzado_anuais_ids:
                df_filtrado = df_filtrado.filter(pl.col("id_agrupado").is_in(self._filtro_cruzado_anuais_ids))

            if id_agrup:
                df_filtrado = df_filtrado.filter(pl.col("id_agrupado").cast(pl.Utf8).str.contains(id_agrup))
            if desc:
                col_desc = "descr_padrao" if "descr_padrao" in df_filtrado.columns else "Descr_item"
                df_filtrado = df_filtrado.filter(
                    pl.col(col_desc).cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(desc, literal=True)
                )
            if ncm:
                col_ncm = "ncm_padrao" if "ncm_padrao" in df_filtrado.columns else "Ncm"
                df_filtrado = df_filtrado.filter(
                    pl.col(col_ncm).cast(pl.Utf8, strict=False).fill_null("").str.contains(ncm, literal=True)
                )
            if tipo == "Entradas":
                df_filtrado = df_filtrado.filter(pl.col("Tipo_operacao").cast(pl.Utf8, strict=False).str.contains("ENTRADA", literal=True))
            elif tipo == "Saidas":
                df_filtrado = df_filtrado.filter(pl.col("Tipo_operacao").cast(pl.Utf8, strict=False).str.contains("SAIDA", literal=True))

            df_filtrado = self._filtrar_intervalo_data(df_filtrado, data_col, data_ini, data_fim)
            df_filtrado = self._filtrar_intervalo_numerico(df_filtrado, num_col, num_min, num_max)

            if texto:
                df_filtrado = self._filtrar_texto_em_colunas(df_filtrado, texto)

            self.mov_estoque_model.set_dataframe(df_filtrado)
            self._resize_table_once(self.mov_estoque_table, "mov_estoque")
            if not self._aplicar_preferencias_tabela("mov_estoque", self.mov_estoque_table, self.mov_estoque_model):
                self._aplicar_ordenacao_padrao(
                    self.mov_estoque_table,
                    self.mov_estoque_model,
                    ["ordem_operacoes", "Dt_doc", "Dt_e_s", "id_agrupado"],
                )
                self._aplicar_preset_mov_estoque()
            if "ordem_operacoes" in self.mov_estoque_model.dataframe.columns:
                offset = 1 if getattr(self.mov_estoque_model, "_checkable", False) else 0
                idx_ordem = self.mov_estoque_model.dataframe.columns.index("ordem_operacoes") + offset
                self.mov_estoque_table.setColumnHidden(idx_ordem, False)
            self.lbl_mov_estoque_status.setText(
                f"Movimentacoes: {df_filtrado.height:,} de {self._mov_estoque_df.height:,} linhas."
                + (" (FILTRO CRUZADO ATIVO)" if self._filtro_cruzado_anuais_ids else "")
            )
            self.lbl_mov_estoque_filtros.setText(
                self._formatar_resumo_filtros(
                    [
                        ("id_agrupado", id_agrup),
                        ("descricao", desc),
                        ("ncm", ncm),
                        ("tipo", "" if tipo == "Todos" else tipo),
                        ("data", f"{data_ini.toString('dd/MM/yyyy') if data_ini else ''}..{data_fim.toString('dd/MM/yyyy') if data_fim else ''}" if data_ini or data_fim else ""),
                        (num_col, f"{num_min or ''}..{num_max or ''}" if num_min or num_max else ""),
                        ("texto", texto),
                        ("cruzado", ",".join(self._filtro_cruzado_anuais_ids[:3]) + ("..." if len(self._filtro_cruzado_anuais_ids) > 3 else "") if self._filtro_cruzado_anuais_ids else ""),
                    ]
                )
            )
            self._atualizar_titulo_aba_mov_estoque(df_filtrado.height, self._mov_estoque_df.height)
            self._salvar_preferencias_tabela("mov_estoque", self.mov_estoque_table, self.mov_estoque_model)
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao filtrar mov_estoque: {e}")

    def _atualizar_titulo_aba_mov_estoque(self, visiveis: int | None = None, total: int | None = None) -> None:
        if not hasattr(self, "estoque_tabs") or not hasattr(self, "tab_mov_estoque"):
            return
        idx = self.estoque_tabs.indexOf(self.tab_mov_estoque)
        if idx < 0:
            return
        if visiveis is None:
            self.estoque_tabs.setTabText(idx, "Tabela mov_estoque")
            return
        self.estoque_tabs.setTabText(idx, f"Tabela mov_estoque ({visiveis})")

    def _atualizar_titulo_aba_anual(self, visiveis: int | None = None, total: int | None = None) -> None:
        if not hasattr(self, "estoque_tabs") or not hasattr(self, "tab_aba_anual"):
            return
        idx = self.estoque_tabs.indexOf(self.tab_aba_anual)
        if idx < 0:
            return
        if visiveis is None:
            self.estoque_tabs.setTabText(idx, "Tabela anual")
            return
        if total is None:
            self.estoque_tabs.setTabText(idx, f"Tabela anual ({visiveis})")
            return
        self.estoque_tabs.setTabText(idx, f"Tabela anual ({visiveis}/{total})")

    def _atualizar_titulo_aba_mensal(self, visiveis: int | None = None, total: int | None = None) -> None:
        if not hasattr(self, "estoque_tabs") or not hasattr(self, "tab_aba_mensal"):
            return
        idx = self.estoque_tabs.indexOf(self.tab_aba_mensal)
        if idx < 0:
            return
        if visiveis is None:
            self.estoque_tabs.setTabText(idx, "Tabela mensal")
            return
        if total is None:
            self.estoque_tabs.setTabText(idx, f"Tabela mensal ({visiveis})")
            return
        self.estoque_tabs.setTabText(idx, f"Tabela mensal ({visiveis}/{total})")

    def _atualizar_titulo_aba_produtos_selecionados(self, visiveis: int | None = None, total: int | None = None) -> None:
        if not hasattr(self, "estoque_tabs") or not hasattr(self, "tab_produtos_selecionados"):
            return
        idx = self.estoque_tabs.indexOf(self.tab_produtos_selecionados)
        if idx < 0:
            return
        if visiveis is None:
            self.estoque_tabs.setTabText(idx, "Produtos selecionados")
            return
        if total is None:
            self.estoque_tabs.setTabText(idx, f"Produtos selecionados ({visiveis})")
            return
        self.estoque_tabs.setTabText(idx, f"Produtos selecionados ({visiveis}/{total})")

    def _atualizar_titulo_aba_id_agrupados(self, visiveis: int | None = None, total: int | None = None) -> None:
        if not hasattr(self, "estoque_tabs") or not hasattr(self, "tab_id_agrupados"):
            return
        idx = self.estoque_tabs.indexOf(self.tab_id_agrupados)
        if idx < 0:
            return
        if visiveis is None:
            self.estoque_tabs.setTabText(idx, "id_agrupados")
            return
        if total is None:
            self.estoque_tabs.setTabText(idx, f"id_agrupados ({visiveis})")
            return
        self.estoque_tabs.setTabText(idx, f"id_agrupados ({visiveis}/{total})")

    def _popular_combo_texto(self, combo: QComboBox, valores: list[str], valor_atual: str = "", primeiro_item: str = "") -> None:
        combo.blockSignals(True)
        combo.clear()
        if primeiro_item is not None:
            combo.addItem(primeiro_item)
        combo.addItems([str(v) for v in valores])
        if valor_atual:
            combo.setCurrentText(valor_atual)
        combo.blockSignals(False)

    def _filtrar_texto_em_colunas(self, df: pl.DataFrame, texto: str) -> pl.DataFrame:
        texto = (texto or "").strip().lower()
        if not texto or df.is_empty():
            return df

        colunas_busca = [c for c in df.columns if df.schema[c] in [pl.Utf8, pl.Categorical]]
        if not colunas_busca:
            return df

        expr = None
        for col in colunas_busca:
            atual = pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(texto, literal=True)
            expr = atual if expr is None else (expr | atual)
        return df.filter(expr) if expr is not None else df

    def _valor_qdate_ativo(self, value: QDate) -> QDate | None:
        return None if not value.isValid() or value == QDate(1900, 1, 1) else value

    def _parse_numero_filtro(self, valor: str) -> float | None:
        bruto = (valor or "").strip()
        if not bruto:
            return None
        try:
            return float(bruto.replace(",", "."))
        except Exception:
            return None

    def _filtrar_intervalo_numerico(
        self,
        df: pl.DataFrame,
        coluna: str | None,
        valor_min: str,
        valor_max: str,
    ) -> pl.DataFrame:
        if not coluna or coluna not in df.columns:
            return df

        minimo = self._parse_numero_filtro(valor_min)
        maximo = self._parse_numero_filtro(valor_max)
        if minimo is None and maximo is None:
            return df

        expr_col = pl.col(coluna).cast(pl.Float64, strict=False)
        if minimo is not None:
            df = df.filter(expr_col >= minimo)
        if maximo is not None:
            df = df.filter(expr_col <= maximo)
        return df

    def _filtrar_intervalo_data(
        self,
        df: pl.DataFrame,
        coluna: str | None,
        data_ini: QDate | None,
        data_fim: QDate | None,
    ) -> pl.DataFrame:
        if not coluna or coluna not in df.columns or (data_ini is None and data_fim is None):
            return df

        col_data = (
            pl.col(coluna)
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.replace_all(r"[^0-9]", "")
            .str.slice(0, 8)
            .str.strptime(pl.Date, format="%Y%m%d", strict=False)
        )
        if data_ini is not None:
            df = df.filter(col_data >= pl.lit(data_ini.toPython()))
        if data_fim is not None:
            df = df.filter(col_data <= pl.lit(data_fim.toPython()))
        return df

    def _preferencia_tabela_key(self, aba: str, scope: str | None = None) -> str:
        escopo = scope or (self.state.current_cnpj or "__global__")
        return f"preferencias_tabela::{aba}::{escopo}"

    def _consulta_scope(self) -> str:
        arquivo = self.state.current_file.name if self.state.current_file else "__sem_arquivo__"
        cnpj = self.state.current_cnpj or "__global__"
        return f"{cnpj}::{arquivo}"

    def _carregar_preferencias_tabela(self, aba: str, scope: str | None = None) -> dict:
        prefs = self.selection_service.get_value(self._preferencia_tabela_key(aba, scope), {})
        return prefs if isinstance(prefs, dict) else {}

    def _capturar_estado_tabela(self, table: QTableView, model: PolarsTableModel) -> dict:
        offset = 1 if getattr(model, "_checkable", False) else 0
        colunas = model.dataframe.columns
        header = table.horizontalHeader()
        visiveis = [
            nome
            for _visual, nome in sorted(
                (
                    (header.visualIndex(idx + offset), nome)
                    for idx, nome in enumerate(colunas)
                    if not table.isColumnHidden(idx + offset)
                ),
                key=lambda item: item[0],
            )
        ]
        estado = {
            "visible_columns": visiveis,
            "column_order": visiveis,
            "header_state": self._serializar_estado_header(table),
        }
        if getattr(model, "_last_sort_column", None):
            estado["sort_column"] = model._last_sort_column
            estado["sort_order"] = "desc" if model._last_sort_order == Qt.DescendingOrder else "asc"
        return estado

    def _aplicar_estado_tabela(self, table: QTableView, model: PolarsTableModel, prefs: dict) -> bool:
        if not prefs or model.dataframe.is_empty():
            return False

        aplicado = False
        visiveis = prefs.get("visible_columns")
        if isinstance(visiveis, list) and visiveis:
            self._aplicar_preset_colunas(table, model.dataframe.columns, [str(v) for v in visiveis])
            aplicado = True

        sort_column = prefs.get("sort_column")
        sort_order = Qt.DescendingOrder if prefs.get("sort_order") == "desc" else Qt.AscendingOrder
        if isinstance(sort_column, str) and sort_column in model.dataframe.columns:
            idx = model.dataframe.columns.index(sort_column) + (1 if getattr(model, "_checkable", False) else 0)
            model.sort(idx, sort_order)
            table.sortByColumn(idx, sort_order)
            aplicado = True

        header_state = prefs.get("header_state")
        if isinstance(header_state, str) and header_state:
            aplicado = self._restaurar_estado_header(table, header_state) or aplicado
        return aplicado

    def _colunas_estado_perfil(self, prefs: dict, model: PolarsTableModel) -> list[str] | None:
        if not isinstance(prefs, dict) or model.dataframe.is_empty():
            return None

        raw = prefs.get("visible_columns")
        if not isinstance(raw, list) or not raw:
            return None

        visiveis = ordenar_colunas_perfil(
            list(model.dataframe.columns),
            raw,
            prefs.get("column_order") if isinstance(prefs.get("column_order"), list) else None,
        )
        if not visiveis:
            return None

        header_state = prefs.get("header_state")
        if not isinstance(header_state, str) or not header_state:
            return visiveis

        probe = QTableView()
        try:
            probe.setModel(model)
            if not self._restaurar_estado_header(probe, header_state):
                return visiveis

            offset = 1 if getattr(model, "_checkable", False) else 0
            ordem = [
                nome
                for _visual, nome in sorted(
                    (
                        (probe.horizontalHeader().visualIndex(idx + offset), nome)
                        for idx, nome in enumerate(model.dataframe.columns)
                        if nome in visiveis
                    ),
                    key=lambda item: item[0],
                )
            ]
            return ordenar_colunas_perfil(list(model.dataframe.columns), visiveis, ordem)
        finally:
            probe.setModel(None)

    def _nomes_perfis_nomeados_tabela(self, aba: str, scope: str | None = None) -> list[str]:
        prefs = self._carregar_preferencias_tabela(aba, scope)
        perfis = prefs.get("named_profiles", {})
        if not isinstance(perfis, dict):
            return []
        return sorted([str(nome) for nome in perfis.keys() if str(nome).strip()], key=lambda v: v.lower())

    def _obter_estado_perfil_nomeado(self, aba: str, perfil: str, scope: str | None = None) -> dict | None:
        prefs = self._carregar_preferencias_tabela(aba, scope)
        perfis = prefs.get("named_profiles", {})
        if not isinstance(perfis, dict):
            return None
        estado = perfis.get(perfil)
        return estado if isinstance(estado, dict) else None

    def _atualizar_combo_perfis_tabela(
        self,
        combo: QComboBox,
        aba: str,
        presets: list[str],
        scope: str | None = None,
    ) -> None:
        atual = combo.currentText().strip()
        nomes = presets + [n for n in self._nomes_perfis_nomeados_tabela(aba, scope) if n not in presets]
        combo.blockSignals(True)
        combo.clear()
        combo.addItems(nomes)
        if atual and atual in nomes:
            combo.setCurrentText(atual)
        elif nomes:
            combo.setCurrentIndex(0)
        combo.blockSignals(False)

    def _salvar_perfil_nomeado_tabela(
        self,
        aba: str,
        table: QTableView,
        model: PolarsTableModel,
        nome: str,
        scope: str | None = None,
    ) -> None:
        if model.dataframe.is_empty():
            return
        nome_limpo = nome.strip()
        if not nome_limpo:
            return
        prefs = self._carregar_preferencias_tabela(aba, scope)
        perfis = prefs.get("named_profiles", {})
        if not isinstance(perfis, dict):
            perfis = {}
        perfis[nome_limpo] = self._capturar_estado_tabela(table, model)
        prefs["named_profiles"] = perfis
        self.selection_service.set_value(self._preferencia_tabela_key(aba, scope), prefs)

    def _serializar_estado_header(self, table: QTableView) -> str:
        estado = bytes(table.horizontalHeader().saveState())
        return base64.b64encode(estado).decode("ascii")

    def _restaurar_estado_header(self, table: QTableView, valor: str) -> bool:
        try:
            bruto = base64.b64decode(valor.encode("ascii"))
            return bool(table.horizontalHeader().restoreState(QByteArray(bruto)))
        except Exception:
            return False

    def _salvar_preferencias_tabela(
        self,
        aba: str,
        table: QTableView,
        model: PolarsTableModel,
        scope: str | None = None,
    ) -> None:
        if model.dataframe.is_empty():
            return
        prefs = self._carregar_preferencias_tabela(aba, scope)
        prefs.update(self._capturar_estado_tabela(table, model))
        self.selection_service.set_value(self._preferencia_tabela_key(aba, scope), prefs)

    def _aplicar_preferencias_tabela(
        self,
        aba: str,
        table: QTableView,
        model: PolarsTableModel,
        scope: str | None = None,
    ) -> bool:
        prefs = self._carregar_preferencias_tabela(aba, scope)
        return self._aplicar_estado_tabela(table, model, prefs)

    def _obter_colunas_preset_perfil(self, perfil: str, colunas: list[str], contexto: str) -> list[str]:
        nome = (perfil or "").strip().lower()
        if contexto == "mov_estoque":
            mapa = {
                "padrao": [
                    "ordem_operacoes", "Tipo_operacao", "fonte",
                    "id_agrupado", "descr_padrao", "Descr_item", "Descr_compl", "Cod_item", "Cod_barra", "Ncm", "Cest", "Tipo_item",
                    "Chv_nfe", "mod", "Ser", "num_nfe", "Num_item", "Dt_doc", "Dt_e_s", "nsu", "finnfe", "infprot_cstat", "co_uf_emit", "co_uf_dest",
                    "Cfop", "Cst", "Aliq_icms", "Vl_bc_icms", "Vl_icms", "vl_bc_icms_st", "vl_icms_st", "aliq_st",
                    "Qtd", "q_conv", "Unid", "unid_ref", "fator",
                    "Vl_item", "preco_item", "preco_unit", "custo_medio_anual",
                    "saldo_estoque_anual", "entr_desac_anual", "mov_rep", "excluir_estoque", "dev_simples", "dev_venda", "dev_compra", "dev_ent_simples",
                    "ncm_padrao", "cest_padrao", "co_sefin_agr", "it_pc_interna", "it_in_st", "it_pc_mva", "it_in_mva_ajustado",
                    "it_in_isento_icms", "it_in_reducao", "it_pc_reducao", "it_in_combustivel", "it_in_pmpf", "it_in_reducao_credito",
                ],
                "exportar": ["ordem_operacoes", "Tipo_operacao", "fonte", "id_agrupado", "descr_padrao", "Descr_item", "Dt_doc", "Dt_e_s", "Cfop", "Qtd", "q_conv", "saldo_estoque_anual", "entr_desac_anual", "custo_medio_anual", "preco_item", "preco_unit", "unid_ref", "fator", "mov_rep", "dev_simples", "excluir_estoque"],
                "contribuinte": ["ordem_operacoes", "Tipo_operacao", "Dt_doc", "id_agrupado", "descr_padrao", "Qtd", "q_conv", "unid_ref", "preco_item", "preco_unit", "saldo_estoque_anual", "entr_desac_anual"],
                "auditoria": ["ordem_operacoes", "Tipo_operacao", "fonte", "id_agrupado", "descr_padrao", "Dt_doc", "Dt_e_s", "Cfop", "q_conv", "saldo_estoque_anual", "entr_desac_anual", "mov_rep", "dev_simples", "excluir_estoque"],
                "auditoria fiscal": ["ordem_operacoes", "Tipo_operacao", "fonte", "id_agrupado", "descr_padrao", "Descr_item", "Descr_compl", "Cod_item", "Cod_barra", "Ncm", "Cest", "Tipo_item", "Chv_nfe", "mod", "Ser", "num_nfe", "Num_item", "Dt_doc", "Dt_e_s", "nsu", "finnfe", "infprot_cstat", "co_uf_emit", "co_uf_dest", "Cfop", "Cst", "Aliq_icms", "Vl_bc_icms", "Vl_icms", "vl_bc_icms_st", "vl_icms_st", "aliq_st", "Qtd", "q_conv", "Unid", "unid_ref", "fator", "Vl_item", "preco_item", "preco_unit", "custo_medio_anual", "saldo_estoque_anual", "entr_desac_anual", "mov_rep", "excluir_estoque", "dev_simples", "dev_venda", "dev_compra", "dev_ent_simples", "co_sefin_agr", "it_pc_interna", "it_in_st", "it_pc_mva", "it_in_mva_ajustado", "it_in_isento_icms", "it_in_reducao", "it_pc_reducao", "it_in_combustivel", "it_in_pmpf", "it_in_reducao_credito"],
                "estoque": ["ordem_operacoes", "Tipo_operacao", "id_agrupado", "descr_padrao", "Dt_doc", "q_conv", "saldo_estoque_anual", "unid_ref", "fator"],
                "custos": ["ordem_operacoes", "Tipo_operacao", "id_agrupado", "descr_padrao", "Dt_doc", "q_conv", "preco_item", "preco_unit", "custo_medio_anual", "saldo_estoque_anual"],
            }
        elif contexto in {"agregacao_top", "agregacao_bottom"}:
            mapa = {
                "padrao": [
                    "id_agrupado", "descr_padrao",
                    "ids_origem_agrupamento",
                    "preco_medio_compra", "preco_medio_venda",
                    "total_entradas", "total_saidas", "total_movimentacao",
                    "total_compras", "qtd_compras_total",
                    "total_vendas", "qtd_vendas_total",
                    "ncm_padrao", "cest_padrao", "gtin_padrao",
                    "lista_itens_agrupados",
                    "lista_ncm", "lista_cest", "lista_gtin", "lista_descricoes", "lista_desc_compl",
                    "co_sefin_padrao", "co_sefin_agr", "lista_unidades", "fontes",
                ],
                "auditoria": [
                    "id_agrupado", "descr_padrao",
                    "ids_origem_agrupamento", "lista_itens_agrupados",
                    "ncm_padrao", "cest_padrao", "gtin_padrao",
                    "lista_ncm", "lista_cest", "lista_gtin", "lista_descricoes", "lista_desc_compl",
                    "co_sefin_padrao", "co_sefin_agr", "lista_co_sefin",
                    "co_sefin_divergentes", "lista_unidades", "fontes",
                    "total_entradas", "total_saidas", "total_movimentacao",
                    "total_compras", "qtd_compras_total", "preco_medio_compra",
                    "total_vendas", "qtd_vendas_total", "preco_medio_venda",
                    "lista_chave_produto",
                ],
                "estoque": [
                    "id_agrupado", "descr_padrao",
                    "ids_origem_agrupamento",
                    "total_entradas", "total_saidas", "total_movimentacao",
                    "total_compras", "qtd_compras_total",
                    "total_vendas", "qtd_vendas_total",
                    "lista_unidades", "lista_descricoes", "lista_desc_compl", "lista_itens_agrupados", "fontes",
                    "ncm_padrao", "cest_padrao",
                ],
                "custos": [
                    "id_agrupado", "descr_padrao",
                    "ids_origem_agrupamento",
                    "preco_medio_compra", "preco_medio_venda",
                    "total_entradas", "total_saidas", "total_movimentacao",
                    "total_compras", "qtd_compras_total",
                    "total_vendas", "qtd_vendas_total",
                    "lista_ncm", "lista_cest", "lista_gtin", "lista_descricoes", "lista_desc_compl", "lista_itens_agrupados",
                    "lista_unidades", "fontes",
                ],
            }
        elif nome in {"", "padrao"}:
            return colunas
        elif contexto == "conversao":
            mapa = {
                "auditoria": ["id_agrupado", "id_produtos", "descr_padrao", "lista_descricoes_produto", "unid", "unid_ref", "fator", "fator_calculado", "preco_medio", "preco_medio_ref", "origem_preco"],
                "estoque": ["id_agrupado", "descr_padrao", "unid", "unid_ref", "fator", "fator_calculado"],
                "custos": ["id_agrupado", "descr_padrao", "unid", "unid_ref", "preco_medio", "preco_medio_ref", "fator_calculado", "fator", "origem_preco"],
            }
        elif contexto == "aba_mensal":
            mapa = {
                "exportar": ["ano", "mes", "id_agregado", "descr_padrao", "ST", "it_in_st", "valor_entradas", "qtd_entradas", "pme_mes", "valor_saidas", "qtd_saidas", "pms_mes", "MVA", "MVA_ajustado", "entradas_desacob", "ICMS_entr_desacob", "saldo_mes", "custo_medio_mes", "valor_estoque"],
                "auditoria": ["ano", "mes", "id_agregado", "descr_padrao", "ST", "it_in_st", "valor_entradas", "qtd_entradas", "valor_saidas", "qtd_saidas", "entradas_desacob", "ICMS_entr_desacob", "saldo_mes", "custo_medio_mes", "valor_estoque"],
                "estoque": ["ano", "mes", "id_agregado", "descr_padrao", "unids_ref_mes", "qtd_entradas", "qtd_saidas", "saldo_mes", "custo_medio_mes", "valor_estoque"],
                "custos": ["ano", "mes", "id_agregado", "descr_padrao", "valor_entradas", "pme_mes", "valor_saidas", "pms_mes", "custo_medio_mes", "valor_estoque"],
            }
        elif contexto == "aba_anual":
            mapa = {
                "exportar": ["ano", "id_agregado", "descr_padrao", "unid_ref", "estoque_inicial", "entradas", "saidas", "estoque_final", "saidas_calculadas", "saldo_final", "entradas_desacob", "saidas_desacob", "estoque_final_desacob", "pme", "pms", "ICMS_saidas_desac", "ICMS_estoque_desac"],
                "auditoria": ["ano", "id_agregado", "descr_padrao", "unid_ref", "estoque_inicial", "entradas", "saidas", "estoque_final", "saldo_final", "entradas_desacob", "saidas_desacob", "estoque_final_desacob", "ICMS_saidas_desac"],
                "estoque": ["ano", "id_agregado", "descr_padrao", "unid_ref", "estoque_inicial", "entradas", "saidas", "estoque_final", "saldo_final"],
                "custos": ["ano", "id_agregado", "descr_padrao", "unid_ref", "entradas", "saidas", "pme", "pms", "ICMS_saidas_desac"],
            }
        elif contexto == "nfe_entrada":
            mapa = {
                "auditoria": ["data_classificacao", "fonte_documento", "tipo_operacao", "nnf", "prod_nitem", "id_agrupado", "descr_padrao", "prod_xprod", "prod_ncm", "prod_cest", "co_sefin_inferido", "co_sefin_agr", "it_pc_interna", "it_in_st", "it_in_isento_icms", "it_in_reducao", "it_pc_reducao", "it_pc_mva", "it_in_mva_ajustado", "xnome_emit", "xnome_dest", "chave_acesso"],
                "estoque": ["data_classificacao", "fonte_documento", "id_agrupado", "descr_padrao", "prod_xprod", "prod_ncm", "prod_cest", "co_sefin_agr", "it_pc_interna", "it_in_st", "it_in_isento_icms", "it_in_reducao", "prod_ucom", "prod_qcom", "prod_vprod"],
                "custos": ["data_classificacao", "fonte_documento", "id_agrupado", "descr_padrao", "prod_xprod", "co_sefin_agr", "it_pc_interna", "it_in_st", "it_in_isento_icms", "it_in_reducao", "it_pc_reducao", "it_pc_mva", "it_in_mva_ajustado", "prod_vuncom", "prod_vprod"],
            }
        elif contexto == "produtos_selecionados":
            mapa = {
                "auditoria": ["id_agregado", "descr_padrao", "total_ICMS_entr_desacob", "total_ICMS_saidas_desac", "total_ICMS_estoque_desac", "total_ICMS_total"],
                "estoque": ["id_agregado", "descr_padrao", "total_ICMS_entr_desacob", "total_ICMS_estoque_desac", "total_ICMS_total"],
                "custos": ["id_agregado", "descr_padrao", "total_ICMS_entr_desacob", "total_ICMS_saidas_desac", "total_ICMS_estoque_desac", "total_ICMS_total"],
            }
        elif contexto == "id_agrupados":
            mapa = {
                "auditoria": ["id_agrupado", "descr_padrao", "lista_descricoes", "lista_desc_compl", "lista_codigos", "lista_unidades"],
                "estoque": ["id_agrupado", "descr_padrao", "lista_unidades", "lista_descricoes", "lista_desc_compl"],
                "custos": ["id_agrupado", "descr_padrao", "lista_codigos", "lista_unidades"],
            }
        else:
            mapa = {
                "auditoria": ["cnpj", "periodo", "id_agrupado", "id_agregado", "descr_padrao", "descricao", "descr", "Descr_item", "Ncm", "ncm_padrao", "cest_padrao", "Cfop", "valor_item", "preco_item", "q_conv", "Qtd", "total_compras", "qtd_compras_total", "preco_medio_compra", "total_vendas", "qtd_vendas_total", "preco_medio_venda", "saldo_final", "entradas_desacob"],
                "estoque": ["id_agrupado", "id_agregado", "descr_padrao", "descricao", "Descr_item", "ncm_padrao", "Ncm", "unid_ref", "q_conv", "Qtd", "total_entradas", "total_saidas", "total_movimentacao", "total_compras", "total_vendas", "saldo_final", "estoque_final"],
                "custos": ["id_agrupado", "id_agregado", "descr_padrao", "descricao", "Descr_item", "total_entradas", "total_saidas", "total_movimentacao", "total_compras", "qtd_compras_total", "preco_medio_compra", "total_vendas", "qtd_vendas_total", "preco_medio_venda", "preco_item", "preco_unit", "valor_item", "q_conv", "Qtd", "pme", "pms", "custo_medio_anual"],
            }

        desejadas = mapa.get(nome, colunas)
        selecionadas = [c for c in desejadas if c in colunas]
        return selecionadas or colunas

    def _aplicar_layout_padrao_agregacao(
        self,
        contexto: str,
        table: QTableView,
        model: PolarsTableModel,
        perfil: str,
    ) -> None:
        if contexto not in {"agregacao_top", "agregacao_bottom"} or model.dataframe.is_empty():
            return
        ordem = self._obter_colunas_preset_perfil(perfil, model.dataframe.columns, contexto)
        self._aplicar_ordem_colunas(table, ordem)
        colunas = model.dataframe.columns
        offset = 1 if getattr(model, "_checkable", False) else 0
        larguras = {
            "id_agrupado": 150,
            "descr_padrao": 320,
            "ids_origem_agrupamento": 180,
            "preco_medio_compra": 150,
            "preco_medio_venda": 150,
            "total_entradas": 145,
            "total_saidas": 145,
            "total_movimentacao": 155,
            "total_compras": 140,
            "qtd_compras_total": 140,
            "total_vendas": 140,
            "qtd_vendas_total": 140,
            "lista_ncm": 180,
            "lista_cest": 180,
            "lista_gtin": 180,
            "lista_descricoes": 340,
            "lista_desc_compl": 320,
            "lista_itens_agrupados": 340,
        }
        for nome, largura in larguras.items():
            if nome in colunas:
                table.setColumnWidth(colunas.index(nome) + offset, largura)

    def _abrir_menu_colunas_tabela(self, aba: str, table: QTableView, pos=None, scope: str | None = None) -> None:
        model = table.model()
        if not isinstance(model, PolarsTableModel) or model.dataframe.is_empty():
            return
        offset = 1 if getattr(model, "_checkable", False) else 0
        header = table.horizontalHeader()
        colunas = [
            nome for _visual, nome in sorted(
                ((header.visualIndex(idx + offset), nome) for idx, nome in enumerate(model.dataframe.columns)),
                key=lambda item: item[0],
            )
        ]
        visiveis = [
            nome
            for nome in colunas
            if nome in model.dataframe.columns
            and not table.isColumnHidden(model.dataframe.columns.index(nome) + offset)
        ]
        dialog = ColumnSelectorDialog(colunas, visiveis, self)
        if not dialog.exec():
            return
        selecionadas = dialog.selected_columns()
        if not selecionadas:
            self.show_error("Selecao invalida", "Pelo menos uma coluna deve permanecer visivel.")
            return
        self._aplicar_ordem_colunas(table, dialog.column_order())
        self._aplicar_preset_colunas(table, colunas, selecionadas)
        self._salvar_preferencias_tabela(aba, table, model, scope)

    def _aplicar_perfil_tabela(
        self,
        aba: str,
        table: QTableView,
        model: PolarsTableModel,
        perfil: str,
        contexto: str,
        scope: str | None = None,
    ) -> None:
        if model.dataframe.is_empty():
            return
        perfil_salvo = self._obter_estado_perfil_nomeado(aba, perfil, scope)
        if perfil_salvo is not None:
            self._aplicar_estado_tabela(table, model, perfil_salvo)
            self._salvar_preferencias_tabela(aba, table, model, scope)
            return
        visiveis = self._obter_colunas_preset_perfil(perfil, model.dataframe.columns, contexto)
        self._aplicar_preset_colunas(table, model.dataframe.columns, visiveis)
        self._aplicar_layout_padrao_agregacao(contexto, table, model, perfil)
        self._salvar_preferencias_tabela(aba, table, model, scope)

    def _salvar_perfil_tabela_com_dialogo(
        self,
        aba: str,
        table: QTableView,
        model: PolarsTableModel,
        combo: QComboBox,
        presets: list[str],
        scope: str | None = None,
    ) -> None:
        if model.dataframe.is_empty():
            self.show_info("Salvar perfil", "Nao ha dados carregados para salvar um perfil.")
            return
        nome, ok = QInputDialog.getText(self, "Salvar perfil", "Nome do perfil:")
        nome = (nome or "").strip()
        if not ok or not nome:
            return
        if nome.lower() in {p.lower() for p in presets} and nome.lower() != "exportar":
            self.show_error("Nome invalido", "Escolha um nome diferente dos perfis padrao.")
            return
        self._salvar_perfil_nomeado_tabela(aba, table, model, nome, scope)
        self._atualizar_combo_perfis_tabela(combo, aba, presets, scope)
        combo.setCurrentText(nome)

    def _aplicar_ordenacao_padrao(
        self,
        table: QTableView,
        model: PolarsTableModel,
        colunas_prioritarias: list[str],
        order: Qt.SortOrder = Qt.AscendingOrder,
    ) -> None:
        if model.dataframe.is_empty():
            return

        colunas = model.dataframe.columns
        deslocamento = 1 if getattr(model, "_checkable", False) else 0
        for nome in colunas_prioritarias:
            if nome not in colunas:
                continue
            idx = colunas.index(nome) + deslocamento
            model.sort(idx, order)
            table.sortByColumn(idx, order)
            return

    def _aplicar_preset_colunas(self, table: QTableView, colunas: list[str], visiveis: list[str]) -> None:
        visiveis_set = set(visiveis)
        model = table.model()
        if not isinstance(model, PolarsTableModel):
            return
        offset = 1 if getattr(model, "_checkable", False) else 0
        colunas_modelo = list(model.dataframe.columns)
        for idx, nome in enumerate(colunas_modelo):
            table.setColumnHidden(idx + offset, nome not in visiveis_set)

    def _aplicar_ordem_colunas(self, table: QTableView, ordem_colunas: list[str]) -> None:
        model = table.model()
        if not isinstance(model, PolarsTableModel) or model.dataframe.is_empty():
            return
        header = table.horizontalHeader()
        offset = 1 if getattr(model, "_checkable", False) else 0
        for idx, nome in enumerate(ordem_colunas):
            if nome not in model.dataframe.columns:
                continue
            logical_index = model.dataframe.columns.index(nome) + offset
            visual_atual = header.visualIndex(logical_index)
            visual_destino = idx + offset
            if visual_atual != visual_destino:
                header.moveSection(visual_atual, visual_destino)

    def _dataframe_colunas_visiveis(self, table: QTableView, model: PolarsTableModel, df: pl.DataFrame | None = None) -> pl.DataFrame:
        base_df = df if df is not None else model.dataframe
        if base_df.is_empty():
            return base_df
        offset = 1 if getattr(model, "_checkable", False) else 0
        colunas_modelo = list(model.dataframe.columns)
        header = table.horizontalHeader()
        visiveis = [nome for idx, nome in enumerate(colunas_modelo) if not table.isColumnHidden(idx + offset)]
        ordem_visual = [
            nome
            for _visual, nome in sorted(
                ((header.visualIndex(idx + offset), nome) for idx, nome in enumerate(colunas_modelo)),
                key=lambda item: item[0],
            )
        ]
        visiveis = ordenar_colunas_visiveis(list(base_df.columns), visiveis, ordem_visual)
        return base_df.select(visiveis) if visiveis else base_df

    def _dataframe_colunas_perfil(
        self,
        aba: str,
        contexto: str,
        model: PolarsTableModel,
        df: pl.DataFrame | None = None,
        perfil: str = "Exportar",
        scope: str | None = None,
    ) -> pl.DataFrame:
        base_df = df if df is not None else model.dataframe
        if base_df.is_empty():
            return base_df

        estado_perfil = self._obter_estado_perfil_nomeado(aba, perfil, scope)
        visiveis = self._colunas_estado_perfil(estado_perfil, model)

        if not visiveis:
            visiveis = self._obter_colunas_preset_perfil(perfil, list(base_df.columns), contexto)
            visiveis = [col for col in visiveis if col in base_df.columns]

        return base_df.select(visiveis) if visiveis else base_df

    def _refresh_profile_combos(self) -> None:
        combos = [
            (self.consulta_profile, "consulta", ["Padrao", "Auditoria", "Estoque", "Custos"], self._consulta_scope()),
            (self.top_profile, "agregacao_top", ["Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.bottom_profile, "agregacao_bottom", ["Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.conversao_profile, "conversao", ["Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.mov_profile, "mov_estoque", ["Exportar", "Padrao", "Contribuinte", "Auditoria", "Auditoria Fiscal", "Estoque", "Custos"], None),
            (self.mensal_profile, "aba_mensal", ["Exportar", "Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.anual_profile, "aba_anual", ["Exportar", "Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.produtos_sel_profile, "produtos_selecionados", ["Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.nfe_entrada_profile, "nfe_entrada", ["Padrao", "Auditoria", "Estoque", "Custos"], None),
            (self.id_agrupados_profile, "id_agrupados", ["Padrao", "Auditoria", "Estoque", "Custos"], None),
        ]
        for combo, aba, presets, scope in combos:
            if combo is not None:
                self._atualizar_combo_perfis_tabela(combo, aba, presets, scope)

    def _aplicar_preset_mov_estoque(self) -> None:
        if self.mov_estoque_model.dataframe.is_empty():
            return
        visiveis = self._obter_colunas_preset_perfil("exportar", self.mov_estoque_model.dataframe.columns, "mov_estoque")
        self._aplicar_preset_colunas(self.mov_estoque_table, self.mov_estoque_model.dataframe.columns, visiveis)
        colunas = self.mov_estoque_model.dataframe.columns
        for nome, largura in {"descr_padrao": 320, "Descr_item": 320, "Tipo_operacao": 170, "id_agrupado": 140}.items():
            if nome in colunas:
                self.mov_estoque_table.setColumnWidth(colunas.index(nome), largura)

    def _aplicar_preset_aba_anual(self) -> None:
        if self.aba_anual_model.dataframe.is_empty():
            return
        visiveis = self._obter_colunas_preset_perfil("exportar", self.aba_anual_model.dataframe.columns, "aba_anual")
        self._aplicar_preset_colunas(self.aba_anual_table, self.aba_anual_model.dataframe.columns, visiveis)
        colunas = self.aba_anual_model.dataframe.columns
        offset = 1 if getattr(self.aba_anual_model, "_checkable", False) else 0
        for nome, largura in {"descr_padrao": 320, "id_agregado": 140}.items():
            if nome in colunas:
                self.aba_anual_table.setColumnWidth(colunas.index(nome) + offset, largura)

    def _aplicar_preset_aba_mensal(self) -> None:
        if self.aba_mensal_model.dataframe.is_empty():
            return
        visiveis = self._obter_colunas_preset_perfil("exportar", self.aba_mensal_model.dataframe.columns, "aba_mensal")
        self._aplicar_preset_colunas(self.aba_mensal_table, self.aba_mensal_model.dataframe.columns, visiveis)
        colunas = self.aba_mensal_model.dataframe.columns
        offset = 1 if getattr(self.aba_mensal_model, "_checkable", False) else 0
        for nome, largura in {"descr_padrao": 320, "id_agregado": 150, "unids_mes": 180, "unids_ref_mes": 180}.items():
            if nome in colunas:
                self.aba_mensal_table.setColumnWidth(colunas.index(nome) + offset, largura)

    def _aplicar_perfil_consulta(self) -> None:
        if self.table_model.dataframe.is_empty():
            return
        self._aplicar_perfil_tabela(
            "consulta",
            self.table_view,
            self.table_model,
            self.consulta_profile.currentText(),
            "consulta",
            self._consulta_scope(),
        )

    def _aplicar_perfil_agregacao(
        self,
        aba: str,
        table: QTableView,
        model: PolarsTableModel,
        perfil: str,
    ) -> None:
        if model.dataframe.is_empty():
            return
        self._aplicar_perfil_tabela(aba, table, model, perfil, aba)

    def _carregar_dataset_ui(
        self,
        path: Path,
        conditions: list[FilterCondition] | None = None,
        columns: list[str] | None = None,
    ) -> pl.DataFrame:
        colunas_solicitadas = columns
        if columns is not None:
            schema = set(self.parquet_service.get_schema(path))
            colunas_solicitadas = [coluna for coluna in columns if coluna in schema]
            if not colunas_solicitadas:
                return pl.DataFrame()
        return self.parquet_service.load_dataset(path, conditions or [], colunas_solicitadas)

    def _limpar_aba_resumo_estoque(self, contexto: str, mensagem: str) -> None:
        if contexto == "aba_mensal":
            self.aba_mensal_model.set_dataframe(pl.DataFrame())
            self._aba_mensal_df = pl.DataFrame()
            self.lbl_aba_mensal_status.setText(mensagem)
            self.lbl_aba_mensal_filtros.setText("Filtros ativos: nenhum")
            self._atualizar_titulo_aba_mensal()
            return
        if contexto == "aba_anual":
            self.aba_anual_model.set_dataframe(pl.DataFrame())
            self._aba_anual_df = pl.DataFrame()
            self.lbl_aba_anual_status.setText(mensagem)
            self.lbl_aba_anual_filtros.setText("Filtros ativos: nenhum")
            self._atualizar_titulo_aba_anual()

    def _garantir_resumos_estoque_atualizados(self, cnpj: str) -> bool:
        artefatos_defasados = self.servico_agregacao.artefatos_estoque_defasados(cnpj)
        if not artefatos_defasados:
            return True

        if self._sync_resumos_estoque_cnpj == cnpj:
            return False

        if self.service_worker is not None and self.service_worker.isRunning():
            self.status.showMessage("Aguardando o processamento atual para sincronizar as tabelas mensal/anual.")
            return False

        self._sync_resumos_estoque_cnpj = cnpj
        nomes = {
            "calculos_mensais": "mensal",
            "calculos_anuais": "anual",
        }
        descricoes = ", ".join(nomes.get(item, item) for item in artefatos_defasados)

        def _on_success(ok) -> None:
            self._sync_resumos_estoque_cnpj = None
            if ok:
                self.refresh_file_tree(cnpj)
                self.atualizar_aba_mensal()
                self.atualizar_aba_anual()
                self.atualizar_aba_produtos_selecionados()
                self.atualizar_aba_resumo_global()
                self.status.showMessage(f"Tabelas {descricoes} sincronizadas com a mov_estoque.")
            else:
                self.status.showMessage("Falha ao sincronizar as tabelas mensal/anual com a mov_estoque.")
                self.show_error("Falha na sincronizacao", "Nao foi possivel atualizar as tabelas mensal/anual.")

        def _on_failure(mensagem: str) -> None:
            self._sync_resumos_estoque_cnpj = None
            self.status.showMessage("Erro ao sincronizar as tabelas mensal/anual.")
            self.show_error("Falha na sincronizacao", mensagem)

        iniciado = self._executar_em_worker(
            self.servico_agregacao.recalcular_resumos_estoque,
            cnpj,
            mensagem_inicial=f"Sincronizando tabelas {descricoes} com a mov_estoque...",
            on_success=_on_success,
            on_failure=_on_failure,
        )
        if not iniciado:
            self._sync_resumos_estoque_cnpj = None
            return False
        return False

    def atualizar_aba_anual(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            self._atualizar_titulo_aba_anual()
            return

        if not self._garantir_resumos_estoque_atualizados(cnpj):
            self._limpar_aba_resumo_estoque("aba_anual", "Sincronizando tabela anual com a mov_estoque atual...")
            return
            
        path = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"aba_anual_{cnpj}.parquet"
        if not path.exists():
            self.aba_anual_model.set_dataframe(pl.DataFrame())
            self._aba_anual_df = pl.DataFrame()
            self.lbl_aba_anual_status.setText("Tabela Anual nao encontrada.")
            self._atualizar_titulo_aba_anual()
            self.atualizar_aba_produtos_selecionados()
            return
            
        try:
            self._aba_anual_df = self._carregar_dataset_ui(path)
            self._aba_anual_file_path = path
            self._reset_table_resize_flag("aba_anual")

            id_atual = self.anual_filter_id.currentText()
            ids = self._aba_anual_df.get_column("id_agregado").unique().sort().to_list()
            self._popular_combo_texto(self.anual_filter_id, [str(i) for i in ids], id_atual, "")

            ano_atual = self.anual_filter_ano.currentText()
            anos = self._aba_anual_df.get_column("ano").unique().sort().to_list()
            self._popular_combo_texto(self.anual_filter_ano, [str(a) for a in anos], ano_atual, "Todos")

            self.aplicar_filtros_aba_anual()
            self.atualizar_aba_produtos_selecionados()
            self.atualizar_aba_resumo_global()
        except Exception as e:
            self.show_error("Erro de leitura", f"Falha ao ler aba_anual: {e}")

    def atualizar_aba_mensal(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            self._atualizar_titulo_aba_mensal()
            return

        if not self._garantir_resumos_estoque_atualizados(cnpj):
            self._limpar_aba_resumo_estoque("aba_mensal", "Sincronizando tabela mensal com a mov_estoque atual...")
            return

        path = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"aba_mensal_{cnpj}.parquet"
        if not path.exists():
            self.aba_mensal_model.set_dataframe(pl.DataFrame())
            self._aba_mensal_df = pl.DataFrame()
            self.lbl_aba_mensal_status.setText("Tabela Mensal nao encontrada.")
            self._atualizar_titulo_aba_mensal()
            self.atualizar_aba_produtos_selecionados()
            return

        try:
            self._aba_mensal_df = self._carregar_dataset_ui(path)
            self._aba_mensal_file_path = path
            self._reset_table_resize_flag("aba_mensal")

            id_atual = self.mensal_filter_id.currentText()
            ids = self._aba_mensal_df.get_column("id_agregado").unique().sort().to_list()
            self._popular_combo_texto(self.mensal_filter_id, [str(i) for i in ids], id_atual, "")

            ano_atual = self.mensal_filter_ano.currentText()
            anos = self._aba_mensal_df.get_column("ano").unique().sort().to_list()
            self._popular_combo_texto(self.mensal_filter_ano, [str(a) for a in anos], ano_atual, "Todos")

            self.aplicar_filtros_aba_mensal()
            self.atualizar_aba_produtos_selecionados()
            self.atualizar_aba_resumo_global()
        except Exception as e:
            self.show_error("Erro de leitura", f"Falha ao ler aba_mensal: {e}")

    def aplicar_filtros_aba_mensal(self) -> None:
        if self._aba_mensal_df.is_empty():
            return
        try:
            id_agreg = self.mensal_filter_id.currentText().strip()
            desc = self.mensal_filter_desc.text().strip().lower()
            ano = self.mensal_filter_ano.currentText()
            mes = self.mensal_filter_mes.currentText()
            texto = self.mensal_filter_texto.text().strip().lower()
            num_col = self.mensal_filter_num_col.currentText().strip()
            num_min = self.mensal_filter_num_min.text().strip()
            num_max = self.mensal_filter_num_max.text().strip()

            df_filtrado = self._aba_mensal_df
            if id_agreg:
                df_filtrado = df_filtrado.filter(pl.col("id_agregado").cast(pl.Utf8).str.contains(id_agreg))
            if desc:
                df_filtrado = df_filtrado.filter(
                    pl.col("descr_padrao").cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(desc, literal=True)
                )
            if ano != "Todos":
                df_filtrado = df_filtrado.filter(pl.col("ano").cast(pl.Utf8) == ano)
            if mes != "Todos":
                df_filtrado = df_filtrado.filter(pl.col("mes").cast(pl.Utf8) == mes)

            df_filtrado = self._filtrar_intervalo_numerico(df_filtrado, num_col, num_min, num_max)
            if texto:
                df_filtrado = self._filtrar_texto_em_colunas(df_filtrado, texto)

            self.aba_mensal_model.set_dataframe(df_filtrado)
            self._resize_table_once(self.aba_mensal_table, "aba_mensal")
            if not self._aplicar_preferencias_tabela("aba_mensal", self.aba_mensal_table, self.aba_mensal_model):
                self._aplicar_ordenacao_padrao(
                    self.aba_mensal_table,
                    self.aba_mensal_model,
                    ["ano", "mes", "id_agregado", "descr_padrao"],
                )
                self._aplicar_preset_aba_mensal()
            self.lbl_aba_mensal_status.setText(
                f"Exibindo {df_filtrado.height:,} de {self._aba_mensal_df.height:,} linhas."
            )
            self.lbl_aba_mensal_filtros.setText(
                self._formatar_resumo_filtros(
                    [
                        ("id_agregado", id_agreg),
                        ("descricao", desc),
                        ("ano", "" if ano == "Todos" else ano),
                        ("mes", "" if mes == "Todos" else mes),
                        (num_col, f"{num_min or ''}..{num_max or ''}" if num_min or num_max else ""),
                        ("texto", texto),
                    ]
                )
            )
            self._atualizar_titulo_aba_mensal(df_filtrado.height, self._aba_mensal_df.height)
            self._salvar_preferencias_tabela("aba_mensal", self.aba_mensal_table, self.aba_mensal_model)
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao filtrar aba mensal: {e}")

    def limpar_filtros_aba_mensal(self) -> None:
        self.mensal_filter_id.setCurrentIndex(0)
        self.mensal_filter_desc.clear()
        self.mensal_filter_ano.setCurrentIndex(0)
        self.mensal_filter_mes.setCurrentIndex(0)
        self.mensal_filter_texto.clear()
        self.mensal_filter_num_min.clear()
        self.mensal_filter_num_max.clear()
        self.aplicar_filtros_aba_mensal()

    def exportar_aba_mensal_excel_metodo(self) -> None:
        df = self._dataframe_colunas_perfil("aba_mensal", "aba_mensal", self.aba_mensal_model, self.aba_mensal_model.dataframe, perfil="Exportar")
        if df.is_empty():
            return
        target = self._save_dialog("Exportar Mensal", "Excel (*.xlsx)")
        if not target:
            return
        try:
            self.export_service.export_excel(target, df, sheet_name="Mensal")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def exportar_aba_mensal_excel(self) -> None:
        self.exportar_aba_mensal_excel_metodo()

    def aplicar_filtros_aba_anual(self) -> None:
        if self._aba_anual_df.is_empty():
            return
        try:
            id_agreg = self.anual_filter_id.currentText().strip()
            desc = self.anual_filter_desc.text().strip().lower()
            ano = self.anual_filter_ano.currentText()
            texto = self.anual_filter_texto.text().strip().lower()
            num_col = self.anual_filter_num_col.currentText().strip()
            num_min = self.anual_filter_num_min.text().strip()
            num_max = self.anual_filter_num_max.text().strip()

            df_filtrado = self._aba_anual_df

            if self._filtro_cruzado_anuais_ids:
                df_filtrado = df_filtrado.filter(pl.col("id_agregado").is_in(self._filtro_cruzado_anuais_ids))

            if id_agreg:
                df_filtrado = df_filtrado.filter(pl.col("id_agregado").cast(pl.Utf8).str.contains(id_agreg))
            if desc:
                col_desc = "descr_padrao" if "descr_padrao" in df_filtrado.columns else ("descriCAo" if "descriCAo" in df_filtrado.columns else None)
                if col_desc is not None:
                    df_filtrado = df_filtrado.filter(
                        pl.col(col_desc).cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(desc, literal=True)
                    )
            if ano != "Todos":
                df_filtrado = df_filtrado.filter(pl.col("ano").cast(pl.Utf8) == ano)

            df_filtrado = self._filtrar_intervalo_numerico(df_filtrado, num_col, num_min, num_max)

            if texto:
                df_filtrado = self._filtrar_texto_em_colunas(df_filtrado, texto)

            self.aba_anual_model.set_dataframe(df_filtrado)
            self._resize_table_once(self.aba_anual_table, "aba_anual")
            if not self._aplicar_preferencias_tabela("aba_anual", self.aba_anual_table, self.aba_anual_model):
                self._aplicar_ordenacao_padrao(
                    self.aba_anual_table,
                    self.aba_anual_model,
                    ["ano", "id_agregado", "descr_padrao"],
                )
                self._aplicar_preset_aba_anual()
            self.lbl_aba_anual_status.setText(
                f"Exibindo {df_filtrado.height:,} de {self._aba_anual_df.height:,} linhas."
                + (" (FILTRO CRUZADO ATIVO)" if self._filtro_cruzado_anuais_ids else "")
            )
            self.lbl_aba_anual_filtros.setText(
                self._formatar_resumo_filtros(
                    [
                        ("id_agregado", id_agreg),
                        ("descricao", desc),
                        ("ano", "" if ano == "Todos" else ano),
                        (num_col, f"{num_min or ''}..{num_max or ''}" if num_min or num_max else ""),
                        ("texto", texto),
                        ("cruzado", ",".join(self._filtro_cruzado_anuais_ids[:3]) + ("..." if len(self._filtro_cruzado_anuais_ids) > 3 else "") if self._filtro_cruzado_anuais_ids else ""),
                    ]
                )
            )
            self._atualizar_titulo_aba_anual(df_filtrado.height, self._aba_anual_df.height)
            self._salvar_preferencias_tabela("aba_anual", self.aba_anual_table, self.aba_anual_model)
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao filtrar aba anual: {e}")

    def limpar_filtros_aba_anual(self) -> None:
        self.anual_filter_id.setCurrentIndex(0)
        self.anual_filter_desc.clear()
        self.anual_filter_ano.setCurrentIndex(0)
        self.anual_filter_texto.clear()
        self.anual_filter_num_min.clear()
        self.anual_filter_num_max.clear()
        self.aplicar_filtros_aba_anual()

    def filtrar_estoque_pela_selecao_anual(self) -> None:
        checked_ids = self.aba_anual_model.get_checked_rows()
        if not checked_ids:
            QMessageBox.information(self, "Aviso", "Nenhum produto selecionado.")
            return
        ids_unicos = list(set([str(r.get("id_agregado", "")) for r in checked_ids if r.get("id_agregado")]))
        self._filtro_cruzado_anuais_ids = ids_unicos
        self.aplicar_filtros_aba_anual()
        self.aplicar_filtros_mov_estoque()

    def limpar_filtro_cruzado_anual(self) -> None:
        self._filtro_cruzado_anuais_ids = []
        self.aba_anual_model.clear_checked()
        self.aplicar_filtros_aba_anual()
        self.aplicar_filtros_mov_estoque()

    def exportar_aba_anual_excel_metodo(self) -> None:
        df = self._dataframe_colunas_perfil("aba_anual", "aba_anual", self.aba_anual_model, self.aba_anual_model.dataframe, perfil="Exportar")
        if df.is_empty(): return
        target = self._save_dialog(f"Exportar Anual", "Excel (*.xlsx)")
        if not target: return
        try:
            self.export_service.export_excel(target, df, sheet_name="Anual")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def exportar_aba_anual_excel(self) -> None:
        self.exportar_aba_anual_excel_metodo()

    def exportar_mov_estoque_excel(self) -> None:
        if self.mov_estoque_model.dataframe.is_empty():
            QMessageBox.information(self, "Exportacao", "Nao ha dados filtrados na mov_estoque para exportar.")
            return
        target = self._save_dialog("Exportar Movimentacao de Estoque", "Excel (*.xlsx)")
        if not target:
            return
        try:
            df_to_export = self._dataframe_colunas_visiveis(
                self.mov_estoque_table,
                self.mov_estoque_model,
                self.mov_estoque_model.dataframe,
            )
            self.export_service.export_excel(target, df_to_export, sheet_name="Mov_Estoque")
            # self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}") # Export service shows its own message/bar maybe? I'll add one just in case 
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def atualizar_aba_nfe_entrada(self) -> None:
        self._atualizar_estado_botao_nfe_entrada()
        cnpj = self.state.current_cnpj
        if not cnpj:
            self.nfe_entrada_model.set_dataframe(pl.DataFrame())
            self._nfe_entrada_df = pl.DataFrame()
            self.lbl_nfe_entrada_status.setText("Selecione um CPF/CNPJ para carregar as NFes/NFCes de entrada.")
            return

        path_nfe = CNPJ_ROOT / cnpj / "arquivos_parquet" / f"nfe_agr_{cnpj}.parquet"
        path_nfce = CNPJ_ROOT / cnpj / "arquivos_parquet" / f"nfce_agr_{cnpj}.parquet"
        if not path_nfe.exists() and not path_nfce.exists():
            self.nfe_entrada_model.set_dataframe(pl.DataFrame())
            self._nfe_entrada_df = pl.DataFrame()
            base_nfe = CNPJ_ROOT / cnpj / "arquivos_parquet" / f"nfe_{cnpj}.parquet"
            base_nfce = CNPJ_ROOT / cnpj / "arquivos_parquet" / f"nfce_{cnpj}.parquet"
            if base_nfe.exists() or base_nfce.exists():
                self.lbl_nfe_entrada_status.setText("Arquivos 'nfe_agr'/'nfce_agr' ainda nao foram gerados. Clique em Extrair para preparar a tabela.")
            else:
                self.lbl_nfe_entrada_status.setText("Arquivos 'nfe_agr'/'nfce_agr' nao encontrados para este CPF/CNPJ.")
            return

        try:
            from transformacao.co_sefin import inferir_co_sefin_dataframe
            from transformacao.co_sefin_class import enriquecer_co_sefin_class

            df_partes: list[pl.DataFrame] = []
            if path_nfe.exists():
                df_partes.append(
                    pl.read_parquet(path_nfe).with_columns(pl.lit("NFe").alias("fonte_documento"))
                )
            if path_nfce.exists():
                df_partes.append(
                    pl.read_parquet(path_nfce).with_columns(pl.lit("NFCe").alias("fonte_documento"))
                )
            df_nfe = pl.concat(df_partes, how="diagonal_relaxed") if df_partes else pl.DataFrame()
            self._nfe_entrada_file_path = path_nfe if path_nfe.exists() else path_nfce

            if "tipo_operacao" in df_nfe.columns:
                df_nfe = df_nfe.filter(
                    pl.col("tipo_operacao").cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains("entrada", literal=True)
                )
            elif "co_tp_nf" in df_nfe.columns:
                df_nfe = df_nfe.filter(pl.col("co_tp_nf").cast(pl.Int64, strict=False) == 0)

            if df_nfe.is_empty():
                self._nfe_entrada_df = pl.DataFrame()
                self.nfe_entrada_model.set_dataframe(pl.DataFrame())
                self.lbl_nfe_entrada_status.setText("Nenhuma NFe/NFCe de entrada foi encontrada.")
                return

            for col in ["dhemi", "dhsaient", "prod_cest"]:
                if col not in df_nfe.columns:
                    df_nfe = df_nfe.with_columns(pl.lit(None).alias(col))

            df_nfe = df_nfe.with_columns(
                [
                    pl.col("dhemi").cast(pl.Datetime, strict=False).dt.date().alias("__dhemi_date__"),
                    pl.col("dhsaient").cast(pl.Datetime, strict=False).dt.date().alias("__dhsaient_date__"),
                ]
            ).with_columns(
                pl.coalesce(
                    [
                        pl.max_horizontal(pl.col("__dhemi_date__"), pl.col("__dhsaient_date__")),
                        pl.col("__dhsaient_date__"),
                        pl.col("__dhemi_date__"),
                    ]
                ).alias("data_classificacao")
            )

            df_class = inferir_co_sefin_dataframe(df_nfe, col_ncm="prod_ncm", col_cest="prod_cest", output_col="co_sefin_inferido")
            df_enriquecer = df_class.with_columns(
                [
                    pl.col("prod_ncm").cast(pl.Utf8, strict=False).alias("ncm_padrao"),
                    pl.col("prod_cest").cast(pl.Utf8, strict=False).alias("cest_padrao"),
                    pl.col("__dhemi_date__").alias("Dt_doc"),
                    pl.col("__dhsaient_date__").alias("Dt_e_s"),
                ]
            )
            df_enriquecido = enriquecer_co_sefin_class(df_enriquecer)
            df_enriquecido = df_enriquecido.drop(
                ["__dhemi_date__", "__dhsaient_date__", "Dt_doc", "Dt_e_s", "ncm_padrao", "cest_padrao"],
                strict=False,
            )

            colunas = list(df_enriquecido.columns)
            prioridade = [
                "data_classificacao",
                "fonte_documento",
                "tipo_operacao",
                "nnf",
                "prod_nitem",
                "id_agrupado",
                "descr_padrao",
                "prod_xprod",
                "prod_ncm",
                "prod_cest",
                "co_sefin_inferido",
                "co_sefin_agr",
                "it_pc_interna",
                "it_in_st",
                "it_in_isento_icms",
                "it_in_reducao",
                "it_pc_reducao",
                "it_pc_mva",
                "it_in_mva_ajustado",
                "xnome_emit",
                "xnome_dest",
                "chave_acesso",
            ]
            ordenadas = [c for c in prioridade if c in colunas] + [c for c in colunas if c not in prioridade]
            self._nfe_entrada_df = df_enriquecido.select(ordenadas)

            id_atual = self.nfe_entrada_filter_id.currentText()
            ids = (
                self._nfe_entrada_df.get_column("id_agrupado").cast(pl.Utf8, strict=False).drop_nulls().unique().sort().to_list()
                if "id_agrupado" in self._nfe_entrada_df.columns
                else []
            )
            self._popular_combo_texto(self.nfe_entrada_filter_id, [str(i) for i in ids], id_atual, "")

            self.aplicar_filtros_nfe_entrada()
        except Exception as e:
            self.show_error("Erro de leitura", f"Falha ao montar a aba NFe Entrada: {e}")

    def aplicar_filtros_nfe_entrada(self) -> None:
        if self._nfe_entrada_df.is_empty():
            return
        try:
            id_agrupado = self.nfe_entrada_filter_id.currentText().strip()
            desc = self.nfe_entrada_filter_desc.text().strip().lower()
            ncm = self.nfe_entrada_filter_ncm.text().strip()
            co_sefin = self.nfe_entrada_filter_sefin.text().strip()
            texto = self.nfe_entrada_filter_texto.text().strip().lower()
            data_ini = None if self.nfe_entrada_filter_data_ini.date() == self.nfe_entrada_filter_data_ini.minimumDate() else self.nfe_entrada_filter_data_ini.date()
            data_fim = None if self.nfe_entrada_filter_data_fim.date() == self.nfe_entrada_filter_data_fim.minimumDate() else self.nfe_entrada_filter_data_fim.date()

            df_filtrado = self._nfe_entrada_df
            if id_agrupado and "id_agrupado" in df_filtrado.columns:
                df_filtrado = df_filtrado.filter(pl.col("id_agrupado").cast(pl.Utf8, strict=False).fill_null("").str.contains(id_agrupado, literal=True))
            if desc:
                cols_desc = [c for c in ["descr_padrao", "prod_xprod"] if c in df_filtrado.columns]
                if cols_desc:
                    exprs = [
                        pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(desc, literal=True)
                        for col in cols_desc
                    ]
                    df_filtrado = df_filtrado.filter(pl.any_horizontal(exprs))
            if ncm and "prod_ncm" in df_filtrado.columns:
                df_filtrado = df_filtrado.filter(pl.col("prod_ncm").cast(pl.Utf8, strict=False).fill_null("").str.contains(ncm, literal=True))
            if co_sefin:
                cols_sefin = [c for c in ["co_sefin_agr", "co_sefin_inferido"] if c in df_filtrado.columns]
                if cols_sefin:
                    exprs = [
                        pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.contains(co_sefin, literal=True)
                        for col in cols_sefin
                    ]
                    df_filtrado = df_filtrado.filter(pl.any_horizontal(exprs))

            df_filtrado = self._filtrar_intervalo_data(df_filtrado, "data_classificacao", data_ini, data_fim)

            if texto:
                df_filtrado = self._filtrar_texto_em_colunas(df_filtrado, texto)

            self.nfe_entrada_model.set_dataframe(df_filtrado)
            self._resize_table_once(self.nfe_entrada_table, "nfe_entrada")
            if not self._aplicar_preferencias_tabela("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model):
                self._aplicar_ordenacao_padrao(self.nfe_entrada_table, self.nfe_entrada_model, ["data_classificacao", "nnf", "prod_nitem"], Qt.DescendingOrder)
                self._aplicar_preset_colunas(
                    self.nfe_entrada_table,
                    self.nfe_entrada_model.dataframe.columns,
                    self._obter_colunas_preset_perfil("auditoria", self.nfe_entrada_model.dataframe.columns, "nfe_entrada"),
                )
            self.lbl_nfe_entrada_status.setText(f"Exibindo {df_filtrado.height:,} de {self._nfe_entrada_df.height:,} itens de NFe/NFCe de entrada.")
            periodo = ""
            if data_ini is not None or data_fim is not None:
                periodo = f"{data_ini.toString('dd/MM/yyyy') if data_ini is not None else '...'} ate {data_fim.toString('dd/MM/yyyy') if data_fim is not None else '...'}"
            self.lbl_nfe_entrada_filtros.setText(
                self._formatar_resumo_filtros(
                    [
                        ("id_agrupado", id_agrupado),
                        ("descricao", desc),
                        ("ncm", ncm),
                        ("co_sefin", co_sefin),
                        ("periodo", periodo),
                        ("texto", texto),
                    ]
                )
            )
            self._salvar_preferencias_tabela("nfe_entrada", self.nfe_entrada_table, self.nfe_entrada_model)
        except Exception as e:
            self.show_error("Erro", f"Erro ao filtrar NFe Entrada: {e}")

    def limpar_filtros_nfe_entrada(self) -> None:
        self.nfe_entrada_filter_id.setCurrentIndex(0)
        self.nfe_entrada_filter_desc.clear()
        self.nfe_entrada_filter_ncm.clear()
        self.nfe_entrada_filter_sefin.clear()
        self.nfe_entrada_filter_texto.clear()
        self.nfe_entrada_filter_data_ini.setDate(self.nfe_entrada_filter_data_ini.minimumDate())
        self.nfe_entrada_filter_data_fim.setDate(self.nfe_entrada_filter_data_fim.minimumDate())
        self.aplicar_filtros_nfe_entrada()

    def exportar_nfe_entrada_excel(self) -> None:
        df = self._dataframe_colunas_perfil("nfe_entrada", "nfe_entrada", self.nfe_entrada_model, self.nfe_entrada_model.dataframe, perfil="Exportar")
        if df.is_empty():
            self.show_info("Exportacao", "Nao ha dados de NFe Entrada para exportar.")
            return
        target = self._save_dialog("Exportar NFe Entrada", "Excel (*.xlsx)")
        if not target:
            return
        try:
            self.export_service.export_excel(target, df, sheet_name="NFe_Entrada")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def atualizar_aba_id_agrupados(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            self.id_agrupados_model.set_dataframe(pl.DataFrame())
            self._id_agrupados_df = pl.DataFrame()
            self.lbl_id_agrupados_status.setText("Selecione um CPF/CNPJ para carregar os id_agrupados.")
            self._atualizar_titulo_aba_id_agrupados()
            return

        path = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"id_agrupados_{cnpj}.parquet"
        if not path.exists():
            self.id_agrupados_model.set_dataframe(pl.DataFrame())
            self._id_agrupados_df = pl.DataFrame()
            self.lbl_id_agrupados_status.setText("Arquivo 'id_agrupados' nao encontrado para este CPF/CNPJ.")
            self._atualizar_titulo_aba_id_agrupados()
            return

        try:
            self._id_agrupados_df = pl.read_parquet(path)
            self._id_agrupados_file_path = path
            self._reset_table_resize_flag("id_agrupados")

            id_atual = self.id_agrupados_filter_id.currentText()
            ids = (
                self._id_agrupados_df.get_column("id_agrupado").cast(pl.Utf8, strict=False).drop_nulls().unique().sort().to_list()
                if "id_agrupado" in self._id_agrupados_df.columns
                else []
            )
            self._popular_combo_texto(self.id_agrupados_filter_id, [str(i) for i in ids], id_atual, "")
            self.aplicar_filtros_id_agrupados()
        except Exception as e:
            self.show_error("Erro de leitura", f"Falha ao ler id_agrupados: {e}")

    def aplicar_filtros_id_agrupados(self) -> None:
        if self._id_agrupados_df.is_empty():
            return
        try:
            id_agrupado = self.id_agrupados_filter_id.currentText().strip()
            texto = self.id_agrupados_filter_texto.text().strip().lower()

            df_filtrado = self._id_agrupados_df
            if id_agrupado and "id_agrupado" in df_filtrado.columns:
                df_filtrado = df_filtrado.filter(pl.col("id_agrupado").cast(pl.Utf8, strict=False).fill_null("").str.contains(id_agrupado, literal=True))
            if texto:
                exprs = []
                for col in df_filtrado.columns:
                    dtype = df_filtrado.schema.get(col)
                    if dtype in [pl.Utf8, pl.Categorical]:
                        exprs.append(
                            pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(texto, literal=True)
                        )
                    elif isinstance(dtype, pl.List):
                        exprs.append(
                            pl.col(col)
                            .list.join(" | ", ignore_nulls=True)
                            .cast(pl.Utf8, strict=False)
                            .fill_null("")
                            .str.to_lowercase()
                            .str.contains(texto, literal=True)
                        )
                if exprs:
                    df_filtrado = df_filtrado.filter(pl.any_horizontal(exprs))

            self.id_agrupados_model.set_dataframe(df_filtrado)
            self._resize_table_once(self.id_agrupados_table, "id_agrupados")
            if not self._aplicar_preferencias_tabela("id_agrupados", self.id_agrupados_table, self.id_agrupados_model):
                self._aplicar_ordenacao_padrao(self.id_agrupados_table, self.id_agrupados_model, ["id_agrupado"], Qt.AscendingOrder)
                self._aplicar_preset_colunas(
                    self.id_agrupados_table,
                    self.id_agrupados_model.dataframe.columns,
                    self._obter_colunas_preset_perfil("auditoria", self.id_agrupados_model.dataframe.columns, "id_agrupados"),
                )
            self.lbl_id_agrupados_status.setText(f"Exibindo {df_filtrado.height:,} de {self._id_agrupados_df.height:,} grupos consolidados.")
            self.lbl_id_agrupados_filtros.setText(
                self._formatar_resumo_filtros(
                    [
                        ("id_agrupado", id_agrupado),
                        ("texto", texto),
                    ]
                )
            )
            self._salvar_preferencias_tabela("id_agrupados", self.id_agrupados_table, self.id_agrupados_model)
            self._atualizar_titulo_aba_id_agrupados(df_filtrado.height, self._id_agrupados_df.height)
        except Exception as e:
            self.show_error("Erro", f"Erro ao filtrar id_agrupados: {e}")

    def limpar_filtros_id_agrupados(self) -> None:
        self.id_agrupados_filter_id.setCurrentIndex(0)
        self.id_agrupados_filter_texto.clear()
        self.aplicar_filtros_id_agrupados()

    def exportar_id_agrupados_excel(self) -> None:
        df = self._dataframe_colunas_perfil("id_agrupados", "id_agrupados", self.id_agrupados_model, self.id_agrupados_model.dataframe, perfil="Exportar")
        if df.is_empty():
            self.show_info("Exportacao", "Nao ha dados de id_agrupados para exportar.")
            return
        target = self._save_dialog("Exportar id_agrupados", "Excel (*.xlsx)")
        if not target:
            return
        try:
            wb = Workbook()
            ws_id_agrupados = wb.active
            ws_id_agrupados.title = "id_agrupados"
            self._escrever_planilha_openpyxl(ws_id_agrupados, df)

            df_produtos_sel = self._dataframe_colunas_visiveis(
                self.produtos_sel_table,
                self.produtos_selecionados_model,
            )
            if not df_produtos_sel.is_empty():
                self._escrever_planilha_openpyxl(
                    wb.create_sheet("produtos_selecionados"),
                    df_produtos_sel,
                )

            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target)
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def atualizar_aba_produtos_selecionados(self) -> None:
        if not self.state.current_cnpj:
            self.produtos_selecionados_model.set_dataframe(pl.DataFrame())
            self._produtos_selecionados_df = pl.DataFrame()
            self._produtos_selecionados_mov_df = pl.DataFrame()
            self._produtos_selecionados_mensal_df = pl.DataFrame()
            self._produtos_selecionados_anual_df = pl.DataFrame()
            self.lbl_produtos_sel_status.setText("Selecione um CNPJ para consolidar os produtos analisados.")
            self.lbl_produtos_sel_resumo.setText("Recorte atual: mov_estoque 0 | mensal 0 | anual 0")
            self._atualizar_titulo_aba_produtos_selecionados()
            return

        try:
            df_produtos = self._coletar_base_produtos_selecionados()
            self._reset_table_resize_flag("produtos_selecionados")

            id_atual = self.produtos_sel_filter_id.currentText()
            ids = df_produtos.get_column("id_agregado").cast(pl.Utf8, strict=False).drop_nulls().unique().sort().to_list() if "id_agregado" in df_produtos.columns else []
            self._popular_combo_texto(self.produtos_sel_filter_id, [str(i) for i in ids], id_atual, "")

            anos = self._anos_disponiveis_produtos_selecionados()
            ano_ini_atual = self.produtos_sel_filter_ano_ini.currentText()
            ano_fim_atual = self.produtos_sel_filter_ano_fim.currentText()
            anos_texto = [str(a) for a in anos]
            self._popular_combo_texto(self.produtos_sel_filter_ano_ini, anos_texto, ano_ini_atual, "Todos")
            self._popular_combo_texto(self.produtos_sel_filter_ano_fim, anos_texto, ano_fim_atual, "Todos")

            self.aplicar_filtros_produtos_selecionados()
        except Exception as e:
            self.show_error("Erro de leitura", f"Falha ao consolidar produtos selecionados: {e}")

    def _coletar_base_produtos_selecionados(self) -> pl.DataFrame:
        bases: list[pl.DataFrame] = []
        if not self._aba_mensal_df.is_empty() and {"id_agregado", "descr_padrao"}.issubset(set(self._aba_mensal_df.columns)):
            bases.append(self._aba_mensal_df.select(["id_agregado", "descr_padrao"]))
        if not self._aba_anual_df.is_empty() and {"id_agregado", "descr_padrao"}.issubset(set(self._aba_anual_df.columns)):
            bases.append(self._aba_anual_df.select(["id_agregado", "descr_padrao"]))
        if not self._mov_estoque_df.is_empty():
            col_id = "id_agregado" if "id_agregado" in self._mov_estoque_df.columns else ("id_agrupado" if "id_agrupado" in self._mov_estoque_df.columns else None)
            col_desc = "descr_padrao" if "descr_padrao" in self._mov_estoque_df.columns else None
            if col_id and col_desc:
                bases.append(
                    self._mov_estoque_df.select(
                        [
                            pl.col(col_id).cast(pl.Utf8, strict=False).alias("id_agregado"),
                            pl.col(col_desc).cast(pl.Utf8, strict=False).alias("descr_padrao"),
                        ]
                    )
                )
        if not bases:
            return pl.DataFrame({"id_agregado": [], "descr_padrao": []}, schema={"id_agregado": pl.Utf8, "descr_padrao": pl.Utf8})
        return pl.concat(bases, how="vertical_relaxed").unique(subset=["id_agregado"]).sort("id_agregado")

    def _anos_disponiveis_produtos_selecionados(self) -> list[int]:
        anos: set[int] = set()
        for df in (self._aba_mensal_df, self._aba_anual_df):
            if not df.is_empty() and "ano" in df.columns:
                try:
                    anos.update(int(a) for a in df.get_column("ano").drop_nulls().unique().to_list())
                except Exception:
                    pass
        if not anos and not self._mov_estoque_df.is_empty():
            for col in ("Dt_e_s", "Dt_doc"):
                if col in self._mov_estoque_df.columns:
                    try:
                        serie = self._mov_estoque_df.get_column(col).cast(pl.Date, strict=False).drop_nulls().dt.year().unique().to_list()
                        anos.update(int(a) for a in serie if a is not None)
                    except Exception:
                        pass
        return sorted(anos)

    def _intervalo_anos_produtos_selecionados(self) -> tuple[int | None, int | None]:
        ano_ini_txt = self.produtos_sel_filter_ano_ini.currentText().strip()
        ano_fim_txt = self.produtos_sel_filter_ano_fim.currentText().strip()
        ano_ini = int(ano_ini_txt) if ano_ini_txt and ano_ini_txt != "Todos" else None
        ano_fim = int(ano_fim_txt) if ano_fim_txt and ano_fim_txt != "Todos" else None
        if ano_ini is not None and ano_fim is not None and ano_ini > ano_fim:
            ano_ini, ano_fim = ano_fim, ano_ini
        return ano_ini, ano_fim

    def _intervalo_datas_produtos_selecionados(self) -> tuple[QDate | None, QDate | None]:
        data_ini = self._valor_qdate_ativo(self.produtos_sel_filter_data_ini.date())
        data_fim = self._valor_qdate_ativo(self.produtos_sel_filter_data_fim.date())
        if data_ini is not None and data_fim is not None and data_ini > data_fim:
            data_ini, data_fim = data_fim, data_ini
        return data_ini, data_fim

    def _filtrar_dataframe_por_ids(self, df: pl.DataFrame, ids: list[str]) -> pl.DataFrame:
        if df.is_empty() or not ids:
            return df
        if "id_agregado" in df.columns:
            return df.filter(pl.col("id_agregado").cast(pl.Utf8, strict=False).is_in(ids))
        if "id_agrupado" in df.columns:
            return df.filter(pl.col("id_agrupado").cast(pl.Utf8, strict=False).is_in(ids))
        return df

    def _filtrar_dataframe_por_ano(self, df: pl.DataFrame, ano_ini: int | None, ano_fim: int | None) -> pl.DataFrame:
        if df.is_empty() or (ano_ini is None and ano_fim is None):
            return df
        if "ano" in df.columns:
            ano_expr = pl.col("ano").cast(pl.Int32, strict=False)
            if ano_ini is not None:
                df = df.filter(ano_expr >= ano_ini)
            if ano_fim is not None:
                df = df.filter(ano_expr <= ano_fim)
            return df
        data_col = None
        for col in ("Dt_e_s", "Dt_doc"):
            if col in df.columns:
                data_col = col
                break
        if data_col is None:
            return df
        ano_expr = pl.col(data_col).cast(pl.Date, strict=False).dt.year()
        if ano_ini is not None:
            df = df.filter(ano_expr >= ano_ini)
        if ano_fim is not None:
            df = df.filter(ano_expr <= ano_fim)
        return df

    def _filtrar_dataframe_produtos_selecionados_por_data(
        self,
        df: pl.DataFrame,
        data_ini: QDate | None,
        data_fim: QDate | None,
        tipo_base: str,
    ) -> pl.DataFrame:
        if df.is_empty() or (data_ini is None and data_fim is None):
            return df

        if tipo_base == "mensal" and {"ano", "mes"}.issubset(set(df.columns)):
            df_tmp = df.with_columns(
                pl.concat_str(
                    [
                        pl.col("ano").cast(pl.Int32, strict=False).cast(pl.Utf8),
                        pl.lit("-"),
                        pl.col("mes").cast(pl.Int32, strict=False).cast(pl.Utf8).str.zfill(2),
                        pl.lit("-01"),
                    ]
                ).str.strptime(pl.Date, "%Y-%m-%d", strict=False).alias("__data_ref_filtro__")
            )
            df_tmp = self._filtrar_intervalo_data(df_tmp, "__data_ref_filtro__", data_ini, data_fim)
            return df_tmp.drop("__data_ref_filtro__", strict=False)

        if tipo_base == "anual" and "ano" in df.columns:
            df_tmp = df.with_columns(
                pl.concat_str(
                    [
                        pl.col("ano").cast(pl.Int32, strict=False).cast(pl.Utf8),
                        pl.lit("-12-31"),
                    ]
                ).str.strptime(pl.Date, "%Y-%m-%d", strict=False).alias("__data_ref_filtro__")
            )
            df_tmp = self._filtrar_intervalo_data(df_tmp, "__data_ref_filtro__", data_ini, data_fim)
            return df_tmp.drop("__data_ref_filtro__", strict=False)

        return self._filtrar_intervalo_data(df, "Dt_e_s" if "Dt_e_s" in df.columns else "Dt_doc", data_ini, data_fim)

    def _ids_produtos_selecionados_para_exportacao(self) -> list[str]:
        checked = self.produtos_selecionados_model.get_checked_rows()
        ids = [str(r.get("id_agregado") or "").strip() for r in checked if r.get("id_agregado")]
        if ids:
            return sorted(set(ids))
        return []

    def aplicar_filtros_produtos_selecionados(self) -> None:
        try:
            base = self._coletar_base_produtos_selecionados()
            if base.is_empty():
                self.produtos_selecionados_model.set_dataframe(pl.DataFrame())
                self._produtos_selecionados_df = pl.DataFrame()
                self._produtos_selecionados_mov_df = pl.DataFrame()
                self._produtos_selecionados_mensal_df = pl.DataFrame()
                self._produtos_selecionados_anual_df = pl.DataFrame()
                self.lbl_produtos_sel_status.setText("Nenhum dado de estoque/mensal/anual foi encontrado para consolidacao.")
                self.lbl_produtos_sel_resumo.setText("Recorte atual: mov_estoque 0 | mensal 0 | anual 0")
                self._atualizar_titulo_aba_produtos_selecionados(0, 0)
                return

            id_agregado = self.produtos_sel_filter_id.currentText().strip()
            desc = self.produtos_sel_filter_desc.text().strip().lower()
            texto = self.produtos_sel_filter_texto.text().strip().lower()
            ano_ini, ano_fim = self._intervalo_anos_produtos_selecionados()
            data_ini, data_fim = self._intervalo_datas_produtos_selecionados()

            df_produtos = base
            if id_agregado:
                df_produtos = df_produtos.filter(pl.col("id_agregado").cast(pl.Utf8, strict=False).fill_null("").str.contains(id_agregado, literal=True))
            if desc:
                df_produtos = df_produtos.filter(
                    pl.col("descr_padrao").cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(desc, literal=True)
                )
            if texto:
                df_produtos = self._filtrar_texto_em_colunas(df_produtos, texto)

            ids_filtrados = (
                df_produtos.get_column("id_agregado").cast(pl.Utf8, strict=False).drop_nulls().unique().sort().to_list()
                if "id_agregado" in df_produtos.columns else []
            )

            df_mensal = self._filtrar_dataframe_por_ids(self._aba_mensal_df, ids_filtrados)
            df_mensal = self._filtrar_dataframe_por_ano(df_mensal, ano_ini, ano_fim)
            df_mensal = self._filtrar_dataframe_produtos_selecionados_por_data(df_mensal, data_ini, data_fim, "mensal")
            if not df_mensal.is_empty() and {"id_agregado", "descr_padrao"}.issubset(set(df_mensal.columns)):
                resumo_mensal = (
                    df_mensal.group_by(["id_agregado", "descr_padrao"])
                    .agg(pl.col("ICMS_entr_desacob").cast(pl.Float64, strict=False).fill_null(0).sum().alias("total_ICMS_entr_desacob"))
                )
            else:
                resumo_mensal = pl.DataFrame(
                    schema={
                        "id_agregado": pl.Utf8,
                        "descr_padrao": pl.Utf8,
                        "total_ICMS_entr_desacob": pl.Float64,
                    }
                )

            df_anual = self._filtrar_dataframe_por_ids(self._aba_anual_df, ids_filtrados)
            df_anual = self._filtrar_dataframe_por_ano(df_anual, ano_ini, ano_fim)
            df_anual = self._filtrar_dataframe_produtos_selecionados_por_data(df_anual, data_ini, data_fim, "anual")
            if not df_anual.is_empty() and {"id_agregado", "descr_padrao"}.issubset(set(df_anual.columns)):
                resumo_anual = (
                    df_anual.group_by(["id_agregado", "descr_padrao"])
                    .agg(
                        [
                            pl.col("ICMS_saidas_desac").cast(pl.Float64, strict=False).fill_null(0).sum().alias("total_ICMS_saidas_desac"),
                            pl.col("ICMS_estoque_desac").cast(pl.Float64, strict=False).fill_null(0).sum().alias("total_ICMS_estoque_desac"),
                        ]
                    )
                )
            else:
                resumo_anual = pl.DataFrame(
                    schema={
                        "id_agregado": pl.Utf8,
                        "descr_padrao": pl.Utf8,
                        "total_ICMS_saidas_desac": pl.Float64,
                        "total_ICMS_estoque_desac": pl.Float64,
                    }
                )

            resumo = (
                df_produtos.join(resumo_mensal, on=["id_agregado", "descr_padrao"], how="left")
                .join(resumo_anual, on=["id_agregado", "descr_padrao"], how="left")
                .with_columns(
                    [
                        pl.col("total_ICMS_entr_desacob").cast(pl.Float64, strict=False).fill_null(0).round(2),
                        pl.col("total_ICMS_saidas_desac").cast(pl.Float64, strict=False).fill_null(0).round(2),
                        pl.col("total_ICMS_estoque_desac").cast(pl.Float64, strict=False).fill_null(0).round(2),
                        (
                            pl.col("total_ICMS_entr_desacob").cast(pl.Float64, strict=False).fill_null(0)
                            + pl.col("total_ICMS_saidas_desac").cast(pl.Float64, strict=False).fill_null(0)
                            + pl.col("total_ICMS_estoque_desac").cast(pl.Float64, strict=False).fill_null(0)
                        ).round(2).alias("total_ICMS_total"),
                    ]
                )
                .sort(["descr_padrao", "id_agregado"], nulls_last=True)
            )

            self._produtos_selecionados_df = resumo
            self._produtos_selecionados_mensal_df = df_mensal
            self._produtos_selecionados_anual_df = df_anual
            self._produtos_selecionados_mov_df = self._filtrar_dataframe_por_ano(
                self._filtrar_dataframe_por_ids(self._mov_estoque_df, ids_filtrados),
                ano_ini,
                ano_fim,
            )
            self._produtos_selecionados_mov_df = self._filtrar_dataframe_produtos_selecionados_por_data(
                self._produtos_selecionados_mov_df,
                data_ini,
                data_fim,
                "mov",
            )

            self.produtos_selecionados_model.set_dataframe(resumo)
            if self.state.current_cnpj and self._produtos_sel_preselecionado_cnpj != self.state.current_cnpj:
                top_ids = (
                    resumo
                    .sort(["total_ICMS_total", "id_agregado"], descending=[True, False], nulls_last=True)
                    .head(20)
                    .get_column("id_agregado")
                    .cast(pl.Utf8, strict=False)
                    .drop_nulls()
                    .to_list()
                ) if "id_agregado" in resumo.columns else []
                self.produtos_selecionados_model.set_checked_keys(
                    {(str(item_id),) for item_id in top_ids if item_id is not None}
                )
                self._produtos_sel_preselecionado_cnpj = self.state.current_cnpj
            self._resize_table_once(self.produtos_sel_table, "produtos_selecionados")
            self._aplicar_preferencias_tabela("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model)

            self.lbl_produtos_sel_status.setText(
                f"Exibindo {resumo.height:,} produtos consolidados para o periodo selecionado."
            )
            periodo = ""
            if data_ini is not None or data_fim is not None:
                periodo = f"{data_ini.toString('dd/MM/yyyy') if data_ini is not None else '...'} ate {data_fim.toString('dd/MM/yyyy') if data_fim is not None else '...'}"
            elif ano_ini is not None or ano_fim is not None:
                periodo = f"{ano_ini or '...'} ate {ano_fim or '...'}"
            self.lbl_produtos_sel_filtros.setText(
                self._formatar_resumo_filtros(
                    [
                        ("id_agregado", id_agregado),
                        ("descricao", desc),
                        ("periodo", periodo),
                        ("texto", texto),
                    ]
                )
            )
            self.lbl_produtos_sel_resumo.setText(
                f"Recorte atual: mov_estoque {self._produtos_selecionados_mov_df.height:,} | mensal {self._produtos_selecionados_mensal_df.height:,} | anual {self._produtos_selecionados_anual_df.height:,}"
            )
            self._atualizar_titulo_aba_produtos_selecionados(resumo.height, base.height)
            self._salvar_preferencias_tabela("produtos_selecionados", self.produtos_sel_table, self.produtos_selecionados_model)
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao consolidar produtos selecionados: {e}")

    def limpar_filtros_produtos_selecionados(self) -> None:
        self.produtos_sel_filter_id.setCurrentIndex(0)
        self.produtos_sel_filter_desc.clear()
        self.produtos_sel_filter_ano_ini.setCurrentIndex(0)
        self.produtos_sel_filter_ano_fim.setCurrentIndex(0)
        self.produtos_sel_filter_data_ini.setDate(self.produtos_sel_filter_data_ini.minimumDate())
        self.produtos_sel_filter_data_fim.setDate(self.produtos_sel_filter_data_fim.minimumDate())
        self.produtos_sel_filter_texto.clear()
        self.aplicar_filtros_produtos_selecionados()

    def _escrever_planilha_openpyxl(self, ws, df: pl.DataFrame) -> None:
        fonte_padrao = OpenPyxlFont(name="Arial", size=8)
        fonte_header = OpenPyxlFont(name="Arial", size=8, bold=True)
        formato_num = "#,##0.00"
        formato_inteiro = "#,##0"
        formato_ano = "0"
        formato_data = "dd/mm/yyyy"
        formato_data_hora = "dd/mm/yyyy hh:mm:ss"

        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = fonte_header

        for row in df.iter_rows():
            linha_excel = []
            for coluna, valor in zip(df.columns, row):
                if is_excel_text_identifier_column_name(coluna):
                    linha_excel.append(formatar_identificador_excel_texto(valor))
                elif hasattr(valor, "hour"):
                    linha_excel.append(valor)
                elif hasattr(valor, "day") and hasattr(valor, "month"):
                    linha_excel.append(valor)
                elif (
                    isinstance(valor, str)
                    and (
                        is_excel_datetime_column_name(coluna)
                        or is_excel_date_column_name(coluna)
                    )
                ):
                    valor_data = parse_data_iso_texto(valor)
                    linha_excel.append(valor_data if valor_data is not None else display_cell(valor, coluna))
                elif isinstance(valor, (int, float)) and not isinstance(valor, bool):
                    linha_excel.append(valor)
                else:
                    linha_excel.append(display_cell(valor, coluna))
            ws.append(linha_excel)

        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.font = fonte_header
                else:
                    cell.font = fonte_padrao
                if cell.row == 1:
                    continue
                nome_coluna = str(ws.cell(row=1, column=cell.column).value or "")
                if is_year_column_name(nome_coluna) and isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = formato_ano
                elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    if float(cell.value).is_integer():
                        cell.number_format = formato_inteiro
                    else:
                        cell.number_format = formato_num
                elif hasattr(cell.value, "hour"):
                    cell.number_format = formato_data_hora
                elif hasattr(cell.value, "day") and hasattr(cell.value, "month"):
                    cell.number_format = formato_data

        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

    def _gerar_resumo_global(self, mensal: pl.DataFrame, anual: pl.DataFrame, anos_base: list[int] | None = None) -> pl.DataFrame:
        if anos_base is None:
            anos_base = []
            for df in (mensal, anual):
                if not df.is_empty() and "ano" in df.columns:
                    anos_base.extend(
                        df.get_column("ano").cast(pl.Int32, strict=False).drop_nulls().unique().sort().to_list()
                    )
            anos_base = sorted({int(ano) for ano in anos_base})

        competencias = [f"{ano:04d}-{mes:02d}" for ano in anos_base for mes in range(1, 13)]
        if not competencias:
            return pl.DataFrame({"Ano/Mes": []}).with_columns(
                [
                    pl.lit(0.0).alias("ICMS_entr_desacob"),
                    pl.lit(0.0).alias("ICMS_saidas_desac"),
                    pl.lit(0.0).alias("ICMS_estoque_desac"),
                    pl.lit(0.0).alias("Total"),
                ]
            )

        base_competencias = pl.DataFrame({"Ano/Mes": competencias})

        if not mensal.is_empty():
            mensal_base = (
                mensal
                .select(
                    [
                        pl.concat_str(
                            [
                                pl.col("ano").cast(pl.Utf8, strict=False),
                                pl.lit("-"),
                                pl.col("mes").cast(pl.Utf8, strict=False).str.zfill(2),
                            ]
                        ).alias("Ano/Mes"),
                        pl.col("ICMS_entr_desacob").cast(pl.Float64, strict=False).fill_null(0.0).alias("ICMS_entr_desacob"),
                    ]
                )
                .group_by("Ano/Mes")
                .agg(pl.col("ICMS_entr_desacob").sum().alias("ICMS_entr_desacob"))
            )
        else:
            mensal_base = pl.DataFrame(
                schema={
                    "Ano/Mes": pl.Utf8,
                    "ICMS_entr_desacob": pl.Float64,
                }
            )

        if not anual.is_empty():
            anual_base = (
                anual
                .select(
                    [
                        pl.concat_str(
                            [
                                pl.col("ano").cast(pl.Utf8, strict=False),
                                pl.lit("-12"),
                            ]
                        ).alias("Ano/Mes"),
                        pl.col("ICMS_saidas_desac").cast(pl.Float64, strict=False).fill_null(0.0).alias("ICMS_saidas_desac"),
                        pl.col("ICMS_estoque_desac").cast(pl.Float64, strict=False).fill_null(0.0).alias("ICMS_estoque_desac"),
                    ]
                )
                .group_by("Ano/Mes")
                .agg(
                    [
                        pl.col("ICMS_saidas_desac").sum().alias("ICMS_saidas_desac"),
                        pl.col("ICMS_estoque_desac").sum().alias("ICMS_estoque_desac"),
                    ]
                )
            )
        else:
            anual_base = pl.DataFrame(
                schema={
                    "Ano/Mes": pl.Utf8,
                    "ICMS_saidas_desac": pl.Float64,
                    "ICMS_estoque_desac": pl.Float64,
                }
            )

        consolidado = (
            base_competencias
            .join(mensal_base, on="Ano/Mes", how="left")
            .join(anual_base, on="Ano/Mes", how="left")
            .with_columns(
                [
                    pl.col("ICMS_entr_desacob").cast(pl.Float64, strict=False).fill_null(0.0).round(2),
                    pl.col("ICMS_saidas_desac").cast(pl.Float64, strict=False).fill_null(0.0).round(2),
                    pl.col("ICMS_estoque_desac").cast(pl.Float64, strict=False).fill_null(0.0).round(2),
                ]
            )
            .with_columns(
                (
                    pl.col("ICMS_entr_desacob")
                    + pl.col("ICMS_saidas_desac")
                    + pl.col("ICMS_estoque_desac")
                ).round(2).alias("Total")
            )
        )

        return consolidado.select(
            ["Ano/Mes", "ICMS_entr_desacob", "ICMS_saidas_desac", "ICMS_estoque_desac", "Total"]
        ).sort(
            ["Ano/Mes"],
            descending=[False],
            nulls_last=True,
        )

    def atualizar_aba_resumo_global(self) -> None:
        if self._aba_mensal_df.is_empty() and self._aba_anual_df.is_empty():
            self.resumo_global_model.set_dataframe(pl.DataFrame())
            self._resumo_global_df = pl.DataFrame()
            self.lbl_resumo_global_status.setText("Aguarde a aba mensal e a aba anual estarem processadas.")
            return

        try:
            resumo = self._gerar_resumo_global(self._aba_mensal_df, self._aba_anual_df)
            self._resumo_global_df = resumo
            self.resumo_global_model.set_dataframe(resumo)
            self.lbl_resumo_global_status.setText(f"Resumo global calculado com base em {resumo.height} competencias.")
        except Exception as e:
            self.show_error("Erro de consoliacao", f"Erro ao calcular Resumo Global: {e}")

    def exportar_resumo_global_excel(self) -> None:
        if self._resumo_global_df is None or self._resumo_global_df.is_empty():
            QMessageBox.information(self, "Exportacao", "Nao ha dados globais para exportar.")
            return
        target = self._save_dialog("Exportar Resumo Global", "Excel (*.xlsx)")
        if not target:
            return
        try:
            df_to_export = self._dataframe_colunas_perfil(
                "resumo_global",
                "resumo_global",
                self.resumo_global_model,
                self._resumo_global_df,
                perfil="Exportar"
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumo Global"
            self._escrever_planilha_openpyxl(ws, df_to_export)
            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target)
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def _montar_valores_consolidados_produtos_selecionados(self, ids: list[str]) -> pl.DataFrame:
        resumo = self._filtrar_dataframe_por_ids(self._produtos_selecionados_df, ids)
        mensal = self._filtrar_dataframe_por_ids(self._produtos_selecionados_mensal_df, ids)
        anual = self._filtrar_dataframe_por_ids(self._produtos_selecionados_anual_df, ids)

        if resumo.is_empty():
            return pl.DataFrame(
                schema={
                    "Ano/Mes": pl.Utf8,
                    "ICMS_entr_desacob": pl.Float64,
                    "ICMS_saidas_desac": pl.Float64,
                    "ICMS_estoque_desac": pl.Float64,
                    "Total": pl.Float64,
                }
            )

        anos_base: list[int] = []
        ano_ini, ano_fim = self._intervalo_anos_produtos_selecionados()
        if ano_ini is not None and ano_fim is not None:
            anos_base = list(range(int(ano_ini), int(ano_fim) + 1))
        else:
            anos_base = None

        return self._gerar_resumo_global(mensal, anual, anos_base)

    def exportar_produtos_selecionados_excel(self) -> None:
        if self._produtos_selecionados_df.is_empty():
            QMessageBox.information(self, "Exportacao", "Nao ha dados consolidados para exportar.")
            return
        target = self._save_dialog("Exportar mov_estoque, mensal e anual", "Excel (*.xlsx)")
        if not target:
            return
        try:
            ids = self._ids_produtos_selecionados_para_exportacao()
            if not ids:
                QMessageBox.information(self, "Exportacao", "Marque pelo menos um id_agregado em Produtos selecionados.")
                return
            mensal = self._dataframe_colunas_perfil(
                "aba_mensal",
                "aba_mensal",
                self.aba_mensal_model,
                self._filtrar_dataframe_por_ids(self._produtos_selecionados_mensal_df, ids),
                perfil="Exportar",
            )
            anual = self._dataframe_colunas_perfil(
                "aba_anual",
                "aba_anual",
                self.aba_anual_model,
                self._filtrar_dataframe_por_ids(self._produtos_selecionados_anual_df, ids),
                perfil="Exportar",
            )
            mov = self._dataframe_colunas_perfil(
                "mov_estoque",
                "mov_estoque",
                self.mov_estoque_model,
                self._filtrar_dataframe_por_ids(self._produtos_selecionados_mov_df, ids),
                perfil="Exportar",
            )
            valores_consolidados = self._montar_valores_consolidados_produtos_selecionados(ids)

            produtos_selecionados_tabela = self._dataframe_colunas_perfil(
                "produtos_selecionados",
                "produtos_selecionados",
                self.produtos_selecionados_model,
                self._filtrar_dataframe_por_ids(self._produtos_selecionados_df, ids),
                perfil="Exportar"
            )

            wb = Workbook()
            ws_produtos = wb.active
            ws_produtos.title = "Produtos_Selecionados"
            self._escrever_planilha_openpyxl(ws_produtos, produtos_selecionados_tabela)

            self._escrever_planilha_openpyxl(wb.create_sheet("Mov_Estoque"), mov)
            self._escrever_planilha_openpyxl(wb.create_sheet("Mensal"), mensal)
            self._escrever_planilha_openpyxl(wb.create_sheet("Anual"), anual)
            self._escrever_planilha_openpyxl(wb.create_sheet("ICMS_devido"), valores_consolidados)
            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target)
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as e:
            self.show_error("Erro de exportacao", str(e))

    def _aba_mensal_foreground(self, row: dict, _col_name: str):
        entradas_desacob = float(row.get("entradas_desacob") or 0)
        icms_entr = float(row.get("ICMS_entr_desacob") or 0)
        if entradas_desacob > 0 or icms_entr > 0:
            return "#fff7ed"
        return "#f5f5f5"

    def _aba_mensal_background(self, row: dict, _col_name: str):
        entradas_desacob = float(row.get("entradas_desacob") or 0)
        icms_entr = float(row.get("ICMS_entr_desacob") or 0)
        if entradas_desacob > 0 or icms_entr > 0:
            return "#5b3a06"
        mes = int(row.get("mes") or 0)
        return "#1f1f1f" if (mes % 2) == 0 else "#262626"

    def _aba_anual_foreground(self, row: dict, _col_name: str):
        entradas_desacob = float(row.get("entradas_desacob") or 0)
        saidas_desacob = float(row.get("saidas_desacob") or 0)
        estoque_final_desacob = float(row.get("estoque_final_desacob") or 0)
        if entradas_desacob > 0 or saidas_desacob > 0 or estoque_final_desacob > 0:
            return "#fff7ed"
        return "#f5f5f5"

    def _aba_anual_background(self, row: dict, _col_name: str):
        entradas_desacob = float(row.get("entradas_desacob") or 0)
        saidas_desacob = float(row.get("saidas_desacob") or 0)
        estoque_final_desacob = float(row.get("estoque_final_desacob") or 0)
        if entradas_desacob > 0 or saidas_desacob > 0 or estoque_final_desacob > 0:
            return "#5b3a06"
        val = str(row.get("id_agregado", ""))
        import hashlib
        h = int(hashlib.md5(val.encode()).hexdigest(), 16)
        return "#1f1f1f" if (h % 2) == 0 else "#262626"

    def _mov_estoque_foreground(self, row: dict, _col_name: str):
        tipo = str(row.get("Tipo_operacao") or "").upper()
        if float(row.get("entr_desac_anual") or 0) > 0:
            return "#fdba74"
        if str(row.get("excluir_estoque", "")).strip().upper() in {"TRUE", "1", "S", "Y", "SIM"}:
            return "#94a3b8"
        if "ESTOQUE FINAL" in tipo:
            return "#fde68a"
        if "ESTOQUE INICIAL" in tipo:
            return "#bfdbfe"
        if "ENTRADA" in tipo:
            return "#93c5fd"
        if "SAIDA" in tipo:
            return "#fca5a5"
        return None

    def _mov_estoque_font(self, row: dict, _col_name: str):
        if float(row.get("entr_desac_anual") or 0) > 0:
            fonte = QFont()
            fonte.setBold(True)
            return fonte
        return None

    def _mov_estoque_background(self, row: dict, _col_name: str):
        tipo = str(row.get("Tipo_operacao") or "").upper()
        if float(row.get("entr_desac_anual") or 0) > 0:
            return "#431407"
        if str(row.get("excluir_estoque", "")).strip().upper() in {"TRUE", "1", "S", "Y", "SIM"}:
            return "#1e293b"
        if str(row.get("mov_rep", "")).strip().upper() in {"TRUE", "1", "S", "Y", "SIM"}:
            return "#111827"
        if "ESTOQUE FINAL" in tipo:
            return "#3f2f10"
        if "ESTOQUE INICIAL" in tipo:
            return "#0f172a"
        if "ENTRADA" in tipo:
            return "#10213f"
        if "SAIDA" in tipo:
            return "#3b1212"
        return None

    def _formatar_resumo_filtros(self, pares: list[tuple[str, str]]) -> str:
        ativos = [f"{rotulo}: {valor}" for rotulo, valor in pares if valor]
        return "Filtros ativos: " + (" | ".join(ativos) if ativos else "nenhum")

    def _obter_cnpj_valido(self) -> str | None:
        if not self.state.current_cnpj:
            self.show_error("CNPJ nao selecionado", "Selecione um CNPJ na lista a esquerda.")
            return None
        return self.state.current_cnpj

    def _executar_em_worker(
        self,
        func: Callable,
        *args,
        mensagem_inicial: str,
        on_success: Callable[[object], None],
        on_failure: Callable[[str], None] | None = None,
        **kwargs,
    ) -> bool:
        if self.service_worker is not None and self.service_worker.isRunning():
            self.show_error("Aguarde", "Ja existe um processamento pesado em execucao.")
            return False

        self.status.showMessage(mensagem_inicial)
        worker = ServiceTaskWorker(func, *args, **kwargs)
        self.service_worker = worker
        worker.progress.connect(self.status.showMessage)

        def _finalizar_ok(resultado) -> None:
            self.service_worker = None
            on_success(resultado)

        def _finalizar_erro(mensagem: str) -> None:
            self.service_worker = None
            if on_failure is not None:
                on_failure(mensagem)
            else:
                self.show_error("Erro", mensagem)

        worker.finished_ok.connect(_finalizar_ok)
        worker.failed.connect(_finalizar_erro)
        self._registrar_limpeza_worker("service_worker", worker)
        worker.start()
        return True

    def on_cnpj_selected(self) -> None:
        item = self.cnpj_list.currentItem()
        if not item:
            return
        cnpj = item.text()
        self.state.current_cnpj = cnpj
        self._produtos_sel_preselecionado_cnpj = None
        self._atualizar_estado_botao_nfe_entrada()
        self._reset_table_resize_flag("conversao")
        self._reset_table_resize_flag("mov_estoque")
        self._reset_table_resize_flag("aba_mensal")
        self._reset_table_resize_flag("aba_anual")
        self._reset_table_resize_flag("nfe_entrada")
        self._reset_table_resize_flag("produtos_selecionados")
        self._reset_table_resize_flag("agregacao_top")
        self._reset_table_resize_flag("agregacao_bottom")
        self.status.showMessage(f"CNPJ selecionado: {cnpj}")
        self._refresh_profile_combos()
        self.refresh_file_tree(cnpj)
        self.atualizar_aba_conversao()
        self.atualizar_aba_mov_estoque()
        self.atualizar_aba_mensal()
        self.atualizar_aba_anual()
        self.atualizar_aba_nfe_entrada()
        self.atualizar_aba_id_agrupados()
        self.atualizar_tabelas_agregacao()
        self.recarregar_historico_agregacao(cnpj)

        # O limite padrao de extracao deve permanecer sempre na data atual.
        # Nao sobrescrevemos mais com a ultima entrega da EFD para evitar que
        # o campo nasca com datas futuras ou historicas ao trocar de CNPJ.
        self.date_input.setDate(QDate.currentDate())


    def refresh_file_tree(self, cnpj: str) -> None:
        self.file_tree.clear()
        
        root_path = self.parquet_service.cnpj_dir(cnpj)
        
        cat_brutas = QTreeWidgetItem(["Tabelas brutas (SQL)", str(root_path / "arquivos_parquet")])
        cat_analises = QTreeWidgetItem(["Analises de Produtos", str(root_path / "analises" / "produtos")])
        cat_outros = QTreeWidgetItem(["Outros Parquets", str(root_path)])
        
        self.file_tree.addTopLevelItem(cat_brutas)
        self.file_tree.addTopLevelItem(cat_analises)
        self.file_tree.addTopLevelItem(cat_outros)

        first_leaf: QTreeWidgetItem | None = None
        
        for path in self.parquet_service.list_parquet_files(cnpj):
            # Identificar categoria
            if "arquivos_parquet" in str(path.parent):
                parent = cat_brutas
            elif "analises" in str(path.parent) or "produtos" in str(path.parent):
                parent = cat_analises
            else:
                parent = cat_outros
                
            item = QTreeWidgetItem([path.name, str(path.parent)])
            item.setData(0, Qt.UserRole, str(path))
            parent.addChild(item)
            if first_leaf is None:
                first_leaf = item
                
        cat_brutas.setExpanded(True)
        cat_analises.setExpanded(True)
        
        # Limpar categorias vazias
        for cat in [cat_brutas, cat_analises, cat_outros]:
            if cat.childCount() == 0:
                self.file_tree.takeTopLevelItem(self.file_tree.indexOfTopLevelItem(cat))

        if first_leaf is not None:
            self.file_tree.setCurrentItem(first_leaf)
            self.on_file_activated(first_leaf, 0)

    def on_file_activated(self, item: QTreeWidgetItem, _column: int) -> None:
        raw_path = item.data(0, Qt.UserRole)
        if not raw_path:
            return
        self.state.current_file = Path(raw_path)
        self._reset_table_resize_flag("consulta")
        self.state.filters = []
        self.current_page_df_all = pl.DataFrame()
        self.current_page_df_visible = pl.DataFrame()
        self.load_current_file(reset_columns=True)
        self.tabs.setCurrentIndex(0)

    def load_current_file(self, reset_columns: bool = False) -> None:
        if self.state.current_file is None:
            return
        try:
            all_columns = self.parquet_service.get_schema(self.state.current_file)
        except Exception as exc:
            self.show_error("Erro ao abrir Parquet", str(exc))
            return
        self.state.all_columns = all_columns
        prefs = self._carregar_preferencias_tabela("consulta", self._consulta_scope())
        self._refresh_profile_combos()
        pref_visiveis = prefs.get("visible_columns") if isinstance(prefs, dict) else None
        if reset_columns:
            self.state.visible_columns = pref_visiveis if isinstance(pref_visiveis, list) and pref_visiveis else all_columns[:]
        elif not self.state.visible_columns:
            self.state.visible_columns = pref_visiveis if isinstance(pref_visiveis, list) and pref_visiveis else all_columns[:]
        self.filter_column.clear()
        self.filter_column.addItems(all_columns)
        self.reload_table()

    def reload_table(self, update_main_view: bool = True) -> None:
        if self.state.current_file is None:
            return
        try:
            self.current_page_df_all = self.parquet_service.load_dataset(
                self.state.current_file,
                self.state.filters or [],
            )
            colunas_visiveis = self.state.visible_columns or self.current_page_df_all.columns
            self.current_page_df_visible = self.current_page_df_all.select(
                [coluna for coluna in colunas_visiveis if coluna in self.current_page_df_all.columns]
            )
            self.state.total_rows = self.current_page_df_all.height

            if update_main_view:
                self.table_model.set_dataframe(self.current_page_df_visible)
                self._update_page_label()
                self._update_context_label()
                self._refresh_filter_list_widget()
                self._resize_table_once(self.table_view, "consulta")
                self._aplicar_preferencias_tabela("consulta", self.table_view, self.table_model, self._consulta_scope())
        except Exception as exc:
            self.show_error("Erro ao carregar dados", str(exc))

    def _update_page_label(self) -> None:
        self.lbl_page.setText(f"Linhas filtradas: {self.state.total_rows}")

    def _update_context_label(self) -> None:
        if self.state.current_file is None:
            self.lbl_context.setText("Nenhum arquivo selecionado")
            return
        self.lbl_context.setText(
            f"CNPJ: {self.state.current_cnpj or '-'} | Arquivo: {self.state.current_file.name} | "
            f"Colunas visiveis: {len(self.state.visible_columns or [])}/{len(self.state.all_columns or [])}"
        )

    def add_filter_from_form(self) -> None:
        column = self.filter_column.currentText().strip()
        operator = self.filter_operator.currentText().strip()
        value = self.filter_value.text().strip()
        if not column:
            self.show_error("Filtro invalido", "Selecione uma coluna para filtrar.")
            return
        if operator not in {"e nulo", "nao e nulo"} and value == "":
            self.show_error("Filtro invalido", "Informe um valor para o filtro escolhido.")
            return
        self.state.filters = self.state.filters or []
        self.state.filters.append(FilterCondition(column=column, operator=operator, value=value))
        self.filter_value.clear()
        self.reload_table()

    def clear_filters(self) -> None:
        self.state.filters = []
        self.reload_table()

    def remove_selected_filter(self) -> None:
        row = self.filter_list.currentRow()
        if row < 0 or not self.state.filters:
            return
        self.state.filters.pop(row)
        self.reload_table()

    def _refresh_filter_list_widget(self) -> None:
        self.filter_list.clear()
        for cond in self.state.filters or []:
            text = f"{cond.column} {cond.operator} {cond.value}".strip()
            self.filter_list.addItem(text)

    def choose_columns(self) -> None:
        if not self.state.all_columns:
            return
        dialog = ColumnSelectorDialog(self.state.all_columns, self.state.visible_columns or self.state.all_columns, self)
        if dialog.exec():
            selected = dialog.selected_columns()
            if not selected:
                self.show_error("Selecao invAlida", "Pelo menos uma coluna deve permanecer visivel.")
                return
            self.state.visible_columns = selected
            prefs = self._carregar_preferencias_tabela("consulta", self._consulta_scope())
            prefs["visible_columns"] = selected
            prefs.pop("header_state", None)
            self.selection_service.set_value(self._preferencia_tabela_key("consulta", self._consulta_scope()), prefs)
            self.reload_table()

    def _save_dialog(self, title: str, pattern: str) -> Path | None:
        filename, _ = QFileDialog.getSaveFileName(self, title, str(CONSULTAS_ROOT), pattern)
        return Path(filename) if filename else None

    def _filters_text(self) -> str:
        return " | ".join(f"{f.column} {f.operator} {f.value}".strip() for f in self.state.filters or [])

    def _dataset_for_export(self, mode: str) -> pl.DataFrame:
        if self.state.current_file is None:
            raise ValueError("Nenhum arquivo selecionado.")
        if mode == "full":
            return self.parquet_service.load_dataset(self.state.current_file)
        if mode == "filtered":
            return self.parquet_service.load_dataset(self.state.current_file, self.state.filters or [])
        if mode == "visible":
            return self.parquet_service.load_dataset(
                self.state.current_file,
                self.state.filters or [],
                self.state.visible_columns or [],
            )
        raise ValueError(f"Modo de exportacao nao suportado: {mode}")

    def export_excel(self, mode: str) -> None:
        try:
            df = self._dataset_for_export(mode)
            if mode != "visible":
                df = self._dataframe_colunas_visiveis(self.table_view, self.table_model, df)
            target = self._save_dialog("Salvar Excel", "Excel (*.xlsx)")
            if not target:
                return
            self.export_service.export_excel(target, df, sheet_name=self.state.current_file.stem if self.state.current_file else "Dados")
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as exc:
            self.show_error("Falha na exportacao para Excel", str(exc))

    def export_docx(self) -> None:
        try:
            if self.state.current_file is None:
                raise ValueError("Nenhum arquivo selecionado.")
            df = self.parquet_service.load_dataset(self.state.current_file, self.state.filters or [], self.state.visible_columns or [])
            target = self._save_dialog("Salvar relatorio Word", "Word (*.docx)")
            if not target:
                return
            self.export_service.export_docx(
                target,
                title="Relatorio Padronizado de AnAlise Fiscal",
                cnpj=self.state.current_cnpj or "",
                table_name=self.state.current_file.name,
                df=df,
                filters_text=self._filters_text(),
                visible_columns=self.state.visible_columns or [],
            )
            self.show_info("Relatorio gerado", f"Arquivo gerado em:\n{target}")
        except Exception as exc:
            self.show_error("Falha na exportacao para Word", str(exc))

    def export_txt_html(self) -> None:
        try:
            if self.state.current_file is None:
                raise ValueError("Nenhum arquivo selecionado.")
            df = self.parquet_service.load_dataset(self.state.current_file, self.state.filters or [], self.state.visible_columns or [])
            html_report = self.export_service.build_html_report(
                title="Relatorio Padronizado de AnAlise Fiscal",
                cnpj=self.state.current_cnpj or "",
                table_name=self.state.current_file.name,
                df=df,
                filters_text=self._filters_text(),
                visible_columns=self.state.visible_columns or [],
            )
            target = self._save_dialog("Salvar TXT com HTML", "TXT (*.txt)")
            if not target:
                return
            self.export_service.export_txt_with_html(target, html_report)
            self.show_info("Relatorio HTML/TXT gerado", f"Arquivo gerado em:\n{target}")
        except Exception as exc:
            self.show_error("Falha na exportacao TXT/HTML", str(exc))

    def open_editable_aggregation_table(self) -> None:
        if not self.state.current_cnpj:
            self.show_error("CNPJ nao selecionado", "Selecione um CNPJ na lista.")
            return
        try:
            target = self.servico_agregacao.carregar_tabela_editavel(self.state.current_cnpj)
            self._aggregation_file_path = target
            self._aggregation_filters = []
            self._aggregation_results_filters = []
            self._aggregation_relational_mode = None
            self._aggregation_results_relational_mode = None
            self._load_aggregation_table()
            self.recarregar_historico_agregacao(self.state.current_cnpj)
        except Exception as exc:
            self.show_error("Falha ao abrir tabela editAvel", str(exc))
            return

        self.tabs.setCurrentIndex(2) # Switch to Agregacao tab (0-indexed: Consulta, SQL, Agregacao, Logs)

    def _abrir_tabela_agrupada(self) -> None:
        self.open_editable_aggregation_table()

    def _desfazer_agregacao(self) -> None:
        self.aggregation_table_model.clear_checked()
        self.results_table_model.clear_checked()
        self.status.showMessage("Selecao de agregacao limpa.")

    def _obter_ids_agrupados_para_reversao(self) -> list[str]:
        rows = self.results_table_model.get_checked_rows()
        if not rows:
            selecao = self.results_table.selectionModel()
            if selecao is not None:
                df = self.results_table_model.get_dataframe()
                rows = [
                    df.row(index.row(), named=True)
                    for index in selecao.selectedRows()
                    if 0 <= index.row() < df.height
                ]

        ids: list[str] = []
        vistos: set[str] = set()
        for row in rows:
            valor = str(row.get("id_agrupado") or "").strip()
            if not valor or valor in vistos:
                continue
            vistos.add(valor)
            ids.append(valor)
        return ids

    def reverter_agregacao(self) -> None:
        if not self.state.current_cnpj:
            self.show_error("CNPJ nao selecionado", "Selecione um CNPJ antes de reverter agrupamentos.")
            return

        ids_reversao = self._obter_ids_agrupados_para_reversao()
        if not ids_reversao:
            self.show_error(
                "Selecao insuficiente",
                "Marque ou selecione na tabela inferior o agrupamento que deve ser revertido.",
            )
            return

        mensagem = (
            "Isso vai restaurar os grupos de origem do(s) agrupamento(s) selecionado(s):\n"
            + "\n".join(ids_reversao)
            + "\n\nProsseguir?"
        )
        if QMessageBox.question(self, "Reverter agrupamento", mensagem) != QMessageBox.StandardButton.Yes:
            return

        try:
            resultados = [
                self.servico_agregacao.reverter_agrupamento(self.state.current_cnpj, id_agrupado)
                for id_agrupado in ids_reversao
            ]
            self.atualizar_tabelas_agregacao()
            self.recarregar_historico_agregacao(self.state.current_cnpj)
            self.atualizar_aba_id_agrupados()
            self.refresh_logs()

            total_restaurado = sum(int(item.get("qtd_grupos_restaurados", 0)) for item in resultados)
            self.show_info(
                "Agrupamento revertido",
                f"Foram restaurados {total_restaurado} grupos a partir de {len(ids_reversao)} agrupamento(s).",
            )
        except Exception as exc:
            self.show_error("Erro ao reverter agrupamento", str(exc))

    def _load_aggregation_table(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            return
        if self._aggregation_file_path is None:
            return
        self._aggregation_file_path = self.servico_agregacao.carregar_tabela_editavel(cnpj)
        df = self.parquet_service.load_dataset(self._aggregation_file_path, self._aggregation_filters or [])
        df = self._aplicar_modo_relacional_agregacao_df(df, self._aggregation_relational_mode)
        self.aggregation_table_model.set_dataframe(df)
        self._resize_table_once(self.aggregation_table_view, "agregacao_top")
        if not self._aplicar_preferencias_tabela("agregacao_top", self.aggregation_table, self.aggregation_table_model):
            self._aplicar_perfil_agregacao(
                "agregacao_top",
                self.aggregation_table,
                self.aggregation_table_model,
                self.top_profile.currentText(),
            )

    def execute_aggregation(self) -> None:
        if not self.state.current_cnpj:
            self.show_error("CNPJ nao selecionado", "Selecione um CNPJ antes de agregar.")
            return

        rows_top = self.aggregation_table_model.get_checked_rows()
        rows_bottom = self.results_table_model.get_checked_rows()
        
        # Merge and de-duplicate
        combined = []
        seen = set()
        for r in (rows_top + rows_bottom):
            key = str(r.get("id_agrupado") or "").strip()
            if not key:
                key = str(r.get("chave_produto") or "").strip()
            if not key:
                key = str(r.get("chave_item") or "").strip()
            if not key:
                key = str(r.get("descr_padrao") or r.get("descricao") or "").strip().upper()
            if key not in seen:
                seen.add(key)
                combined.append(r)

        if len(combined) < 2:
            self.show_error("Selecao insuficiente", "Marque pelo menos duas linhas com 'Visto' (pode ser em ambas as tabelas) para agregar.")
            return

        try:
            # Novo: Passar lista de IDs agrupados para o servico
            ids_selecionados = [str(r.get("id_agrupado") or "") for r in combined if r.get("id_agrupado")]
            
            if len(ids_selecionados) < 2:
                 self.show_error("Selecao insuficiente", "Nao foi possivel identificar IDs unicos para os grupos selecionados.")
                 return

            self.servico_agregacao.agregar_linhas(
                cnpj=self.state.current_cnpj,
                ids_agrupados_selecionados=ids_selecionados,
            )
            # Update the tables to reflect the changes
            self.atualizar_tabelas_agregacao()
            self.recarregar_historico_agregacao(self.state.current_cnpj)
            self.refresh_logs()
            
            self.show_info(
                "Agregacao concluida",
                f"As {len(combined)} descricoes foram unificadas."
            )
        except Exception as e:
            import traceback
            from utilitarios.perf_monitor import registrar_evento_performance
            registrar_evento_performance("main_window.agregacao_erro", contexto={"erro": str(e), "traceback": traceback.format_exc()}, status="error")
            self.show_error("Erro na agregacao", "Ocorreu um erro interno ao agregar. Consulte os logs internos para mais detalhes.")
            
            # Clear checks and reload top table
            self.aggregation_table_model.clear_checked()
            self.results_table_model.clear_checked()
            self.open_editable_aggregation_table()

    def apply_quick_filters(self) -> None:
        idx = self.tabs.currentIndex()
        if idx == 0: # Consulta
            fields = {
                "descricao_normalizada": self.qf_norm.text().strip(),
                "descricao": self.qf_desc.text().strip(),
                "ncm_padrao": self.qf_ncm.text().strip(),
                "cest_padrao": self.qf_cest.text().strip(),
            }
        elif idx == 2: # Agregacao (Index 2 is "Agregacao", Index 1 is "SQL")
            fields = {
                "descricao_normalizada": self.aqf_norm.text().strip(),
                "descricao": self.aqf_desc.text().strip(),
                "ncm_padrao": self.aqf_ncm.text().strip(),
                "cest_padrao": self.aqf_cest.text().strip(),
            }
        else:
            return

        def split_terms(value: str) -> list[str]:
            texto = (value or "").strip()
            if not texto:
                return []
            # Permite buscar varios trechos no mesmo campo.
            # Ex.: "buch 18", "buch;18" ou "buch, 18".
            partes = re.split(r"[;,]+|\s{2,}", texto)
            if len(partes) == 1 and " " in texto:
                partes = texto.split()
            return [p.strip() for p in partes if p and p.strip()]

        # Mapas de colunas equivalentes por tipo de filtro rapido.
        # Inclui colunas usadas na aba de Agregacao (ex.: descr_padrao).
        alternatives = {
            "ncm_padrao": ["ncm_padrao", "NCM_padrao", "lista_ncm", "ncm_final", "ncm"],
            "cest_padrao": ["cest_padrao", "CEST_padrao", "lista_cest", "cest_final", "cest"],
            "descricao_normalizada": [
                "descricao_normalizada",
                "descricao",
                "descr_norm",
                "descr_padrao",
                "descricao_final",
            ],
            "descricao": [
                "descricao",
                "lista_descricoes",
                "lista_desc_compl",
                "lista_itens_agrupados",
                "descr",
                "descr_padrao",
                "descricao_final",
            ],
        }

        # Remove filtros rapidos antigos (inclusive quando ficaram com nome de coluna "alias").
        quick_target_cols = set(fields.keys())
        for key in fields.keys():
            quick_target_cols.update(alternatives.get(key, []))

        # Na aba de Agregacao, o filtro rapido deve ser deterministico:
        # substitui totalmente os filtros anteriores para evitar "filtros ocultos".
        if idx == 2:
            new_filters = []
        else:
            new_filters = [f for f in (self.state.filters or []) if f.column not in quick_target_cols]
        
        available_columns = self.state.all_columns or []
        if idx == 2 and self._aggregation_file_path is not None:
            try:
                available_columns = self.parquet_service.get_schema(self._aggregation_file_path)
            except Exception:
                available_columns = list(self.aggregation_table_model.dataframe.columns)

        for col, val in fields.items():
            termos = split_terms(val)
            if termos:
                # Need to be flexible with column names as they might differ across files
                # We'll use the one present in the schema
                actual_col = col
                if available_columns:
                    # Match case-sensitive in alias map first.
                    if col in alternatives:
                        for alt in alternatives[col]:
                            if alt in available_columns:
                                actual_col = alt
                                break

                    # Fallback: match case/acento-insensitive
                    if actual_col not in available_columns:
                        target_clean = remove_accents(col).lower()
                        for c in available_columns:
                            if remove_accents(c).lower() == target_clean:
                                actual_col = c
                                break

                # Usa operador ASCII para evitar problemas de encoding no caminho UI -> servico.
                # Cada termo vira um filtro proprio; como os filtros sao encadeados,
                # a busca exige que todos os trechos estejam presentes.
                for termo in termos:
                    new_filters.append(FilterCondition(column=actual_col, operator="contem", value=termo))
        
        if idx == 2:
            self._aggregation_filters = new_filters
            self._load_aggregation_table()
        else:
            self.state.filters = new_filters
            self.reload_table(update_main_view=True)

    def refresh_logs(self) -> None:
        import json
        logs = [json.dumps(log) for log in self.servico_agregacao.ler_linhas_log()]
        self.log_view.setPlainText("\n".join(logs))

    def apply_aggregation_results_filters(self) -> None:
        if self.tabs.currentIndex() != 2:
            return

        fields = {
            "descricao_normalizada": self.bqf_norm.text().strip(),
            "descricao": self.bqf_desc.text().strip(),
            "ncm_padrao": self.bqf_ncm.text().strip(),
            "cest_padrao": self.bqf_cest.text().strip(),
        }

        def split_terms(value: str) -> list[str]:
            texto = (value or "").strip()
            if not texto:
                return []
            partes = re.split(r"[;,]+|\s{2,}", texto)
            if len(partes) == 1 and " " in texto:
                partes = texto.split()
            return [p.strip() for p in partes if p and p.strip()]

        alternatives = {
            "ncm_padrao": ["ncm_padrao", "NCM_padrao", "lista_ncm", "ncm_final", "ncm"],
            "cest_padrao": ["cest_padrao", "CEST_padrao", "lista_cest", "cest_final", "cest"],
            "descricao_normalizada": [
                "descricao_normalizada",
                "descricao",
                "descr_norm",
                "descr_padrao",
                "descricao_final",
            ],
            "descricao": [
                "descricao",
                "lista_descricoes",
                "lista_desc_compl",
                "lista_itens_agrupados",
                "descr",
                "descr_padrao",
                "descricao_final",
            ],
        }

        new_filters: list[FilterCondition] = []
        cnpj = self.state.current_cnpj
        if not cnpj:
            self._aggregation_results_filters = []
            self.recarregar_historico_agregacao("")
            return

        path = self.servico_agregacao.caminho_tabela_agregadas(cnpj)
        available_columns = []
        if path.exists():
            try:
                available_columns = self.parquet_service.get_schema(path)
            except Exception:
                available_columns = list(self.results_table_model.dataframe.columns)

        for col, val in fields.items():
            termos = split_terms(val)
            if not termos:
                continue

            actual_col = col
            if available_columns:
                if col in alternatives:
                    for alt in alternatives[col]:
                        if alt in available_columns:
                            actual_col = alt
                            break
                if actual_col not in available_columns:
                    target_clean = remove_accents(col).lower()
                    for c in available_columns:
                        if remove_accents(c).lower() == target_clean:
                            actual_col = c
                            break

            for termo in termos:
                new_filters.append(FilterCondition(column=actual_col, operator="contem", value=termo))

        self._aggregation_results_filters = new_filters
        self.recarregar_historico_agregacao(cnpj)

    def _obter_linha_selecionada_tabela(self, table: QTableView, model: PolarsTableModel) -> dict | None:
        df = model.get_dataframe()
        if df.is_empty():
            return None

        indice = table.currentIndex()
        if not indice.isValid():
            indices = table.selectionModel().selectedIndexes() if table.selectionModel() else []
            if not indices:
                return None
            indice = indices[0]

        linha = indice.row()
        if linha < 0 or linha >= df.height:
            return None
        return df.row(linha, named=True)

    def _resolver_coluna_agregacao(self, aliases: list[str], available_columns: list[str]) -> str | None:
        for alias in aliases:
            if alias in available_columns:
                return alias

        normalizadas = {remove_accents(col).lower(): col for col in available_columns}
        for alias in aliases:
            chave = remove_accents(alias).lower()
            if chave in normalizadas:
                return normalizadas[chave]
        return None

    def _aplicar_modo_relacional_agregacao_df(self, df: pl.DataFrame, modo: str | None) -> pl.DataFrame:
        if df.is_empty() or not modo:
            return df

        aliases_map = {
            "ncm": ["ncm_padrao", "NCM_padrao", "lista_ncm", "ncm_final", "ncm"],
            "cest": ["cest_padrao", "CEST_padrao", "lista_cest", "cest_final", "cest"],
            "gtin": ["gtin_padrao", "GTIN_padrao", "gtin", "cod_barra", "cod_barras"],
        }

        available_columns = list(df.columns)
        col_ncm = self._resolver_coluna_agregacao(aliases_map["ncm"], available_columns)
        col_cest = self._resolver_coluna_agregacao(aliases_map["cest"], available_columns)
        if not col_ncm or not col_cest:
            return df

        chaves: list[tuple[str, str]] = [("ncm", col_ncm), ("cest", col_cest)]
        if modo == "ncm_cest_gtin":
            col_gtin = self._resolver_coluna_agregacao(aliases_map["gtin"], available_columns)
            if not col_gtin:
                return df.head(0)
            chaves.append(("gtin", col_gtin))

        temporarias = [f"__rel_{nome}" for nome, _col in chaves]
        df_rel = df.with_row_index("__row_pos").with_columns(
            [
                pl.col(col)
                .cast(pl.Utf8, strict=False)
                .fill_null("")
                .str.strip_chars()
                .alias(f"__rel_{nome}")
                for nome, col in chaves
            ]
        )

        for coluna_tmp in temporarias:
            df_rel = df_rel.filter(pl.col(coluna_tmp) != "")

        if df_rel.is_empty():
            return df.head(0)

        df_repetidos = (
            df_rel
            .group_by(temporarias)
            .agg(pl.len().alias("__match_count"))
            .filter(pl.col("__match_count") >= 2)
        )

        if df_repetidos.is_empty():
            return df.head(0)

        return (
            df_rel
            .join(df_repetidos, on=temporarias, how="inner")
            .sort("__row_pos")
            .drop(["__row_pos", "__match_count", *temporarias], strict=False)
        )

    def _aplicar_filtro_relacional_agregacao(self, destino: str, include_gtin: bool) -> None:
        if self.tabs.currentIndex() != 2:
            return

        modo = "ncm_cest_gtin" if include_gtin else "ncm_cest"

        if destino == "top":
            if self._aggregation_file_path is None:
                self.show_error("Tabela indisponivel", "Abra a tabela de agregacao antes de aplicar o filtro.")
                return
            self._aggregation_relational_mode = modo
            self._load_aggregation_table()
        else:
            cnpj = self.state.current_cnpj
            if not cnpj:
                self.show_error("CNPJ nao selecionado", "Selecione um CNPJ antes de aplicar o filtro.")
                return
            self._aggregation_results_relational_mode = modo
            self.recarregar_historico_agregacao(cnpj)

        rotulo = "NCM+CEST+GTIN iguais" if include_gtin else "NCM+CEST iguais"
        self.status.showMessage(f"Filtro relacional ativo: {rotulo}.")

    def clear_top_aggregation_filters(self) -> None:
        for widget in [self.aqf_norm, self.aqf_desc, self.aqf_ncm, self.aqf_cest]:
            widget.clear()
        self._aggregation_filters = []
        self._aggregation_relational_mode = None
        self._load_aggregation_table()

    def clear_bottom_aggregation_filters(self) -> None:
        for widget in [self.bqf_norm, self.bqf_desc, self.bqf_ncm, self.bqf_cest]:
            widget.clear()
        self._aggregation_results_filters = []
        self._aggregation_results_relational_mode = None
        cnpj = self.state.current_cnpj or ""
        self.recarregar_historico_agregacao(cnpj)

    def open_cnpj_folder(self) -> None:
        if not self.state.current_cnpj:
            self.show_error("CNPJ nao selecionado", "Selecione um CNPJ para abrir a pasta.")
            return
        target = self.parquet_service.cnpj_dir(self.state.current_cnpj)
        if not target.exists():
            self.show_error("Pasta inexistente", f"A pasta {target} ainda nao foi criada.")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(target)))

    def _on_conversion_selection_changed(self, selected, deselected) -> None:
        indexes = self.conversion_table.selectionModel().selectedIndexes()
        if not indexes or self._conversion_df_full.is_empty():
            self.lbl_produto_sel.setText("Nenhum produto selecionado")
            self.combo_unid_ref.clear()
            self.combo_unid_ref.setEnabled(False)
            self.btn_apply_unid_ref.setEnabled(False)
            self._current_selected_id_produto = None
            return
            
        row = indexes[0].row()
        df = self.conversion_model.dataframe
        if row < 0 or row >= df.height:
            return
            
        id_prod = df.item(row, df.columns.index("id_produtos"))
        descr = df.item(row, df.columns.index("descr_padrao"))
        
        self.lbl_produto_sel.setText(f"{id_prod} - {descr}")
        self._current_selected_id_produto = id_prod
        
        # Obter unidades unicas originais vinculadas a este ID
        try:
            unidades_s = self._conversion_df_full.filter(pl.col("id_produtos") == id_prod).get_column("unid").drop_nulls().cast(pl.Utf8)
            unidades = unidades_s.unique().to_list() if not unidades_s.is_empty() else []
        except Exception:
            unidades = []
        
        self.combo_unid_ref.clear()
        if unidades:
            self.combo_unid_ref.addItems(sorted(unidades))
            self.combo_unid_ref.setEnabled(True)
            self.btn_apply_unid_ref.setEnabled(True)
        else:
            self.combo_unid_ref.setEnabled(False)
            self.btn_apply_unid_ref.setEnabled(False)

    def _apply_unid_ref_to_all(self) -> None:
        id_prod = getattr(self, "_current_selected_id_produto", None)
        nova_unid = self.combo_unid_ref.currentText()
        if not id_prod or not nova_unid or self._conversion_df_full.is_empty():
            return
            
        # Determinar o preco medio da nova unidade de referencia
        df_prod = self._conversion_df_full.filter(pl.col("id_produtos") == id_prod)
        row_ref = df_prod.filter(pl.col("unid") == nova_unid)
        
        novo_preco_ref = None
        if not row_ref.is_empty():
            val = row_ref.get_column("preco_medio")[0]
            if val is not None:
                try:
                    novo_preco_ref = float(val)
                except Exception:
                    pass

        # Atualizar unid_ref para as linhas do produto
        self._conversion_df_full = self._conversion_df_full.with_columns(
            pl.when(pl.col("id_produtos") == id_prod)
            .then(pl.lit(nova_unid))
            .otherwise(pl.col("unid_ref"))
            .alias("unid_ref")
        )
        self._conversion_df_full = self._conversion_df_full.with_columns(
            [
                pl.when(pl.col("id_produtos") == id_prod)
                .then(pl.lit(True))
                .otherwise(pl.col("unid_ref_manual").cast(pl.Boolean, strict=False).fill_null(False))
                .alias("unid_ref_manual"),
                pl.when(pl.col("id_produtos") == id_prod)
                .then(pl.lit(False))
                .otherwise(pl.col("fator_manual").cast(pl.Boolean, strict=False).fill_null(False))
                .alias("fator_manual"),
            ]
        )
        
        # Recalcular fatores de conversao das unidades relativas ao novo preco alvo
        if novo_preco_ref is not None and novo_preco_ref > 0:
            self._conversion_df_full = self._conversion_df_full.with_columns(
                pl.when(pl.col("id_produtos") == id_prod)
                .then(
                    pl.when(pl.col("preco_medio").is_not_null())
                    .then(pl.col("preco_medio").cast(pl.Float64) / novo_preco_ref)
                    .otherwise(1.0)
                )
                .otherwise(pl.col("fator"))
                .alias("fator")
            )
        else:
            # Caso a nova unidade selecionada nao tenha preco medio valido, forcamos fator 1.0 para todo o produto
            self._conversion_df_full = self._conversion_df_full.with_columns(
                pl.when(pl.col("id_produtos") == id_prod)
                .then(pl.lit(1.0))
                .otherwise(pl.col("fator"))
                .alias("fator")
            )
        
        # Salvar as alteracoes matematicas
        if self._conversion_file_path:
            (
                self._preparar_dataframe_para_salvar_conversao(self._conversion_df_full)
                .write_parquet(self._conversion_file_path)
            )
            
        self.status.showMessage(f"Unidade {nova_unid} e fatores recalculados aplicados para {id_prod}.")
        self.atualizar_aba_conversao()
        self._marcar_recalculo_conversao_pendente("Clique em 'Recalcular fatores' ou mude de tela.")

    def recalcular_derivados_conversao(self, show_popup: bool = True) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            self.show_error("CNPJ nao selecionado", "Selecione um CNPJ antes de recalcular.")
            return
        if self._recalculando_conversao:
            return
        if not self._conversion_recalc_pending and show_popup:
            self.status.showMessage("Nao ha recalculo pendente na aba Conversao.")
            return

        self._recalculando_conversao = True

        def _on_success(ok) -> None:
            self._recalculando_conversao = False
            resumo = self.servico_agregacao.resumo_tempos()
            if ok:
                self._limpar_recalculo_conversao_pendente()
                self.atualizar_aba_mov_estoque()
                self.atualizar_aba_mensal()
                self.atualizar_aba_anual()
                self.atualizar_aba_nfe_entrada()
                self.atualizar_aba_id_agrupados()
                self.refresh_file_tree(cnpj)
                self.status.showMessage(
                    "Conversao aplicada; mov_estoque, mensal e anual recalculadas."
                    + (f" {resumo}" if resumo else "")
                )
                if show_popup:
                    self.show_info(
                        "Conversao aplicada",
                        "Fatores salvos; mov_estoque, mensal e anual foram recalculadas."
                        + (f"\n\nTempos: {resumo}" if resumo else ""),
                    )
            else:
                self.status.showMessage("Falha ao recalcular derivados da conversao.")
                if show_popup:
                    self.show_error(
                        "Falha no recalculo",
                        "Nao foi possivel recalcular mov_estoque, mensal e anual a partir da conversao.",
                    )

        def _on_failure(mensagem: str) -> None:
            self._recalculando_conversao = False
            self.status.showMessage(f"Erro ao recalcular derivados da conversao: {mensagem}")
            if show_popup:
                self.show_error("Falha no recalculo", mensagem)

        iniciado = self._executar_em_worker(
            self.servico_agregacao.recalcular_mov_estoque,
            cnpj,
            mensagem_inicial="Recalculando mov_estoque, mensal e anual...",
            on_success=_on_success,
            on_failure=_on_failure,
        )
        if not iniciado:
            self._recalculando_conversao = False

    def _enriquecer_dataframe_conversao(self, df: pl.DataFrame) -> pl.DataFrame:
        if df.is_empty():
            return df
        if not {"id_agrupado", "unid", "unid_ref"}.issubset(set(df.columns)):
            return df

        df_base = df.drop(["preco_medio_ref", "fator_calculado"], strict=False)
        ref_price = (
            df_base
            .filter(pl.col("unid").cast(pl.Utf8, strict=False) == pl.col("unid_ref").cast(pl.Utf8, strict=False))
            .group_by("id_agrupado")
            .agg(pl.col("preco_medio").cast(pl.Float64, strict=False).drop_nulls().mean().alias("preco_medio_ref"))
        )
        df_enriquecido = (
            df_base
            .join(ref_price, on="id_agrupado", how="left")
            .with_columns(
                pl.when(pl.col("preco_medio_ref").cast(pl.Float64, strict=False) > 0)
                .then(
                    pl.col("preco_medio").cast(pl.Float64, strict=False)
                    / pl.col("preco_medio_ref").cast(pl.Float64, strict=False)
                )
                .otherwise(1.0)
                .alias("fator_calculado")
            )
        )

        colunas = list(df_enriquecido.columns)
        if "fator" in colunas:
            novas = [c for c in ["preco_medio_ref", "fator_calculado"] if c in colunas]
            for nome in novas:
                colunas.remove(nome)
            idx_fator = colunas.index("fator")
            for deslocamento, nome in enumerate(novas, start=1):
                colunas.insert(idx_fator + deslocamento, nome)
            df_enriquecido = df_enriquecido.select(colunas)

        return df_enriquecido

    def _montar_descricoes_exibicao_por_grupo(self, df_descricoes: pl.DataFrame) -> pl.DataFrame:
        """
        Normaliza a lista exibida na aba de conversao sem alterar a
        fonte canonica persistida em parquet.
        """
        if df_descricoes.is_empty() or not {"id_agrupado", "descricao_item"}.issubset(set(df_descricoes.columns)):
            return pl.DataFrame()

        return (
            df_descricoes
            .select(
                [
                    pl.col("id_agrupado").cast(pl.Utf8, strict=False),
                    pl.col("descricao_item").cast(pl.Utf8, strict=False),
                ]
            )
            .with_columns(
                pl.col("descricao_item").fill_null("").str.strip_chars().alias("descricao_item")
            )
            .filter(pl.col("descricao_item") != "")
            .unique(subset=["id_agrupado", "descricao_item"])
            .sort(["id_agrupado", "descricao_item"], nulls_last=True)
            .group_by("id_agrupado")
            .agg(pl.col("descricao_item").alias("__lista_descricoes"))
            .with_columns(
                pl.col("__lista_descricoes").list.join(" | ").alias("lista_descricoes_produto")
            )
            .select(["id_agrupado", "lista_descricoes_produto"])
        )

    def _carregar_descr_padrao_canonico_conversao(self, cnpj: str) -> pl.DataFrame:
        """
        Garante que a aba use o descr_padrao atual do agrupamento.
        """
        arquivos_canonicos = [
            CNPJ_ROOT / cnpj / "analises" / "produtos" / f"produtos_agrupados_{cnpj}.parquet",
            CNPJ_ROOT / cnpj / "analises" / "produtos" / f"id_agrupados_{cnpj}.parquet",
            CNPJ_ROOT / cnpj / "analises" / "produtos" / f"produtos_final_{cnpj}.parquet",
        ]

        for caminho in arquivos_canonicos:
            if not caminho.exists():
                continue
            try:
                df_origem = self._carregar_dataset_ui(caminho, columns=["id_agrupado", "descr_padrao"])
            except Exception:
                continue

            if df_origem.is_empty() or not {"id_agrupado", "descr_padrao"}.issubset(set(df_origem.columns)):
                continue

            return (
                df_origem
                .select(
                    [
                        pl.col("id_agrupado").cast(pl.Utf8, strict=False),
                        pl.col("descr_padrao").cast(pl.Utf8, strict=False).fill_null("").str.strip_chars().alias("descr_padrao_canonico"),
                    ]
                )
                .filter(pl.col("descr_padrao_canonico") != "")
                .unique(subset=["id_agrupado"], keep="first")
            )

        return pl.DataFrame()

    def _carregar_descricoes_canonicas_conversao(self, cnpj: str) -> pl.DataFrame:
        """
        Prefere a lista consolidada do ETL; so recorre ao fallback via
        produtos_final quando o schema antigo ainda nao foi regenerado.
        """
        arquivos_canonicos = [
            CNPJ_ROOT / cnpj / "analises" / "produtos" / f"produtos_agrupados_{cnpj}.parquet",
            CNPJ_ROOT / cnpj / "analises" / "produtos" / f"id_agrupados_{cnpj}.parquet",
        ]

        for caminho in arquivos_canonicos:
            if not caminho.exists():
                continue
            try:
                df_origem = self._carregar_dataset_ui(caminho, columns=["id_agrupado", "lista_descricoes"])
            except Exception:
                continue

            if df_origem.is_empty() or not {"id_agrupado", "lista_descricoes"}.issubset(set(df_origem.columns)):
                continue

            df_descricoes = self._montar_descricoes_exibicao_por_grupo(
                df_origem
                .select(
                    [
                        pl.col("id_agrupado").cast(pl.Utf8, strict=False),
                        pl.col("lista_descricoes").cast(pl.List(pl.Utf8), strict=False).alias("descricao_item"),
                    ]
                )
                .explode("descricao_item")
            )
            if not df_descricoes.is_empty():
                return df_descricoes

        return pl.DataFrame()

    def _reconstruir_descricoes_conversao_via_produtos_final(self, cnpj: str) -> pl.DataFrame:
        """Fallback legado para CNPJs ainda nao regenerados."""
        arq_prod_final = CNPJ_ROOT / cnpj / "analises" / "produtos" / f"produtos_final_{cnpj}.parquet"
        if not arq_prod_final.exists():
            return pl.DataFrame()

        try:
            df_prod = self._carregar_dataset_ui(
                arq_prod_final,
                columns=["id_agrupado", "descricao", "descricao_final", "lista_desc_compl"],
            )
        except Exception:
            return pl.DataFrame()

        if df_prod.is_empty() or "id_agrupado" not in df_prod.columns:
            return pl.DataFrame()

        partes_descricoes: list[pl.DataFrame] = []
        if "descricao" in df_prod.columns:
            partes_descricoes.append(
                df_prod.select(
                    [
                        pl.col("id_agrupado").cast(pl.Utf8, strict=False),
                        pl.col("descricao").cast(pl.Utf8, strict=False).alias("descricao_item"),
                    ]
                )
            )
        if "descricao_final" in df_prod.columns:
            partes_descricoes.append(
                df_prod.select(
                    [
                        pl.col("id_agrupado").cast(pl.Utf8, strict=False),
                        pl.col("descricao_final").cast(pl.Utf8, strict=False).alias("descricao_item"),
                    ]
                )
            )
        if "lista_desc_compl" in df_prod.columns:
            partes_descricoes.append(
                df_prod.select(
                    [
                        pl.col("id_agrupado").cast(pl.Utf8, strict=False),
                        pl.col("lista_desc_compl").cast(pl.List(pl.Utf8), strict=False).alias("descricao_item"),
                    ]
                ).explode("descricao_item")
            )

        if not partes_descricoes:
            return pl.DataFrame()

        return self._montar_descricoes_exibicao_por_grupo(
            pl.concat(partes_descricoes, how="vertical_relaxed")
        )

    def _preparar_dataframe_para_salvar_conversao(self, df: pl.DataFrame) -> pl.DataFrame:
        """
        Remove colunas derivadas da UI antes de persistir fatores_conversao.
        """
        return df.drop(
            [
                "__row_id__",
                "preco_medio_ref",
                "fator_calculado",
                "lista_descricoes_produto",
                "descr_padrao_canonico",
            ],
            strict=False,
        )

    def _enriquecer_descricoes_conversao(self, cnpj: str, df: pl.DataFrame) -> pl.DataFrame:
        if df.is_empty() or "id_agrupado" not in df.columns:
            return df

        df_descr_padrao = self._carregar_descr_padrao_canonico_conversao(cnpj)
        df_descricoes_base = self._carregar_descricoes_canonicas_conversao(cnpj)
        if df_descricoes_base.is_empty():
            df_descricoes_base = self._reconstruir_descricoes_conversao_via_produtos_final(cnpj)
        if df_descricoes_base.is_empty() and df_descr_padrao.is_empty():
            return df

        df_out = df.drop(["lista_descricoes_produto", "descr_padrao_canonico"], strict=False)
        if not df_descr_padrao.is_empty():
            df_out = (
                df_out
                .join(df_descr_padrao, on="id_agrupado", how="left")
                .with_columns(
                    pl.coalesce([pl.col("descr_padrao_canonico"), pl.col("descr_padrao")]).alias("descr_padrao")
                )
            )
        if not df_descricoes_base.is_empty():
            df_out = (
                df_out
                .join(df_descricoes_base, on="id_agrupado", how="left")
                .with_columns(
                    pl.col("lista_descricoes_produto")
                    .cast(pl.Utf8, strict=False)
                    .fill_null("")
                    .alias("lista_descricoes_produto")
                )
            )
        else:
            df_out = df_out.with_columns(pl.lit("").alias("lista_descricoes_produto"))

        colunas = list(df_out.columns)
        if "descr_padrao" in colunas and "lista_descricoes_produto" in colunas:
            colunas.remove("lista_descricoes_produto")
            idx_descr = colunas.index("descr_padrao")
            colunas.insert(idx_descr + 1, "lista_descricoes_produto")
            df_out = df_out.select(colunas)
        return df_out

    def atualizar_aba_conversao(self) -> None:
        """Carrega os fatores de conversao do CNPJ atual."""
        cnpj = self.state.current_cnpj
        if not cnpj:
            self._atualizar_titulo_aba_conversao()
            return

        pasta_produtos = CNPJ_ROOT / cnpj / "analises" / "produtos"
        arq_conversao = pasta_produtos / f"fatores_conversao_{cnpj}.parquet"

        if not arq_conversao.exists():
            self.conversion_model.set_dataframe(pl.DataFrame())
            self._conversion_df_full = pl.DataFrame()
            self._conversion_file_path = None
            self._atualizar_titulo_aba_conversao()
            return

        try:
            df = self._carregar_dataset_ui(arq_conversao)
            if "fator_manual" not in df.columns:
                df = df.with_columns(pl.lit(False).alias("fator_manual"))
            if "unid_ref_manual" not in df.columns:
                df = df.with_columns(pl.lit(False).alias("unid_ref_manual"))
            df = self._enriquecer_dataframe_conversao(df)
            df = self._enriquecer_descricoes_conversao(cnpj, df)
            df = df.with_row_index("__row_id__")
            self._conversion_df_full = df
            self._conversion_file_path = arq_conversao
            self._limpar_recalculo_conversao_pendente()
            self._reset_table_resize_flag("conversao")
            id_atual = self.conv_filter_id.currentText()
            ids = self._conversion_df_full.get_column("id_agrupado").cast(pl.Utf8, strict=False).drop_nulls().unique().sort().to_list() if "id_agrupado" in self._conversion_df_full.columns else []
            self._popular_combo_texto(self.conv_filter_id, [str(i) for i in ids], id_atual, "")
            self.aplicar_filtros_conversao()
        except Exception as e:
            self._updating_conversion_model = False
            self._atualizar_titulo_aba_conversao()
            QMessageBox.warning(self, "Erro", f"Erro ao carregar fatores de conversao: {e}")

    def aplicar_filtros_conversao(self) -> None:
        if self._conversion_df_full.is_empty():
            return

        try:
            total_bruto = self._conversion_df_full.height
            df_vis = self._conversion_df_full

            mostrar_unidade_unica = getattr(self, "chk_show_single_unit", None)
            mostrar_unidade_unica = bool(mostrar_unidade_unica and mostrar_unidade_unica.isChecked())
            if (not mostrar_unidade_unica) and {"id_produtos", "unid"}.issubset(set(df_vis.columns)):
                df_multi_unid = (
                    df_vis.group_by("id_produtos")
                    .agg(pl.col("unid").cast(pl.Utf8, strict=False).drop_nulls().n_unique().alias("qtd_unid"))
                    .filter(pl.col("qtd_unid") > 1)
                    .select("id_produtos")
                )
                if df_multi_unid.height > 0:
                    df_vis = df_vis.join(df_multi_unid, on="id_produtos", how="inner")
                else:
                    df_vis = pl.DataFrame(schema=df_vis.schema)

            id_agrupado = self.conv_filter_id.currentText().strip()
            descr = self.conv_filter_desc.text().strip().lower()

            if id_agrupado and "id_agrupado" in df_vis.columns:
                df_vis = df_vis.filter(
                    pl.col("id_agrupado").cast(pl.Utf8, strict=False).fill_null("").str.contains(id_agrupado, literal=True)
                )
            if descr and "descr_padrao" in df_vis.columns:
                df_vis = df_vis.filter(
                    pl.col("descr_padrao").cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(descr, literal=True)
                )

            self._updating_conversion_model = True
            self.conversion_model.set_dataframe(df_vis)
            self._updating_conversion_model = False
            self._resize_table_once(self.conversion_table, "conversao")
            self._aplicar_preferencias_tabela("conversao", self.conversion_table, self.conversion_model)
            for col_oculta in ["__row_id__", "fator_manual", "unid_ref_manual"]:
                if col_oculta in self.conversion_model.dataframe.columns:
                    col_idx = self.conversion_model.dataframe.columns.index(col_oculta)
                    self.conversion_table.setColumnHidden(col_idx, True)
            self._salvar_preferencias_tabela("conversao", self.conversion_table, self.conversion_model)
            self._atualizar_titulo_aba_conversao(df_vis.height, total_bruto)
        except Exception as e:
            self._updating_conversion_model = False
            QMessageBox.warning(self, "Erro", f"Erro ao filtrar fatores de conversao: {e}")

    def _on_conversion_model_changed(self, top_left, bottom_right, _roles) -> None:
        if self._updating_conversion_model:
            return
        if self._conversion_file_path is None or self._conversion_df_full.is_empty():
            return

        df_vis = self.conversion_model.dataframe
        if df_vis.is_empty() or "__row_id__" not in df_vis.columns:
            return

        col_ini = top_left.column()
        col_fim = bottom_right.column()
        touched_cols = set(df_vis.columns[col_ini : col_fim + 1])
        if not ("fator" in touched_cols or "unid_ref" in touched_cols):
            return

        row_ini = max(0, top_left.row())
        row_fim = min(df_vis.height - 1, bottom_right.row())
        
        updates_row_id = []
        updates_fator = []
        updates_unid_ref = []
        updates_fator_manual = []
        updates_unid_ref_manual = []
        
        for r in range(row_ini, row_fim + 1):
            row_id = df_vis.item(r, df_vis.columns.index("__row_id__"))
            
            # Fator
            try:
                fator = df_vis.item(r, df_vis.columns.index("fator"))
                fator_val = None if fator is None else float(fator)
            except Exception:
                fator_val = None
                
            # Unidade de Referencia
            try:
                unid_ref = df_vis.item(r, df_vis.columns.index("unid_ref"))
                unid_ref_val = None if unid_ref is None else str(unid_ref).strip()
            except Exception:
                unid_ref_val = None
                
            updates_row_id.append(int(row_id))
            updates_fator.append(fator_val)
            updates_unid_ref.append(unid_ref_val)
            updates_fator_manual.append("fator" in touched_cols)
            updates_unid_ref_manual.append("unid_ref" in touched_cols)

        if not updates_row_id:
            return

        df_updates = pl.DataFrame({
            "__row_id__": updates_row_id, 
            "fator_editado": updates_fator,
            "unid_ref_editado": updates_unid_ref,
            "fator_manual_editado": updates_fator_manual,
            "unid_ref_manual_editado": updates_unid_ref_manual,
        })
        
        self._conversion_df_full = (
            self._conversion_df_full
            .join(df_updates, on="__row_id__", how="left")
            .with_columns([
                pl.coalesce([pl.col("fator_editado"), pl.col("fator")]).alias("fator"),
                pl.coalesce([pl.col("unid_ref_editado"), pl.col("unid_ref")]).alias("unid_ref"),
                pl.when(pl.col("fator_manual_editado").fill_null(False))
                .then(pl.lit(True))
                .otherwise(pl.col("fator_manual").cast(pl.Boolean, strict=False).fill_null(False))
                .alias("fator_manual"),
                pl.when(pl.col("unid_ref_manual_editado").fill_null(False))
                .then(pl.lit(True))
                .otherwise(pl.col("unid_ref_manual").cast(pl.Boolean, strict=False).fill_null(False))
                .alias("unid_ref_manual"),
            ])
            .drop(["fator_editado", "unid_ref_editado", "fator_manual_editado", "unid_ref_manual_editado"])
        )
        self._conversion_df_full = self._enriquecer_dataframe_conversao(self._conversion_df_full)

        (
            self._preparar_dataframe_para_salvar_conversao(self._conversion_df_full)
            .write_parquet(self._conversion_file_path)
        )
        self.status.showMessage("Fator e/ou unidade de referencia atualizados e salvos.")
        self._marcar_recalculo_conversao_pendente("Clique em 'Recalcular fatores' ou mude de tela.")

    def _atualizar_titulo_aba_conversao(self, visiveis: int | None = None, total: int | None = None) -> None:
        if not hasattr(self, "tabs") or not hasattr(self, "tab_conversao"):
            return
        idx = self.tabs.indexOf(self.tab_conversao)
        if idx < 0:
            return
        if visiveis is None or total is None:
            self.tabs.setTabText(idx, "Conversao")
            return
        self.tabs.setTabText(idx, f"Conversao ({visiveis}/{total})")

    def exportar_conversao_excel(self) -> None:
        """Exporta os fatores de conversao para Excel para edicao."""
        df = self._dataframe_colunas_visiveis(self.conversion_table, self.conversion_model)
        df = df.drop(["__row_id__", "preco_medio_ref", "fator_calculado", "fator_manual", "unid_ref_manual"], strict=False)
        if df.is_empty():
            QMessageBox.information(self, "Aviso", "Nao hA dados para exportar.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Salvar Excel", f"fator_conversao_{self.state.current_cnpj}.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        try:
            df.write_excel(path)
            QMessageBox.information(self, "Sucesso", f"Arquivo salvo com sucesso:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar: {e}")

    def importar_conversao_excel(self) -> None:
        """Importa fatores de conversao do Excel, sobrescrevendo o Parquet."""
        cnpj = self.state.current_cnpj
        if not cnpj:
            return

        path, _ = QFileDialog.getOpenFileName(self, "Abrir Excel", "", "Excel (*.xlsx)")
        if not path:
            return

        try:
            df_excel = pl.read_excel(path)
            # Validacao conforme documentacao: id_produtos, descr_padrao, unid, unid_ref, fator
            mapping = {
                "id_produtos": "id_produtos",
                "descr_padrao": "descr_padrao",
                "unid": "unid",
                "unid_ref": "unid_ref",
                "fator": "fator"
            }
            cols_obrigatorias = list(mapping.keys())
            if not all(c in df_excel.columns for c in cols_obrigatorias):
                raise ValueError(f"O Excel deve conter as colunas: {cols_obrigatorias}")

            pasta_produtos = CNPJ_ROOT / cnpj / "analises" / "produtos"
            nome_saida = f"fatores_conversao_{cnpj}.parquet"
            
            # Renomeia para colunas internas e garante tipos
            df_imp = df_excel.select(cols_obrigatorias).rename({c: mapping[c] for c in cols_obrigatorias})
            df_imp = df_imp.with_columns([
                pl.col("fator").cast(pl.Float64),
                pl.lit(True).alias("fator_manual"),
                pl.lit(True).alias("unid_ref_manual"),
            ])

            df_imp.write_parquet(pasta_produtos / nome_saida)
            self.atualizar_aba_conversao()
            self._marcar_recalculo_conversao_pendente("Clique em 'Recalcular fatores' ou mude de tela.")
            QMessageBox.information(self, "Sucesso", "Fatores de conversao importados com sucesso.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao importar: {e}")

    def reprocessar_agregacao(self) -> None:
        cnpj = self.state.current_cnpj
        if not cnpj:
            return

        ret = QMessageBox.question(
            self,
            "Reprocessar",
            "Isso vai reprocessar a agregacao inteira: padroes, totais, produtos_final, tabelas _agr, precos medios, fatores de conversao, c170_xml, c176_xml, mov_estoque, mensal, anual e Produtos selecionados.\nProsseguir?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if ret == QMessageBox.StandardButton.No:
            return

        def _on_success(ok) -> None:
            self.status.showMessage("Pronto.")
            if ok:
                self.atualizar_tabelas_agregacao()
                self.recarregar_historico_agregacao(cnpj)
                self.atualizar_aba_conversao()
                self.atualizar_aba_mov_estoque()
                self.atualizar_aba_mensal()
                self.atualizar_aba_anual()
                self.atualizar_aba_nfe_entrada()
                self.atualizar_aba_id_agrupados()
                self.atualizar_aba_produtos_selecionados()
                self.refresh_file_tree(cnpj)
                if self.state.current_file is not None:
                    self.load_current_file(reset_columns=False)
                resumo = self.servico_agregacao.resumo_tempos()
                QMessageBox.information(
                    self,
                    "Sucesso",
                    "Reprocessamento concluido com sucesso."
                    + (f"\n\nTempos: {resumo}" if resumo else ""),
                )
            else:
                QMessageBox.warning(self, "Aviso", "Nao foi possivel concluir o reprocessamento.")

        def _on_failure(mensagem: str) -> None:
            self.status.showMessage("Pronto.")
            QMessageBox.critical(self, "Erro", f"Erro ao reprocessar: {mensagem}")

        self._executar_em_worker(
            self.servico_agregacao.reprocessar_agregacao,
            cnpj,
            mensagem_inicial="Reprocessando agregacao, precos medios, fatores e tabelas derivadas...",
            on_success=_on_success,
            on_failure=_on_failure,
        )

    def recalcular_padroes_agregacao(self) -> None:
        """Invoca o servico para recalcular todos os padroes do CNPJ atual."""
        cnpj = self.state.current_cnpj
        if not cnpj:
            return

        self.reprocessar_agregacao()

    def recalcular_totais_agregacao(self) -> None:
        self.reprocessar_agregacao()

    def refazer_tabelas_agr_agregacao(self) -> None:
        self.reprocessar_agregacao()

    def refazer_fontes_produtos_agregacao(self) -> None:
        """Alias legado para refazer_tabelas_agr_agregacao."""
        self.refazer_tabelas_agr_agregacao()

    def recarregar_historico_agregacao(self, cnpj: str) -> None:
        """Carrega a tabela de descricoes agregadas no painel inferior."""
        try:
            path = self.servico_agregacao.carregar_tabela_editavel(cnpj)
            if path.exists():
                df_agregadas = self.parquet_service.load_dataset(path, self._aggregation_results_filters or [])
                df_agregadas = self._aplicar_modo_relacional_agregacao_df(df_agregadas, self._aggregation_results_relational_mode)
            else:
                df_agregadas = pl.DataFrame()
            self.results_table_model.set_dataframe(df_agregadas)
            self._resize_table_once(self.results_table_view, "agregacao_bottom")
            if not self._aplicar_preferencias_tabela("agregacao_bottom", self.results_table, self.results_table_model):
                self._aplicar_perfil_agregacao(
                    "agregacao_bottom",
                    self.results_table,
                    self.results_table_model,
                    self.bottom_profile.currentText(),
                )
        except Exception:
            self.results_table_model.set_dataframe(pl.DataFrame())

    def atualizar_tabelas_agregacao(self) -> None:
        """Atualiza os modelos das tabelas de agregacao."""
        cnpj = self.state.current_cnpj
        if not cnpj: return
        self._aggregation_file_path = self.servico_agregacao.carregar_tabela_editavel(cnpj)
        if self._aggregation_file_path.exists():
            self._load_aggregation_table()
            
    # ==================================================================
    # Consulta SQL - metodos de suporte
    # ==================================================================
    def _populate_sql_combo(self) -> None:
        """Carrega a lista de arquivos SQL disponiveis no combo."""
        self._sql_files = self.sql_service.list_sql_files()
        self.sql_combo.blockSignals(True)
        self.sql_combo.clear()
        self.sql_combo.addItem("- Selecione uma consulta -")
        for info in self._sql_files:
            self.sql_combo.addItem(f"{info.display_name}  [{info.source_dir}]", info.sql_id)
        self.sql_combo.blockSignals(False)

    def _on_sql_selected(self, index: int) -> None:
        """Ao selecionar um SQL no combo: le, exibe e gera o formulario de parametros."""
        if index <= 0:
            self.sql_text_view.setPlainText("")
            self._clear_param_form()
            self._sql_current_sql = ""
            return
        path_str = self.sql_combo.itemData(index)
        if not path_str:
            return
        try:
            sql_text = self.sql_service.read_sql(path_str)
        except Exception as exc:
            self.show_error("Erro ao ler SQL", str(exc))
            return
        self._sql_current_sql = sql_text
        self.sql_text_view.setPlainText(sql_text)
        params = self.sql_service.extract_params(sql_text)
        self._rebuild_param_form(params)

    def _clear_param_form(self) -> None:
        """Remove todos os campos do formulario de parametros."""
        while self.sql_param_form.rowCount() > 0:
            self.sql_param_form.removeRow(0)
        self._sql_param_widgets.clear()

    def _rebuild_param_form(self, params: list[ParamInfo]) -> None:
        """Reconstroi o formulario de parametros conforme os parametros detectados."""
        self._clear_param_form()
        for param in params:
            label = QLabel(f":{param.name}")
            label.setStyleSheet("font-weight: bold; color: #1e40af;")
            if param.widget_type == WIDGET_DATE:
                widget = QDateEdit()
                widget.setCalendarPopup(True)
                widget.setDate(QDate.currentDate())
                widget.setDisplayFormat("dd/MM/yyyy")
            else:
                widget = QLineEdit()
                if param.placeholder:
                    widget.setPlaceholderText(param.placeholder)
                # Pre-preencher CNPJ se disponAvel
                if "cnpj" in param.name.lower() and self.state.current_cnpj:
                    widget.setText(self.state.current_cnpj)
            self.sql_param_form.addRow(label, widget)
            self._sql_param_widgets[param.name] = widget

    def _collect_param_values(self) -> dict[str, str]:
        """Coleta os valores do formulario de parametros."""
        values: dict[str, str] = {}
        for name, widget in self._sql_param_widgets.items():
            if isinstance(widget, QDateEdit):
                values[name] = widget.date().toString("dd/MM/yyyy")
            elif isinstance(widget, QLineEdit):
                values[name] = widget.text().strip()
            else:
                values[name] = ""
        return values

    def _execute_sql_query(self) -> None:
        """Executa a consulta SQL em thread separada."""
        if not self._sql_current_sql:
            self.show_error("Nenhum SQL", "Selecione um arquivo SQL antes de executar.")
            return
        if self.query_worker is not None and self.query_worker.isRunning():
            self.show_error("Aguarde", "Uma consulta ja estA em execucao.")
            return

        values = self._collect_param_values()
        binds = self.sql_service.build_binds(self._sql_current_sql, values)

        self.btn_sql_execute.setEnabled(False)
        self._set_sql_status("a3 Conectando ao Oracle...", "#fef9c3", "#92400e")

        self.query_worker = QueryWorker(self._sql_current_sql, binds)
        self.query_worker.progress.connect(lambda msg: self._set_sql_status(f"a3 {msg}", "#fef9c3", "#92400e"))
        self.query_worker.finished_ok.connect(self._on_query_finished)
        self.query_worker.failed.connect(self._on_query_failed)
        self._registrar_limpeza_worker("query_worker", self.query_worker)
        self.query_worker.start()

    def _on_query_finished(self, df: pl.DataFrame) -> None:
        """Callback quando a consulta Oracle finaliza com sucesso."""
        self.btn_sql_execute.setEnabled(True)
        self._sql_result_df = df
        self._reset_table_resize_flag("sql_result")
        if df.height == 0:
            self._set_sql_status("a1i   Consulta retornou 0 resultados.", "#e0e7ff", "#3730a3")
            self.sql_result_model.set_dataframe(pl.DataFrame())
            self.sql_result_page_label.setText("Total: 0")
        else:
            self._set_sql_status(
                f"a... {df.height:,} linhas, {df.width} colunas.",
                "#dcfce7", "#166534"
            )
            self._mostrar_resultado_sql(df)

    def _on_query_failed(self, message: str) -> None:
        """Callback quando a consulta Oracle falha."""
        self.btn_sql_execute.setEnabled(True)
        self._set_sql_status(f"a Erro: {message[:200]}", "#fee2e2", "#991b1b")

    def _set_sql_status(self, text: str, bg: str, fg: str) -> None:
        self.sql_status_label.setText(text)
        self.sql_status_label.setStyleSheet(
            f"QLabel {{ padding: 4px 8px; background: {bg}; border-radius: 4px; "
            f"border: 1px solid {bg}; color: {fg}; font-weight: bold; }}"
        )

    def _mostrar_resultado_sql(self, dataframe: pl.DataFrame) -> None:
        """Exibe todo o resultado SQL na grade, usando apenas a rolagem da tabela."""

        self.sql_result_model.set_dataframe(dataframe)
        self._resize_table_once(self.sql_result_table, "sql_result")
        self.sql_result_page_label.setText(f"Total: {dataframe.height:,}")

    def _filter_sql_results(self) -> None:
        """Aplica filtro textual global sobre os resultados SQL."""
        search = self.sql_result_search.text().strip().lower()
        if not search or self._sql_result_df.height == 0:
            self._mostrar_resultado_sql(self._sql_result_df)
            return
        # Filtrar em todas as colunas (cast para string)
        exprs = [
            pl.col(c).cast(pl.Utf8, strict=False).fill_null("").str.to_lowercase().str.contains(search, literal=True)
            for c in self._sql_result_df.columns
        ]
        combined = exprs[0]
        for e in exprs[1:]:
            combined = combined | e
        filtered = self._sql_result_df.filter(combined)
        if filtered.height == 0:
            self._set_sql_status(f"Busca '{search}' nao encontrou resultados.", "#e0e7ff", "#3730a3")
            self.sql_result_model.set_dataframe(pl.DataFrame())
            self.sql_result_page_label.setText("Total filtrado: 0")
        else:
            self._set_sql_status(
                f"a... Busca '{search}': {filtered.height:,} de {self._sql_result_df.height:,} linhas.",
                "#dcfce7", "#166534"
            )
            self._mostrar_resultado_sql(filtered)
            self.sql_result_page_label.setText(f"Total filtrado: {filtered.height:,}")

    def _export_sql_results(self) -> None:
        """Exporta os resultados da consulta SQL para Excel."""
        if self._sql_result_df.height == 0:
            self.show_error("Sem dados", "Execute uma consulta antes de exportar.")
            return
        target = self._save_dialog("Exportar resultados SQL para Excel", "Excel (*.xlsx)")
        if not target:
            return
        try:
            sql_name = self.sql_combo.currentText().split("[")[0].strip() or "consulta_sql"
            df = self._dataframe_colunas_visiveis(self.sql_result_table, self.sql_result_model, self._sql_result_df)
            self.export_service.export_excel(target, df, sheet_name=sql_name[:31])
            self.show_info("Exportacao concluida", f"Arquivo gerado em:\n{target}")
        except Exception as exc:
            self.show_error("Falha na exportacao", str(exc))
